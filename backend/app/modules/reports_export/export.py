"""Weekly Activity Report XLSX builder.

One layout: one row per Employee + Date, with dynamic activity column groups
(Project Code, … then 2, 3 …) repeated up to the max activities recorded on any
single day. The preview (flat rows) and this export share the same flat data,
but the export adds per-employee sections so each block is self-contained: a
merged employee title row, the full column header row, that employee's data
rows, then a blank spacer row before the next employee. With a single employee
the title/spacer are omitted and one header sits at the top.

Styling mirrors the company template: Arial 10 bold white header on teal
(FF76A5AF), thin borders, centered count columns, wrapped Day Remarks, real
Excel dates. Employee title rows are bold on a light teal tint (FFD9E2E1).
Sheet: 'Weekly Activity Report'."""
from decimal import Decimal
from io import BytesIO
from itertools import groupby

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.modules.activity_master.benchmark_exception import (
    VALID_BENCHMARK_EXCEPTION_CODES,
    export_exception_remark,
)

SHEET_NAME = "Weekly Activity Report"

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF76A5AF")
_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
_GROUP_FILL = PatternFill(fill_type="solid", fgColor="FFD9E2E1")
_GROUP_FONT = Font(name="Arial", size=10, bold=True)
_DATA_FONT = Font(name="Arial", size=10)
_THIN = Side(style="thin")
_BORDER = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)

# One activity block: (label, width, centered).
_BLOCK = [
    ("Project Code", 16.4, False),
    ("Activity Type", 22.7, False),
    ("Sub Activity Type", 22.0, False),
    ("No. of Tags", 11.0, True),
    ("No. of Docs", 11.0, True),
    ("No. of BOM HEADER", 16.0, True),
    ("No. of Spares", 12.0, True),
    ("No. of Pages", 11.0, True),
    ("No. of Records", 12.0, True),
]
_FIXED_LEFT = [
    ("Employee ID & Name", 24.0, False),
    ("Date", 12.0, False),
    ("Day Status", 11.0, False),
]
_REMARKS = ("Day Remarks", 68.4, False)


def date_range_label(start, end) -> str:
    """Human filename range like "03 JUL - 09 JUL" (zero-padded day, uppercase
    month), used in the download filenames of both XLSX exports."""
    return f"{start.strftime('%d %b').upper()} - {end.strftime('%d %b').upper()}"


def _new_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    return wb, ws


def _write_header(ws, row: int, columns: list[tuple[str, float, bool]]) -> None:
    for idx, (label, width, center) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_group_header(ws, row: int, total_cols: int, label: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    for col in range(1, total_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = _GROUP_FILL
        c.border = _BORDER
    head = ws.cell(row=row, column=1, value=label)
    head.font = _GROUP_FONT
    head.alignment = Alignment(horizontal="left", vertical="center")


def _style_data_cell(cell, center: bool, wrap: bool, is_date: bool) -> None:
    cell.font = _DATA_FONT
    cell.border = _BORDER
    cell.alignment = Alignment(
        horizontal="center" if center else "left", vertical="top", wrap_text=wrap
    )
    if is_date:
        cell.number_format = "yyyy-mm-dd"


def _finalize(wb) -> BytesIO:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


PENDING_SHEET_NAME = "Pending Benchmark"

# --- Benchmark report styling — matched cell-for-cell to the company reference
# workbook (BENCHMARK REPORT 03 JUL - 09 JUL). Only three colours exist in this
# sheet: the yellow header, and the green/red shade on the DIFFERENCE % cell.
# Everything else is white/no fill with black Arial 10 text and thin borders.
_PB_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF00")
# No colour= -> automatic (black), exactly as the reference stores it.
_PB_HEADER_FONT = Font(name="Arial", size=10, bold=True)
_PB_TOTAL_FONT = Font(name="Arial", size=10, bold=True)
# DIFFERENCE % cell shade, keyed off ACHIEVEMENT %. Nothing else is ever shaded.
_DIFF_GREEN = PatternFill(fill_type="solid", fgColor="FFC6EFCE")  # > 100%: ahead
_DIFF_RED = PatternFill(fill_type="solid", fgColor="FFFFC7CE")    # < 95%: needs attention
# ACCEPTED-EXCEPTION amber — an ACTUAL COMPLETED cell whose row (or, on a TOTAL
# row, at least one contributing row) carries a validated benchmark exception:
# all available work was completed, there was simply less of it than the target.
# Deliberately NOT the red "needs attention" shade — this marks an accepted
# outcome, and it lands on the ACTUAL cell alone, never the target, the pending
# or the whole row. The value under it stays the REAL count.
_EXC_AMBER_FILL = PatternFill(fill_type="solid", fgColor="FFFFF2CC")
_EXC_AMBER_FONT = Font(name="Arial", size=10, bold=True, color="FF7F6000")
# Thin amber border in the same family as the fill/font (Excel's darker amber).
_EXC_AMBER_SIDE = Side(style="thin", color="FFBF8F00")
_EXC_AMBER_BORDER = Border(
    top=_EXC_AMBER_SIDE, bottom=_EXC_AMBER_SIDE,
    left=_EXC_AMBER_SIDE, right=_EXC_AMBER_SIDE,
)
# SHORTFALL red — an ACTUAL COMPLETED cell below its target with NO accepted
# exception: work that is genuinely outstanding. The matched pair to the amber
# above, in Excel's own "Bad" palette (the same FFC7CE the DIFFERENCE % cell
# already uses, with its companion dark red), so the two markers read as one
# system: amber = accepted shortfall, red = real shortfall. Lands on the ACTUAL
# cell alone — never the target, the pending, or the whole row — and never
# changes the value under it.
_SHORT_RED_FILL = PatternFill(fill_type="solid", fgColor="FFFFC7CE")
_SHORT_RED_FONT = Font(name="Arial", size=10, bold=True, color="FF9C0006")
_SHORT_RED_SIDE = Side(style="thin", color="FF9C0006")
_SHORT_RED_BORDER = Border(
    top=_SHORT_RED_SIDE, bottom=_SHORT_RED_SIDE,
    left=_SHORT_RED_SIDE, right=_SHORT_RED_SIDE,
)
_PB_HEADER_ROW_HEIGHT = {1: 15.0, 2: 25.5}
_PB_DEFAULT_ROW_HEIGHT = 15.0

# Final column order (29 columns, A..AC): the identity columns — DAY PART
# directly after DATE — with the two percentage columns early (so they read
# beside the activity rather than off past the wide SUB ACTIVITY / PROJECT
# cells), then REMARKS, PROJECT, the three 6-unit groups, then the cycle
# bounds. No ROW TYPE, no per-group TOTAL sub-column, no EMPLOYEE TOTAL.
#
# Widths travel with the SEMANTIC field, not with a column letter: SUB ACTIVITY
# keeps 118.140625 and PROJECT keeps 86.0 wherever they sit.
_PB_LEFT = [
    ("EMP CODE & NAME", 26.0),
    ("DATE", 12.0),
    # Wide enough for its longest value, "HALF DAY (LEGACY)".
    ("DAY PART", 18.0),
    ("ACTIVITY", 22.0),
    ("ACHIEVEMENT %", 18.85546875),
    ("DIFFERENCE %", 15.0),
    ("SUB ACTIVITY", 118.140625),
    ("REMARKS", 50.0),
    ("PROJECT CODE & TITLE", 86.0),
]
_PB_DATE_COL = 2      # DATE
# DAY PART — FULL DAY / FIRST HALF / SECOND HALF / HALF DAY (LEGACY), repeated
# on every detail row of its period (never merged). Blank on a TOTAL row: a
# total spans the cycle, not one period.
_PB_DAY_PART_COL = 3
_PB_ACTIVITY_COL = 4  # ACTIVITY
_PB_ACH_COL = 5       # ACHIEVEMENT % — decides the shade, never wears it
_PB_DIFF_COL = 6      # DIFFERENCE % — the only shaded cell in the body
_PB_SUB_COL = 7       # SUB ACTIVITY
# REMARKS — the remark belonging to the row's OWN period (header remark for
# FULL DAY / HALF DAY (LEGACY), that half's period remark for FIRST/SECOND
# HALF), repeated on every detail row so a filtered row still reads on its
# own. Blank on a TOTAL row: a total spans the cycle, not one specific day.
# Never carries Activity Master's benchmark_remarks.
_PB_REMARKS_COL = 8
_PB_PROJECT_COL = 9   # PROJECT CODE & TITLE — carries the "TOTAL" marker
_PB_GROUPS = ["BENCHMARK TARGET", "ACTUAL COMPLETED", "PENDING BENCHMARK"]
# ledger benchmark_unit values — must stay in the same order as, and cover every
# value of, activity_master.models.COUNT_FIELD_BY_UNIT: a unit missing here has
# no column and its rows would silently land nowhere.
_PB_UNITS = ["tags", "docs", "bom", "spares", "pages", "records"]
_PB_UNIT_LABELS = ["TAGS", "DOCS", "BOM", "SPARES", "PAGES", "RECORDS"]  # no group total
_PB_GROUP_WIDTH = len(_PB_UNIT_LABELS)  # 6 columns per group
# Per-group unit widths: the leading TAGS column is widened to carry the merged
# group label above it; the rest stay 12.
_PB_UNIT_WIDTHS = [
    [21.42578125, 12.0, 12.0, 12.0, 12.0, 12.0],  # BENCHMARK TARGET  J:O
    [16.85546875, 12.0, 12.0, 12.0, 12.0, 12.0],  # ACTUAL COMPLETED  P:U
    [17.7109375, 12.0, 12.0, 12.0, 12.0, 12.0],   # PENDING BENCHMARK V:AA
]
_PB_RIGHT = [("CYCLE START", 13.0), ("CYCLE END", 13.0)]
_PB_NUMFMT_PCT = "0.00%"

# Free-text columns whose values can outrun their width. They wrap (left/top) so
# a long value stays readable instead of being clipped by the next column. This
# is a display property only — no width, order, name or value changes with it.
_PB_WRAP_COLS = (
    1,                    # EMP CODE & NAME
    _PB_ACTIVITY_COL,
    _PB_SUB_COL,
    _PB_REMARKS_COL,
    _PB_PROJECT_COL,
)


def _difference_fill(achievement):
    """Shade for the DIFFERENCE % cell, chosen from the ACHIEVEMENT % fraction
    (1.0 == 100%). Strict boundaries: <95% red, 95%..100% inclusive no shade,
    >100% green. `None` (no numeric target — a textual task row) never shades.

    This fill lands on the DIFFERENCE % cell ONLY. The ACHIEVEMENT % cell that
    decides it, and every other cell on the row, stay unfilled."""
    if achievement is None:
        return None
    if achievement > 1.0:
        return _DIFF_GREEN
    if achievement < 0.95:
        return _DIFF_RED
    return None


def _is_numeric(value) -> bool:
    """Strict numeric-value check: a genuine number, never a status string.

    Only genuinely numeric benchmark values feed the totals and the achievement
    %. The sheet's rows are numeric-only by construction now (task-mode
    activities are excluded upstream), so this is a safety net rather than a
    filter: anything non-numeric that ever reached a unit cell would be excluded
    from the totals instead of being coerced to zero."""
    return isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)


def _upper(value):
    """Uppercase a value on its way into a cell — EXPORT-ONLY.

    Every textual value this workbook writes goes through here, so the sheet
    reads in one case throughout: employee, day part, activity, sub-activity,
    remarks, project, headers, group labels and the TOTAL marker.

    What it deliberately does NOT do:
    - It never touches the stored value. The uppercase string exists only inside
      the workbook; PostgreSQL keeps the employee's own casing verbatim.
    - It passes non-strings straight through untouched, so numbers stay numeric,
      dates stay dates, percentages keep their number format and a formula
      string is never invented. `None` stays `None` (an empty cell), and a
      blank/whitespace-only string collapses to `None` rather than writing "".
    """
    if isinstance(value, str):
        stripped = value.strip()
        return stripped.upper() if stripped else None
    return value


def _bold_system_remark(system: str, user: str | None):
    """The composed REMARKS value as Excel RICH TEXT: the bracketed system
    remark in BOLD, the employee's own remark after it in the normal body font.

    The bold run is what makes an exception row findable by eye in a column of
    free text — "[NO FURTHER TAGS WERE AVAILABLE FOR THIS ACTIVITY] | TNR TAG
    NUMBER" reads as a system marker followed by the employee's words, rather
    than one undifferentiated sentence.

    Only the bold run carries an explicit font; the rest of the cell inherits
    the row's Arial 10 (see style_row), so the two halves match in family and
    size and differ only in weight. Reading the file back without
    `rich_text=True` yields the identical plain string, so every existing
    consumer and assertion is unaffected."""
    marker = TextBlock(InlineFont(b=True), system)
    if user is None:
        return CellRichText(marker)
    return CellRichText(marker, f" | {user}")


def _export_remarks(day_remarks, exception_code, unit):
    """The REMARKS cell for one detail row: the system exception remark first,
    then the employee's own remark, joined by " | ". Uppercase, like every other
    text cell.

    - No exception          -> the employee's remark alone (or an empty cell),
                               as a plain string.
    - Exception, no remark  -> "[NO FURTHER TAGS WERE AVAILABLE FOR THIS ACTIVITY]",
                               bold.
    - Exception + remark    -> "[NO FURTHER ...]" bold, then " | THEIR REMARK"
                               in the normal weight (see _bold_system_remark).

    The remark is composed from the STRUCTURED exception code, never from the
    text of the remark itself: an employee who types the sentence by hand gets
    no bracketed prefix, no bold, and no change to any calculation. Composition
    is a pure function of its inputs and re-checks the prefix, so exporting the
    same cycle twice — or feeding an already-composed value back in — can never
    double the system remark. The stored remark is never modified.
    """
    user = _upper(day_remarks)
    system = (
        export_exception_remark(unit)
        if exception_code in VALID_BENCHMARK_EXCEPTION_CODES
        else None
    )
    if system is None:
        return user
    if user is None:
        return _bold_system_remark(system, None)
    if user.startswith(system):
        # Already composed (a re-export, or a value fed back in): re-split it so
        # the marker is bolded exactly once and never duplicated.
        rest = user[len(system):].removeprefix(" | ") or None
        return _bold_system_remark(system, rest)
    return _bold_system_remark(system, user)


def _cell_number(value):
    """A unit-column value written as a WHOLE number wherever it is one.

    Every quantity in this sheet counts real things — tags, documents, BOM lines
    — so a benchmark cell must read 33, never 33.0 and never 32.5. Targets are
    rounded to whole units upstream (activity_master.service.scaled_target) and
    actuals come from integer columns, so the pending derived from them is whole
    too; this writes them as a genuine Excel integer rather than a float that
    merely displays as one.

    A value that is NOT whole is passed through as a float rather than
    truncated: better a visible oddity than a silently altered number."""
    number = float(value)
    return int(number) if number.is_integer() else number


def _is_short(target, actual) -> bool:
    """True when a genuinely numeric actual falls below a positive target — the
    condition the red shortfall marker keys off.

    A missing or non-numeric value on either side is NOT a shortfall: nothing
    was measured, so nothing is outstanding. A zero target likewise, since
    anything is at or above it."""
    if not (_is_numeric(target) and _is_numeric(actual)):
        return False
    return float(target) > 0 and float(actual) < float(target)


def _row_exception_code(row: dict):
    """The row's validated exception code, or None. An unrecognised value is
    ignored rather than trusted — the workbook is the last consumer of the
    field, not a second validator of it."""
    code = row.get("benchmark_exception_code")
    return code if code in VALID_BENCHMARK_EXCEPTION_CODES else None


def build_pending_benchmark_workbook(
    rows: list[dict], cycle_start, cycle_end
) -> BytesIO:
    """Full-cycle Benchmark XLSX, grouped employee -> sub-activity.

    Layout, styling, colours, fonts, borders, number formats, column widths,
    merged header cells, freeze panes and AutoFilter are matched cell-for-cell
    to the company reference workbook. Within an employee, each sub-activity's
    date-wise detail rows are followed by ONE bold TOTAL row for that exact
    sub-activity (never one combined per-employee total).

    NUMERIC ONLY. `rows` carries per-day numeric benchmark rows and nothing
    else — task-mode activities are excluded upstream (see
    benchmarks.service.get_pending_benchmark_export), so no textual
    "FINISHED" / "N DAYS OVERDUE" cell can land in a numeric column and nothing
    task-shaped contributes to a target, an actual, a pending or a percentage.

    ALL TEXT IS UPPERCASED ON THE WAY IN (see _upper) — headers, group labels,
    employee, day part, activity, sub-activity, remarks, project and the TOTAL
    marker. Numbers stay numeric, dates stay dates, percentages keep their
    number format, and no stored value is modified.

    The TOTAL row repeats the exact EMP CODE & NAME, ACTIVITY and SUB ACTIVITY
    of its detail rows (so an Excel employee filter, or a sub-activity filter,
    keeps both the detail rows and the total), writes "TOTAL" in the PROJECT
    column, and leaves DATE blank. Its PENDING columns net the whole cycle per
    unit (MAX(0, cycle_target - cycle_effective_actual)) rather than summing the
    daily shortages, so a day's overachievement offsets another day's shortfall
    — but only within the same employee + sub-activity + unit. Nothing crosses
    sub-activities, units, employees or cycles.

    BENCHMARK EXCEPTIONS (migration 0063). A detail row may carry a validated
    `benchmark_exception_code`: every available unit of work was completed, but
    fewer were available than the target. Such a row is EVALUATED at its target
    while still REPORTING its real actual:

      effective_actual = target   on a valid exception row
                       = actual   otherwise

    - the ACTUAL COMPLETED cell keeps the real count (40 stays 40) and is the
      only cell that changes appearance — amber fill, amber bold font, thin
      amber border, marking an accepted exception rather than a problem;
    - that row's PENDING cell is 0;
    - the TOTAL row sums REAL actuals into ACTUAL COMPLETED (so the sheet still
      adds up to what was produced) but derives ACHIEVEMENT %, DIFFERENCE % and
      PENDING from the effective contributions, and repeats the amber styling on
      its ACTUAL cell for any unit where at least one contributing row was an
      exception;
    - the REMARKS cell is prefixed with the system exception remark (see
      _export_remarks).

    No column is added, removed, renamed or reordered for any of this: the
    effective actual is never a cell, only an intermediate.

    ACHIEVEMENT % = total_effective_actual / total_target summed across the six
    units, uncapped, formatted 0.00%; blank when total_target is 0 (no divide by
    zero). DIFFERENCE % = ABS(achievement - 100%), formatted 0.00%, blank
    whenever ACHIEVEMENT % is. The DIFFERENCE % cell is the only SHADED cell in
    the body (see _difference_fill) — the amber exception styling is a font +
    border + fill on one ACTUAL cell, never a full row.

    Header is two rows: row 1 carries ONLY the merged group labels, row 2 the
    real per-column header the AutoFilter anchors on. `rows` must arrive sorted
    employee -> activity -> sub-activity -> date -> project (the service
    guarantees it)."""
    wb, ws = _new_sheet()
    ws.title = PENDING_SHEET_NAME

    n_left = len(_PB_LEFT)
    n_units = len(_PB_UNITS)
    first_right = n_left + 1 + len(_PB_GROUPS) * _PB_GROUP_WIDTH  # first CYCLE column
    total_cols = first_right + len(_PB_RIGHT) - 1
    date_cols = {_PB_DATE_COL, first_right, first_right + 1}

    def group_start(gi: int) -> int:
        return n_left + 1 + gi * _PB_GROUP_WIDTH

    # Header labels go through _upper like every other string, so the sheet's
    # case is guaranteed by the writer rather than by how the constants happen
    # to be spelled. The labels themselves are unchanged: same names, same
    # order, same merges — no column is added, dropped, renamed or moved.
    for idx, (label, width) in enumerate(_PB_LEFT, start=1):
        ws.cell(2, idx, _upper(label))
        ws.column_dimensions[get_column_letter(idx)].width = width
    for gi, group in enumerate(_PB_GROUPS):
        start = group_start(gi)
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start + n_units - 1)
        ws.cell(1, start, _upper(group))
        for ui, unit_label in enumerate(_PB_UNIT_LABELS):
            ws.cell(2, start + ui, _upper(unit_label))
            ws.column_dimensions[get_column_letter(start + ui)].width = _PB_UNIT_WIDTHS[gi][ui]
    for ri, (label, width) in enumerate(_PB_RIGHT):
        col = first_right + ri
        ws.cell(2, col, _upper(label))
        ws.column_dimensions[get_column_letter(col)].width = width
    # Yellow header across the FULL A1:AC2 block — the cells left blank above the
    # identity/cycle columns carry the same style as the labelled ones.
    for row in (1, 2):
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = _PB_HEADER_FONT
            cell.fill = _PB_HEADER_FILL
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = _PB_HEADER_ROW_HEIGHT[row]
    ws.freeze_panes = "A3"
    ws.sheet_format.defaultRowHeight = _PB_DEFAULT_ROW_HEIGHT

    def style_row(r: int, bold: bool = False) -> None:
        """Body style: Arial 10, thin borders all round, vertical top, no fill.
        A:I and AB:AC left, the three unit groups (J:AA) centered. Bold on a
        TOTAL row. No fill is applied here — the DIFFERENCE % cell is shaded by
        its writer and is the only shaded body cell."""
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=col)
            _style_data_cell(cell, n_left < col < first_right, False, col in date_cols)
            if bold:
                cell.font = _PB_TOTAL_FONT
        # Long free-text columns wrap instead of spilling across (or being
        # clipped by) their neighbours: a 60-character sub-activity, project
        # title or day remark stays fully readable at the column's own width.
        # Left + top alignment is unchanged; only wrapping is added.
        for col in _PB_WRAP_COLS:
            ws.cell(row=r, column=col).alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

    def style_exception_cell(r: int, col: int) -> None:
        """Mark ONE ACTUAL COMPLETED cell as an accepted exception: amber fill,
        amber bold font, thin amber border. Applied after style_row, so it wins;
        the value in the cell is untouched (it stays the real actual). Never
        applied to the target cell, the pending cell, or any other cell of the
        row."""
        cell = ws.cell(row=r, column=col)
        cell.fill = _EXC_AMBER_FILL
        cell.font = _EXC_AMBER_FONT
        cell.border = _EXC_AMBER_BORDER

    def style_shortfall_cell(r: int, col: int) -> None:
        """Mark ONE ACTUAL COMPLETED cell as a genuine shortfall: red fill, dark
        red bold font, thin red border. Same placement rules as the amber
        marker, and mutually exclusive with it — a row is either an accepted
        exception or a real shortfall, never both."""
        cell = ws.cell(row=r, column=col)
        cell.fill = _SHORT_RED_FILL
        cell.font = _SHORT_RED_FONT
        cell.border = _SHORT_RED_BORDER

    unit_col = {u: i for i, u in enumerate(_PB_UNITS)}

    def write_sub_total_row(
        r: int, *, emp_label: str, activity: str, sub_activity: str,
        totals, used, effective, exception_units,
    ) -> None:
        """One bold TOTAL row for a numeric sub-activity. Nets each unit's cycle
        pending in place (MAX(0, target - EFFECTIVE actual)), then writes the
        per-unit target/actual/pending sums, the uncapped ACHIEVEMENT % and the
        DIFFERENCE % — shading the DIFFERENCE % cell alone. Repeats the exact
        emp/activity/sub-activity, writes "TOTAL" in PROJECT, leaves DATE,
        DAY PART and REMARKS blank (a total spans the cycle, not one specific
        day or period).

        `totals[1]` (ACTUAL COMPLETED) stays the sum of the REAL actuals — the
        column must still add up to what was produced. `effective` is the
        parallel sum where an exception row contributed its target instead, and
        it is what PENDING and both percentages are derived from. It is never
        written to a cell of its own."""
        for ui in range(n_units):
            totals[2][ui] = max(0.0, totals[0][ui] - effective[ui])
        ws.cell(r, 1, _upper(emp_label))                 # exact CODE - NAME (filterable)
        # DATE (col _PB_DATE_COL) stays blank on the total row.
        ws.cell(r, _PB_PROJECT_COL, _upper("TOTAL"))     # PROJECT column marks the total
        ws.cell(r, _PB_ACTIVITY_COL, _upper(activity))   # exact activity name
        ws.cell(r, _PB_SUB_COL, _upper(sub_activity))    # exact sub-activity (filterable)
        for gi in range(len(_PB_GROUPS)):
            for ui in range(n_units):
                if used[ui]:
                    # Whole numbers, same rule as the detail rows: a cycle total
                    # of whole daily figures is itself whole.
                    ws.cell(r, group_start(gi) + ui, _cell_number(totals[gi][ui]))
        ws.cell(r, first_right, cycle_start)
        ws.cell(r, first_right + 1, cycle_end)
        style_row(r, bold=True)
        # Carry the markers up to the total, per unit. Any unit whose cycle
        # total includes at least one exception row gets the amber treatment on
        # its ACTUAL COMPLETED cell, so a reader sees at the total why the
        # percentage reads higher than the raw actual would give. Otherwise a
        # unit whose cycle EVALUATED result still falls short of its target gets
        # the red one — keyed off `effective`, the same figure that drives the
        # PENDING cell and the percentages, so the colour can never disagree
        # with the numbers beside it.
        for ui in range(n_units):
            if not used[ui]:
                continue
            if ui in exception_units:
                style_exception_cell(r, group_start(1) + ui)
            elif _is_short(totals[0][ui], effective[ui]):
                style_shortfall_cell(r, group_start(1) + ui)

        total_target = sum(totals[0])
        total_effective = sum(effective)
        # Uncapped effective/target; blank when target is 0 -> N/A, never a /0.
        # With no exceptions in the group this is identical to the real actual
        # sum, so ordinary totals are bit-for-bit unchanged.
        achievement = (total_effective / total_target) if total_target > 0 else None
        if achievement is None:
            return
        ach_cell = ws.cell(r, _PB_ACH_COL, achievement)
        ach_cell.number_format = _PB_NUMFMT_PCT
        # Distance from target in either direction: 125% and 75% both read 25%.
        # Rounded well past the 2dp the cell shows, purely to keep binary-float
        # noise (0.3999999999999999) out of the formula bar.
        diff_cell = ws.cell(r, _PB_DIFF_COL, round(abs(achievement - 1.0), 10))
        diff_cell.number_format = _PB_NUMFMT_PCT
        # The ONLY shaded cell in the body of the sheet.
        fill = _difference_fill(achievement)
        if fill is not None:
            diff_cell.fill = fill

    r = 3
    for _emp_label, emp_rows in groupby(rows, key=lambda x: x["employee_label"]):
        for _gkey, sub_rows in groupby(emp_rows, key=lambda x: x["group_key"]):
            sub_rows = list(sub_rows)
            totals = [[0.0] * n_units for _ in _PB_GROUPS]
            used = [False] * n_units
            # Calculation-only twin of totals[1] (ACTUAL COMPLETED): an
            # exception row contributes its TARGET here while still contributing
            # its real actual to totals[1]. Drives the total row's PENDING and
            # percentages; never a column of its own.
            effective = [0.0] * n_units
            # Units whose cycle total includes at least one exception row —
            # these get the amber marker on the TOTAL row's ACTUAL cell.
            exception_units: set[int] = set()
            # All rows in a group share one sub_activity_id -> one activity name
            # and one exact sub-activity name; take them from the first row.
            activity = sub_rows[0]["activity"]
            sub_activity = sub_rows[0]["sub_activity"]

            for row in sub_rows:
                exception_code = _row_exception_code(row)
                ws.cell(r, 1, _upper(row["employee_label"]))
                ws.cell(r, _PB_DATE_COL, row["date"])
                # Repeated (never merged) on every row of the period, so a
                # filtered row still names its own part of the day.
                ws.cell(r, _PB_DAY_PART_COL, _upper(row.get("day_part")))
                ws.cell(r, _PB_PROJECT_COL, _upper(row["project"]))
                ws.cell(r, _PB_ACTIVITY_COL, _upper(row["activity"]))
                ws.cell(r, _PB_SUB_COL, _upper(row["sub_activity"]))
                # Repeated on every detail row of this employee+date: the sheet
                # is filterable, so a row must read on its own. An exception row
                # carries the system remark first, the employee's own after it.
                ws.cell(
                    r,
                    _PB_REMARKS_COL,
                    _export_remarks(row.get("day_remarks"), exception_code, row["unit"]),
                )
                # ACHIEVEMENT % / DIFFERENCE % stay blank on every detail row.
                ui = unit_col.get(row["unit"])
                if ui is not None:
                    target_value = row["target"]
                    actual_value = row["actual"]
                    # An accepted exception is evaluated at target, so this row
                    # has nothing outstanding — its PENDING cell reads 0 while
                    # its ACTUAL cell keeps the real count.
                    pending_value = 0.0 if exception_code else row["pending"]
                    for gi, value in enumerate(
                        (target_value, actual_value, pending_value)
                    ):
                        ws.cell(
                            r,
                            group_start(gi) + ui,
                            value if isinstance(value, str) else _cell_number(value),
                        )
                    if exception_code:
                        exception_units.add(ui)
                    # Accumulate cycle target/actual only, and only for genuinely
                    # numeric benchmark values; the pending is derived from them
                    # on the total row so a day's overachievement nets another
                    # day's shortfall per unit.
                    for gi, key in enumerate(("target", "actual")):
                        total_value = row[f"{key}_total"]
                        if _is_numeric(total_value):
                            totals[gi][ui] += float(total_value)
                            used[ui] = True
                    # The parallel CALCULATION-ONLY actual: the target where a
                    # valid exception applies, the real actual everywhere else.
                    # Never written to a cell — it feeds the total row's PENDING
                    # and both percentages, and nothing else.
                    actual_total = row["actual_total"]
                    if _is_numeric(actual_total):
                        effective[ui] += float(
                            row["target_total"]
                            if exception_code and _is_numeric(row["target_total"])
                            else actual_total
                        )
                ws.cell(r, first_right, cycle_start)
                ws.cell(r, first_right + 1, cycle_end)
                style_row(r)
                # After style_row, which would otherwise reset the font and
                # border back to the plain body style. ACTUAL COMPLETED for this
                # row's unit only, and one marker or the other, never both:
                #   amber -> accepted exception (the value is the REAL actual,
                #            never the target);
                #   red   -> a genuine shortfall, nothing excusing it.
                # A row that met or beat its target carries neither.
                if ui is not None:
                    if exception_code:
                        style_exception_cell(r, group_start(1) + ui)
                    elif _is_short(target_value, actual_value):
                        style_shortfall_cell(r, group_start(1) + ui)
                r += 1

            # A numeric sub-activity gets exactly one TOTAL row.
            if any(used):
                write_sub_total_row(
                    r, emp_label=sub_rows[0]["employee_label"], activity=activity,
                    sub_activity=sub_activity, totals=totals, used=used,
                    effective=effective, exception_units=exception_units,
                )
                r += 1

    # Filter on the flattened header row (2) across all data rows.
    ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}{max(r - 1, 2)}"

    return _finalize(wb)


def build_workbook(rows: list[dict], max_activities: int) -> BytesIO:
    wb, ws = _new_sheet()

    # Columns: Employee | Date | Day Status | (block × max) | Day Remarks.
    columns = list(_FIXED_LEFT)
    for i in range(1, max_activities + 1):
        suffix = "" if i == 1 else f" {i}"
        for label, width, center in _BLOCK:
            columns.append((f"{label}{suffix}", width, center))
    columns.append(_REMARKS)
    total_cols = len(columns)
    centers = {idx for idx, (_, _, c) in enumerate(columns, start=1) if c}
    remarks_col = total_cols
    n_left = len(_FIXED_LEFT)

    def write_data_row(r: int, row: dict) -> None:
        ws.cell(r, 1, row["employee_label"])
        ws.cell(r, 2, row["report_date"])
        ws.cell(r, 3, row["day_status"])
        for i, act in enumerate(row["activities"][:max_activities]):
            base = n_left + i * len(_BLOCK)
            ws.cell(r, base + 1, act["project_code"])
            ws.cell(r, base + 2, act["activity_type"])
            ws.cell(r, base + 3, act["sub_activity_type"])
            ws.cell(r, base + 4, act["tags"])
            ws.cell(r, base + 5, act["docs"])
            ws.cell(r, base + 6, act["bom"])
            ws.cell(r, base + 7, act["spares"])
            # Legacy rows without the new counts fall back to 0, matching the
            # NOT NULL DEFAULT 0 the other four count columns carry.
            ws.cell(r, base + 8, act.get("pages") or 0)
            ws.cell(r, base + 9, act.get("records") or 0)
        ws.cell(r, remarks_col, row["remarks"])
        for col in range(1, total_cols + 1):
            _style_data_cell(ws.cell(r, col), col in centers, col == remarks_col, col == 2)

    # rows are ordered by employee_code → contiguous employee sections.
    employees = [(label, list(grp)) for label, grp in groupby(rows, key=lambda r: r["employee_label"])]

    # Single employee (or none): one top header, no title/spacer rows.
    if len(employees) <= 1:
        _write_header(ws, 1, columns)
        ws.freeze_panes = "A2"
        r = 2
        for _, emp_rows in employees:
            for row in emp_rows:
                write_data_row(r, row)
                r += 1
        return _finalize(wb)

    # Multiple employees: a self-contained section per employee.
    r = 1
    for label, emp_rows in employees:
        _write_group_header(ws, r, total_cols, label)
        r += 1
        _write_header(ws, r, columns)
        r += 1
        for row in emp_rows:
            write_data_row(r, row)
            r += 1
        r += 1  # blank spacer row before the next employee

    return _finalize(wb)


# ---------------------------------------------------------------------------
# Project Weekly Report (Phase 7) — the Head's per-project weekly workbook.
#
# Styled after the company's existing hand-made "Head Report" workbook, which is
# the format the Heads already read: a yellow header band (FFFFFF00) in bold
# Calibri 11 over plain body rows, one row per reported activity line, an SL.NO
# running down the left and a wide REMARKS column on the right. The reference
# workbook's own widths are kept where the column survived (SL.NO 6.29, DATE
# 14.14, NAME 26.29, PROJECT 23.14, REMARKS 68.0).
#
# Two columns the hand-made sheet could not have: WORK PERIOD is a real column
# here instead of a "FIRST HALF-" prefix glued onto the activity text, and
# ACTIVITY / SUB-ACTIVITY are separate columns instead of one "MTL / MTL-ASSET
# PHOTO..." string. Both come from real fields (work_report_periods.day_part and
# the two Activity Master snapshots), so neither is parsed out of free text.
#
# The rows handed in are ALREADY the finished report (see
# projects/weekly_report.py). This builder only renders them: it does not sort,
# filter, aggregate or recompute anything, which is exactly why the workbook and
# the browser preview cannot disagree.
_WR_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF00")
_WR_HEADER_FONT = Font(name="Calibri", size=11, bold=True)
_WR_BODY_FONT = Font(name="Calibri", size=11)
_WR_SHEET_NAME = "Weekly Report"
# The blank-cell placeholder, matching the reference workbook's own "-" and the
# browser preview. Never a 0: "no docs were reported" and "0 docs were counted
# against the target" are different statements.
_WR_BLANK = "-"

# (header, width, kind) — kind drives alignment/format only.
#   "text"   left, no wrap        "wrap"   left, wrapped (long free text)
#   "num"    centered number      "date"   real Excel date
_WR_COLUMNS = [
    ("SL.NO", 6.28515625, "num"),
    ("DATE", 14.140625, "date"),
    ("WORK PERIOD", 16.0, "text"),
    ("EMPLOYEE NAME", 26.28515625, "text"),
    ("PROJECT", 23.140625, "text"),
    ("ACTIVITY", 26.0, "wrap"),
    ("SUB-ACTIVITY", 53.42578125, "wrap"),
    ("BENCHMARK", 16.0, "num"),
    ("TAGS", 10.0, "num"),
    ("DOCS", 10.0, "num"),
    ("BOM", 10.0, "num"),
    ("SPARES", 11.0, "num"),
    ("PAGES", 10.0, "num"),
    ("RECORDS", 11.0, "num"),
    ("TASK STATUS", 16.0, "text"),
    ("REMARKS", 68.0, "wrap"),
]
# The six unit columns, in the order COUNT_FIELDS declares them.
_WR_UNITS = ["tags", "docs", "bom", "spares", "pages", "records"]


def _wr_cell_values(row: dict, serial: int) -> list:
    """One report row -> the 16 cell values, in column order.

    Numbers stay numbers (so Excel can sort and filter them), dates stay real
    dates, and anything that does not apply becomes the "-" placeholder rather
    than a misleading zero or an empty cell.
    """
    # Benchmark: the numeric target when there is one, else the textual label
    # ("Lump Sum") for a completion-only task, else blank. Exactly the rule the
    # preview renders, because both read the same two fields.
    benchmark = row.get("benchmark")
    if benchmark is None:
        benchmark = row.get("benchmark_label") or _WR_BLANK
    counts = [
        row.get(unit) if row.get(unit) is not None else _WR_BLANK for unit in _WR_UNITS
    ]
    return [
        serial,
        row["report_date"],
        row.get("work_period_label") or _WR_BLANK,
        row.get("employee_name") or _WR_BLANK,
        # PROJECT = code only. The project name never appears in this column.
        row.get("project_code") or _WR_BLANK,
        row.get("activity_name") or _WR_BLANK,
        row.get("sub_activity_name") or _WR_BLANK,
        benchmark,
        *counts,
        row.get("task_status_label") or _WR_BLANK,
        row.get("remarks") or _WR_BLANK,
    ]


def build_project_weekly_report_workbook(report: dict) -> BytesIO:
    """Render one project's weekly report (the dict from
    projects.weekly_report.get_project_weekly_report) as .xlsx.

    An empty week still produces a valid workbook with its header row — the same
    thing every other CoreOps export does rather than returning an error or an
    empty file. The UI keeps the Download button disabled in that case, so the
    header-only file is a fallback, not the normal path.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _WR_SHEET_NAME

    for idx, (label, width, _kind) in enumerate(_WR_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = _WR_HEADER_FONT
        cell.fill = _WR_HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = width
    # The header stays put while a long week scrolls.
    ws.freeze_panes = "A2"

    for offset, row in enumerate(report.get("rows") or []):
        excel_row = offset + 2
        for idx, ((_label, _width, kind), value) in enumerate(
            zip(_WR_COLUMNS, _wr_cell_values(row, offset + 1)), start=1
        ):
            cell = ws.cell(row=excel_row, column=idx, value=value)
            cell.font = _WR_BODY_FONT
            cell.border = _BORDER
            cell.alignment = Alignment(
                horizontal="center" if kind == "num" else "left",
                vertical="top",
                wrap_text=kind == "wrap",
            )
            if kind == "date":
                cell.number_format = "dd-mm-yyyy"

    return _finalize(wb)
