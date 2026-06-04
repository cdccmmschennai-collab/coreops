"""Seed master data from Excel files.

Reads:
  - Datas/ACTIVITY LIST.xlsx      → activity_types table
  - Datas/PROJ & JOB CODES.xlsx   → job_codes table + projects table

Idempotent: upserts by code. Safe to re-run without creating duplicates or
disturbing project_members / project_managers / reporting_pm_id assignments.

Usage (from backend/ directory inside the container):
  python scripts/seed_master_data.py [--dry-run] [--datas-dir /path]
"""
import argparse
import sys
import uuid
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl

from app.core.database import SessionLocal
from app.modules.activity_types.models import (
    ActivityType,
    CATEGORY_GENERAL,
    CATEGORY_PROJECT,
    CATEGORY_TAG_ESTIMATION,
)
from app.modules.job_codes.models import JobCode
from app.modules.projects.models import Project, ProjectStatus


DATAS_DIR = Path(__file__).parent.parent.parent / "Datas"

_CATEGORY_MAP = {
    "GENERAL": CATEGORY_GENERAL,
    "PROJECT J-CODE": CATEGORY_PROJECT,
    "PROJECT": CATEGORY_PROJECT,
    "TAG ESTIMATION": CATEGORY_TAG_ESTIMATION,
    "TAG_ESTIMATION": CATEGORY_TAG_ESTIMATION,
}


def _col(h_lower: list[str], names: list[str]) -> int | None:
    """Find column index by name. Prefers exact match over substring match."""
    # Pass 1: exact match
    for name in names:
        for i, h in enumerate(h_lower):
            if h == name:
                return i
    # Pass 2: substring match
    for name in names:
        for i, h in enumerate(h_lower):
            if name in h and h:
                return i
    return None


def _load_activity_types(wb_path: Path) -> list[dict]:
    """Parse ACTIVITY LIST.xlsx.

    Excel columns: Activity Code | Activity | Job Code
    """
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header = [str(c).strip() if c else "" for c in rows[0]]
    h_lower = [h.lower() for h in header]

    col_code = _col(h_lower, ["activity code"])
    col_name = _col(h_lower, ["activity"])       # exact "Activity", not "Activity Code"
    col_cat  = _col(h_lower, ["job code", "category"])

    if col_code is None or col_name is None:
        raise ValueError(
            f"Cannot find required columns in {wb_path.name}. "
            f"Headers found: {header}"
        )

    records = []
    seen_codes: set[str] = set()

    for row in rows[1:]:
        if not any(c for c in row):
            continue
        code = str(row[col_code]).strip() if row[col_code] else ""
        name = str(row[col_name]).strip() if col_name is not None and row[col_name] else ""
        cat_raw = (
            str(row[col_cat]).strip().upper()
            if col_cat is not None and row[col_cat]
            else "GENERAL"
        )

        if not code or code.lower() in ("none", "activity code"):
            continue
        if code in seen_codes:
            continue
        seen_codes.add(code)

        category = _CATEGORY_MAP.get(cat_raw, CATEGORY_GENERAL)
        records.append({
            "code": code,
            "name": name or f"Activity {code}",
            "category": category,
            "requires_project": category == CATEGORY_PROJECT,
        })

    wb.close()
    print(f"  Parsed {len(records)} activity types from {wb_path.name}")
    return records


def _load_job_codes_and_projects(wb_path: Path) -> tuple[list[dict], list[dict]]:
    """Parse PROJ & JOB CODES.xlsx.

    Excel columns: S. No | Project Number | Job Code | Project Title | Project Name

    Returns:
      (job_codes_list, projects_list)
      job_codes: deduplicated by J-code; name = first Project Title for that code.
      projects:  one row per Project Number; skips duplicates.
                 Each project carries job_code (str) for FK resolution at upsert time.
    """
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header = [str(c).strip() if c else "" for c in rows[0]]
    h_lower = [h.lower() for h in header]

    col_jcode   = _col(h_lower, ["job code"])
    col_projnum = _col(h_lower, ["project number"])
    col_title   = _col(h_lower, ["project title"])
    col_client  = _col(h_lower, ["project name"])   # contractor name
    col_status  = _col(h_lower, ["status"])

    if col_jcode is None or col_projnum is None:
        raise ValueError(f"Cannot find required columns in {wb_path.name}. Headers: {header}")

    job_codes: dict[str, dict] = {}
    projects: list[dict] = []
    seen_proj_nums: set[str] = set()
    skipped: list[str] = []

    for row in rows[1:]:
        if not any(c for c in row):
            continue

        jcode    = str(row[col_jcode]).strip() if col_jcode is not None and row[col_jcode] else ""
        proj_num = str(row[col_projnum]).strip() if col_projnum is not None and row[col_projnum] else ""
        title    = str(row[col_title]).strip() if col_title is not None and row[col_title] else ""
        client   = str(row[col_client]).strip() if col_client is not None and row[col_client] else ""
        status_raw = (
            str(row[col_status]).strip().lower()
            if col_status is not None and row[col_status]
            else ""
        )

        if not proj_num or proj_num.lower() in ("none", "project number"):
            continue
        if proj_num in seen_proj_nums:
            skipped.append(proj_num)
            continue
        seen_proj_nums.add(proj_num)

        # J-code dedup: first occurrence wins for name
        if jcode and jcode not in job_codes:
            job_codes[jcode] = {"code": jcode, "name": title or jcode}

        proj_status = ProjectStatus.active
        if "hold" in status_raw:
            proj_status = ProjectStatus.on_hold
        elif "complet" in status_raw:
            proj_status = ProjectStatus.completed
        elif "archiv" in status_raw:
            proj_status = ProjectStatus.archived
        elif "plan" in status_raw:
            proj_status = ProjectStatus.planning

        projects.append({
            "code": proj_num,
            "name": title or proj_num,
            "client": client or None,
            "description": None,             # no longer stored in description
            "status": proj_status,
            "job_code": jcode or None,       # raw string for FK resolution
        })

    wb.close()
    if skipped:
        print(
            f"  Skipped {len(skipped)} duplicate project numbers: "
            f"{skipped[:5]}{'...' if len(skipped) > 5 else ''}"
        )
    jc_list = list(job_codes.values())
    print(f"  Parsed {len(jc_list)} unique job codes and {len(projects)} projects from {wb_path.name}")
    return jc_list, projects


# ── upsert helpers ────────────────────────────────────────────────────────────

def _upsert_activity_types(db, records: list[dict], dry_run: bool) -> int:
    from sqlalchemy import select
    upserted = 0
    for rec in records:
        existing = db.execute(
            select(ActivityType).where(ActivityType.code == rec["code"])
        ).scalar_one_or_none()
        if existing is None:
            if not dry_run:
                db.add(ActivityType(
                    code=rec["code"],
                    name=rec["name"],
                    category=rec["category"],
                    requires_project=rec["requires_project"],
                    is_active=True,
                ))
            upserted += 1
        else:
            changed = False
            for field in ("name", "category", "requires_project"):
                if getattr(existing, field) != rec[field]:
                    setattr(existing, field, rec[field])
                    changed = True
            if changed:
                if not dry_run:
                    db.add(existing)
                upserted += 1
    return upserted


def _upsert_job_codes(db, records: list[dict], dry_run: bool) -> tuple[int, dict[str, uuid.UUID]]:
    """Returns (count, code→id mapping for FK resolution)."""
    from sqlalchemy import select
    upserted = 0
    id_map: dict[str, uuid.UUID] = {}

    for rec in records:
        existing = db.execute(
            select(JobCode).where(JobCode.code == rec["code"])
        ).scalar_one_or_none()
        if existing is None:
            jc = JobCode(code=rec["code"], name=rec["name"], is_active=True)
            if not dry_run:
                db.add(jc)
                db.flush()  # need the id for the FK map
                id_map[rec["code"]] = jc.id
            else:
                id_map[rec["code"]] = uuid.uuid4()  # placeholder for dry run
            upserted += 1
        else:
            id_map[rec["code"]] = existing.id
            if existing.name != rec["name"]:
                existing.name = rec["name"]
                if not dry_run:
                    db.add(existing)
                upserted += 1
    return upserted, id_map


def _upsert_projects(
    db,
    records: list[dict],
    job_code_id_map: dict[str, uuid.UUID],
    dry_run: bool,
) -> int:
    from sqlalchemy import select
    upserted = 0
    for rec in records:
        jcode_str = rec.pop("job_code", None)
        jc_id = job_code_id_map.get(jcode_str) if jcode_str else None

        existing = db.execute(
            select(Project).where(
                Project.code == rec["code"],
                Project.deleted_at.is_(None),
            )
        ).scalar_one_or_none()

        if existing is None:
            row = Project(
                code=rec["code"],
                name=rec["name"],
                client=rec["client"],
                description=rec["description"],
                status=rec["status"],
                job_code_id=jc_id,
            )
            if not dry_run:
                db.add(row)
            upserted += 1
        else:
            changed = False
            for field in ("name", "client", "description"):
                if getattr(existing, field) != rec[field]:
                    setattr(existing, field, rec[field])
                    changed = True
            if existing.job_code_id != jc_id and jc_id is not None:
                existing.job_code_id = jc_id
                changed = True
            if changed:
                if not dry_run:
                    db.add(existing)
                upserted += 1

    return upserted


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Seed master data from Excel files.")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--datas-dir",
        default=str(DATAS_DIR),
        help=f"Path to the Datas folder (default: {DATAS_DIR})",
    )
    args = parser.parse_args()

    datas_dir = Path(args.datas_dir)
    activity_path = datas_dir / "ACTIVITY LIST.xlsx"
    proj_path = datas_dir / "PROJ & JOB CODES.xlsx"

    for p in (activity_path, proj_path):
        if not p.exists():
            print(f"ERROR: Not found: {p}", file=sys.stderr)
            sys.exit(1)

    print("Parsing Excel files...")
    activity_records = _load_activity_types(activity_path)
    jc_records, proj_records = _load_job_codes_and_projects(proj_path)

    if args.dry_run:
        print(f"\nDRY RUN — would upsert:")
        print(f"  {len(activity_records)} activity types")
        print(f"  {len(jc_records)} job codes")
        print(f"  {len(proj_records)} projects")
        return

    print("\nWriting to database...")
    with SessionLocal() as db:
        at_count = _upsert_activity_types(db, activity_records, dry_run=False)
        jc_count, jc_id_map = _upsert_job_codes(db, jc_records, dry_run=False)
        db.flush()   # ensure job_code rows have ids before project FK
        pr_count = _upsert_projects(db, proj_records, jc_id_map, dry_run=False)
        db.commit()

    print(f"Done:")
    print(f"  activity_types : {at_count} inserted/updated")
    print(f"  job_codes      : {jc_count} inserted/updated")
    print(f"  projects       : {pr_count} inserted/updated")


if __name__ == "__main__":
    main()
