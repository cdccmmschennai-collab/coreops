"""Import employees + login accounts from the HR spreadsheet.

Source: "EMP DETAILS.xlsx" — the authoritative employee list supplied by HR.
Columns:
    EMP ID | NAME | EMAIL ID | AVAILABLE LEAVE

Each row creates (idempotently):
  - a User (identity) with the provided EMAIL ID as the login email, the shared
    temporary password (cdc@2026), and role=employee.
  - an Employee linked to that user, with EMP ID as employee_code, the cleaned
    display name split into first/last name, and EMAIL ID as work_email.

The AVAILABLE LEAVE column is intentionally ignored: there is no leave_balances
table in the schema yet (see project memory).

Name normalisation (per the HR formatting rules):
  - Replace "." separators with a space  (SANTHAKUMAR.J -> SANTHAKUMAR J).
  - Preserve the source capitalisation; do not abbreviate or drop middle names.
  - Collapse/trim extra whitespace (including non-breaking spaces).
The cleaned value is the display name (full_name). It is split into
first_name (everything before the last token) and last_name (the last token)
so Employee.full_name reproduces the cleaned display name exactly.

Email normalisation:
  - Strip surrounding whitespace, including stray NBSP/U+00A0 characters that
    appear in the sheet.

Force password change on first login:
  - If the users table has a `must_change_password` column it is set TRUE for
    every imported user. If the column does not exist yet the script prints a
    warning and continues (the temporary password still works); add the column
    + login gate separately to fully enforce the rule.

Idempotent: looks up the User by email (case-insensitive) and the Employee by
employee_code. Re-running does NOT create duplicates — existing rows are left
in place (names/links are refreshed, but passwords are NOT reset on re-run so a
user who already changed theirs is not clobbered).

Conflict safety: if an employee_code already exists but is linked to a
DIFFERENT email than the sheet, the row is treated as a conflict and SKIPPED
in full (no new login is created, the existing record is not overwritten). A
matching email is treated as a normal idempotent re-import.

Usage (inside the backend container):
  python -m scripts.import_employees --file /tmp/emp_details.xlsx [--dry-run] \
      [--office Chennai] [--no-office]

--dry-run prints the full transformation/duplicate report and rolls back.
"""
import argparse
import re
import sys
import unicodedata
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook
from sqlalchemy import inspect as sa_inspect, select

from app.core.database import SessionLocal
from app.core.security import hash_password
from app.modules.employees.models import Employee, EmployeeStatus
from app.modules.offices.models import Office
from app.modules.users.models import User, UserRole

TEMP_PASSWORD = "cdc@2026"
DEFAULT_OFFICE = "Chennai"

# Header label -> internal key. Matched case-insensitively after stripping.
COLUMNS = {
    "EMP ID": "code",
    "NAME": "name",
    "EMAIL ID": "email",
}


def _clean_cell(value) -> str:
    """Strip a cell to text, normalising unicode whitespace (incl. NBSP)."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = unicodedata.normalize("NFKC", str(value))
    # NFKC turns NBSP (U+00A0) into a normal space; collapse runs and trim.
    return re.sub(r"\s+", " ", text).strip()


def clean_name(raw: str) -> str:
    """Apply the HR name-formatting rules to a raw NAME cell."""
    text = unicodedata.normalize("NFKC", str(raw or ""))
    text = text.replace(".", " ")          # "." separators -> space
    text = re.sub(r"\s+", " ", text).strip()  # collapse + trim
    return text


def split_name(display: str) -> tuple[str, str]:
    """Split a cleaned display name into (first_name, last_name).

    Last whitespace-delimited token is the last name; everything before it is
    the first name (so full_name == display). Single-token names get an empty
    last name.
    """
    parts = display.split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]


def read_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [_clean_cell(c).upper() for c in rows[0]]
    idx = {}
    for label, key in COLUMNS.items():
        if label not in header:
            raise SystemExit(
                f"Missing required column '{label}' in {path.name}. Found: {header}"
            )
        idx[key] = header.index(label)

    out = []
    for raw in rows[1:]:
        code = _clean_cell(raw[idx["code"]])
        name_raw = raw[idx["name"]]
        email = _clean_cell(raw[idx["email"]])
        if not code and not _clean_cell(name_raw) and not email:
            continue  # blank/spacer row
        display = clean_name(name_raw)
        first, last = split_name(display)
        out.append({
            "code": code,
            "raw_name": "" if name_raw is None else str(name_raw),
            "display_name": display,
            "first_name": first,
            "last_name": last,
            "email": email,
        })
    return out


def find_duplicates(rows: list[dict]) -> tuple[dict, dict]:
    """Return (dup_emails, dup_codes): value -> list of row indexes (1-based)."""
    emails: dict[str, list[int]] = {}
    codes: dict[str, list[int]] = {}
    for i, r in enumerate(rows, start=1):
        if r["email"]:
            emails.setdefault(r["email"].lower(), []).append(i)
        if r["code"]:
            codes.setdefault(r["code"], []).append(i)
    dup_emails = {k: v for k, v in emails.items() if len(v) > 1}
    dup_codes = {k: v for k, v in codes.items() if len(v) > 1}
    return dup_emails, dup_codes


def print_report(rows: list[dict], dup_emails: dict, dup_codes: dict) -> None:
    print(f"\nParsed {len(rows)} employee rows.\n")
    print(f"{'#':>3}  {'EMP ID':<14} {'RAW NAME':<24} {'-> DISPLAY NAME':<24} {'EMAIL':<34}")
    print("-" * 104)
    for i, r in enumerate(rows, start=1):
        raw = r["raw_name"].strip()
        flag = ""
        if not r["email"]:
            flag = "  [NO EMAIL]"
        if not r["code"]:
            flag += "  [NO EMP ID]"
        print(f"{i:>3}  {r['code']:<14} {raw:<24} {r['display_name']:<24} {r['email']:<34}{flag}")

    print("\nName split (first | last):")
    for i, r in enumerate(rows, start=1):
        print(f"{i:>3}  first='{r['first_name']}'  last='{r['last_name']}'")

    print("\nDuplicate emails:")
    if dup_emails:
        for email, idxs in dup_emails.items():
            print(f"  ! {email}  -> rows {idxs}")
    else:
        print("  (none)")

    print("\nDuplicate employee IDs:")
    if dup_codes:
        for code, idxs in dup_codes.items():
            print(f"  ! {code}  -> rows {idxs}")
    else:
        print("  (none)")


def main() -> int:
    parser = argparse.ArgumentParser(description="Import employees + login accounts from the HR spreadsheet.")
    parser.add_argument("--file", required=True, help="Path to EMP DETAILS.xlsx")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print the transformation/duplicate report and roll back.")
    parser.add_argument("--office", default=DEFAULT_OFFICE,
                        help=f"Office name to assign imported employees to (default: {DEFAULT_OFFICE}).")
    parser.add_argument("--no-office", action="store_true",
                        help="Leave office_id null instead of assigning an office.")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        raise SystemExit(f"File not found: {path}")

    rows = read_rows(path)
    dup_emails, dup_codes = find_duplicates(rows)
    print_report(rows, dup_emails, dup_codes)

    if dup_emails:
        raise SystemExit("\nAborting: duplicate emails in the sheet must be resolved first.")

    with SessionLocal() as db:
        # Does users.must_change_password exist? (set only if so.)
        user_cols = {c["name"] for c in sa_inspect(db.bind).get_columns("users")}
        has_force_flag = "must_change_password" in user_cols
        if not has_force_flag:
            print("\nWARNING: users.must_change_password column not found — the "
                  "force-password-change-on-first-login flag will NOT be set. "
                  "Add the column + login gate to enforce that rule.")

        office = None
        if not args.no_office:
            office = db.execute(
                select(Office).where(Office.name == args.office)
            ).scalar_one_or_none()
            if office is None:
                raise SystemExit(
                    f"Office '{args.office}' not found. Use --no-office or pass a valid --office."
                )

        created_users = linked_users = 0
        created_emps = updated_emps = skipped = 0
        warnings: list[str] = []

        for r in rows:
            if not r["email"]:
                warnings.append(f"  [{r['code']}] no email -> skipped (cannot create a login).")
                skipped += 1
                continue
            if not r["code"]:
                warnings.append(f"  [{r['email']}] no EMP ID -> skipped (cannot create an employee).")
                skipped += 1
                continue

            # --- Employee lookup first, so conflicts are caught before any
            #     login is created. ---
            emp = db.execute(
                select(Employee).where(
                    Employee.employee_code == r["code"], Employee.deleted_at.is_(None)
                )
            ).scalar_one_or_none()
            if emp is not None and (emp.work_email or "").lower() != r["email"].lower():
                warnings.append(
                    f"  [{r['code']}] CONFLICT: already exists as "
                    f"'{emp.full_name}' <{emp.work_email or '-'}> "
                    f"-> skipped (sheet email {r['email']} not applied)."
                )
                skipped += 1
                continue

            # --- User (by email, case-insensitive via CITEXT) ---
            user = db.execute(
                select(User).where(User.email == r["email"], User.deleted_at.is_(None))
            ).scalar_one_or_none()
            if user is None:
                user = User(
                    email=r["email"],
                    password_hash=hash_password(TEMP_PASSWORD),
                    role=UserRole.employee,
                    is_active=True,
                )
                if has_force_flag:
                    user.must_change_password = True
                db.add(user)
                db.flush()
                created_users += 1
            else:
                linked_users += 1  # left as-is; password NOT reset on re-run

            # --- Employee (create or idempotent refresh) ---
            if emp is None:
                emp = Employee(
                    employee_code=r["code"],
                    first_name=r["first_name"],
                    last_name=r["last_name"],
                    work_email=r["email"],
                    user_id=user.id,
                    office_id=office.id if office else None,
                    status=EmployeeStatus.active,
                )
                db.add(emp)
                db.flush()
                created_emps += 1
            else:
                emp.first_name = r["first_name"]
                emp.last_name = r["last_name"]
                emp.work_email = r["email"]
                if emp.user_id is None:
                    emp.user_id = user.id
                if office and emp.office_id is None:
                    emp.office_id = office.id
                db.add(emp)
                updated_emps += 1

        if warnings:
            print("\nRow warnings:")
            print("\n".join(warnings))

        office_label = "none" if office is None else office.name
        print(f"\nUsers:     {created_users} created, {linked_users} already existed.")
        print(f"Employees: {created_emps} created, {updated_emps} updated, {skipped} skipped.")
        print(f"Office assigned: {office_label}")
        print(f"Temporary password for new users: {TEMP_PASSWORD}")
        print(f"Force password change set: {'yes' if has_force_flag else 'NO (column missing)'}")

        if args.dry_run:
            db.rollback()
            print("\nDRY RUN — no changes committed.")
            return 0

        db.commit()
        print("\nDone — committed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
