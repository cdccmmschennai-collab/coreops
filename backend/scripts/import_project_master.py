"""Import the real project master from the PM's spreadsheet.

Source: "PROJECT vs PLANNING PLANT.xlsx" — the authoritative list of real
projects, supplied by the PM. Columns:
    PROJECT CODE | PROJECT TITLE | PLANNING PLANT | DESCRIPTION (PP) | JOB CODE

Each project links to:
  - a Planning Plant, matched by code against the seeded planning_plants table
    (the DESCRIPTION (PP) column is informational; the description is sourced
    from planning_plants so the project autofill stays consistent).
  - a Job Code, matched by code against the seeded job_codes table. "NA"
    (or blank) means no job code -> null.

Maintenance Plants are intentionally NOT set here: they hang off the Planning
Plant (maintenance_plants.planning_plant_id) and are selected at usage time
once that master data is loaded. projects.maintenance_plant_id stays null.

Idempotent: upserts by project code (case-sensitive natural key). Safe to
re-run when the spreadsheet is updated. Does NOT delete projects that are
absent from the sheet — it only inserts/updates the rows it finds.

Usage (from backend/ inside the container):
  python scripts/import_project_master.py --file /tmp/project_master.xlsx [--dry-run]
"""
import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook
from sqlalchemy import select

from app.core.database import SessionLocal
from app.modules.job_codes.models import JobCode
from app.modules.plants.models import PlanningPlant
from app.modules.projects.models import Project, ProjectStatus

# Header label -> internal key. Matched case-insensitively after stripping.
COLUMNS = {
    "PROJECT CODE": "code",
    "PROJECT TITLE": "name",
    "PLANNING PLANT": "planning_plant_code",
    "JOB CODE": "job_code",
}

NO_VALUE = {"", "NA", "N/A", "NONE", "-"}


def _clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _read_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [(_clean(c)).upper() for c in rows[0]]
    idx = {}
    for label, key in COLUMNS.items():
        if label not in header:
            raise SystemExit(f"Missing required column '{label}' in {path.name}. Found: {header}")
        idx[key] = header.index(label)

    out = []
    for raw in rows[1:]:
        code = _clean(raw[idx["code"]])
        if not code:
            continue  # skip blank/spacer rows
        name = _clean(raw[idx["name"]])
        pp_code = _clean(raw[idx["planning_plant_code"]])
        job_code = _clean(raw[idx["job_code"]])
        out.append({
            "code": code,
            "name": name,
            "planning_plant_code": pp_code,
            "job_code": None if job_code.upper() in NO_VALUE else job_code,
        })
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description="Import the real project master from the PM spreadsheet.")
    parser.add_argument("--file", required=True, help="Path to PROJECT vs PLANNING PLANT.xlsx")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        raise SystemExit(f"File not found: {path}")

    rows = _read_rows(path)
    print(f"Read {len(rows)} project rows from {path.name}.")

    with SessionLocal() as db:
        planning_by_code = {
            pp.code: pp for pp in db.execute(select(PlanningPlant)).scalars().all()
        }
        jobcode_by_code = {
            jc.code: jc for jc in db.execute(select(JobCode)).scalars().all()
        }

        inserted = updated = 0
        warnings: list[str] = []

        for r in rows:
            pp = planning_by_code.get(r["planning_plant_code"])
            if pp is None:
                warnings.append(
                    f"  [{r['code']}] planning plant code '{r['planning_plant_code']}' "
                    f"not found in planning_plants -> left null"
                )
            jc = None
            if r["job_code"] is not None:
                jc = jobcode_by_code.get(r["job_code"])
                if jc is None:
                    warnings.append(
                        f"  [{r['code']}] job code '{r['job_code']}' not found in job_codes -> left null"
                    )

            existing = db.execute(
                select(Project).where(Project.code == r["code"], Project.deleted_at.is_(None))
            ).scalar_one_or_none()

            if existing is None:
                db.add(Project(
                    code=r["code"],
                    name=r["name"],
                    status=ProjectStatus.active,
                    planning_plant_id=pp.id if pp else None,
                    job_code_id=jc.id if jc else None,
                    maintenance_plant_id=None,
                ))
                inserted += 1
                print(f"  + INSERT {r['code']:<18} PP={r['planning_plant_code'] or '-':<5} "
                      f"JC={r['job_code'] or '-':<6} {r['name'][:60]}")
            else:
                existing.name = r["name"]
                existing.planning_plant_id = pp.id if pp else None
                existing.job_code_id = jc.id if jc else None
                db.add(existing)
                updated += 1
                print(f"  ~ UPDATE {r['code']:<18} PP={r['planning_plant_code'] or '-':<5} "
                      f"JC={r['job_code'] or '-':<6} {r['name'][:60]}")

        if warnings:
            print("\nWarnings:")
            print("\n".join(warnings))

        if args.dry_run:
            db.rollback()
            print(f"\nDRY RUN — would insert {inserted}, update {updated}. Rolled back.")
            return

        db.commit()
        print(f"\nDone: inserted {inserted}, updated {updated}.")


if __name__ == "__main__":
    main()
