"""Pending Benchmark export — per-period row model (DAY PART / REMARKS).

Phase-2 tests for the split-day export row source: every detail row carries a
day_part label (FULL DAY / FIRST HALF / SECOND HALF / HALF DAY (LEGACY)) and
the remarks of ITS OWN period (header remarks for Full-Day and legacy
half-day rows, per-period remarks for halves). Targets come from the frozen
submit-time benchmark_value_snapshot — never re-halved, never re-read from a
later Activity Master edit — and actuals stay exactly as entered.
"""
import uuid as uuid_mod
from datetime import date
from decimal import Decimal

import pytest
from sqlalchemy import select

from app.core.config import settings
from app.modules.activity_master.models import ActivityMaster
from app.modules.activity_master.service import compute_week_bounds
from app.modules.benchmarks.service import get_pending_benchmark_export
from app.modules.projects.models import ProjectStatus
from app.modules.users.models import UserRole
from app.modules.work_reports.models import WorkReportTask

BASE = "/api/v1/work-reports"
TODAY = date.today()

FULL = "FULL DAY"
FIRST = "FIRST HALF"
SECOND = "SECOND HALF"
LEGACY = "HALF DAY (LEGACY)"


@pytest.fixture()
def day_parts_on():
    prev = settings.REPORT_DAY_PARTS_ENABLED
    settings.REPORT_DAY_PARTS_ENABLED = True
    try:
        yield
    finally:
        settings.REPORT_DAY_PARTS_ENABLED = prev


@pytest.fixture()
def setup_author(make_user, make_employee, make_project, make_project_member, login):
    def _make(*, email="emp@x.com", code="E-1", proj_code="P-1"):
        u = make_user(email, role=UserRole.employee)
        e = make_employee(employee_code=code, user_id=u.id)
        p = make_project(code=proj_code, status=ProjectStatus.active)
        make_project_member(project_id=p.id, employee_id=e.id)
        return {"user": u, "emp": e, "project": p, "header": login(email)}

    return _make


@pytest.fixture()
def pm_header(auth_header):
    return auth_header(email="pm@x.com", role=UserRole.project_manager)


def _make_numeric_sub(client, pm_header, *, value, name, count_field="tags"):
    a = client.post(
        "/api/v1/activity-master/activities",
        json={"name": f"Activity for {name}"}, headers=pm_header,
    ).json()
    sub = client.post(
        f"/api/v1/activity-master/activities/{a['id']}/sub-activities",
        json={
            "name": name, "benchmark_type": "NUMERIC",
            "benchmark_value": value, "relevant_count_field": count_field,
        },
        headers=pm_header,
    ).json()
    return a, sub


def _make_task_sub(client, pm_header, *, name, period=1):
    a = client.post(
        "/api/v1/activity-master/activities",
        json={"name": f"Activity for {name}"}, headers=pm_header,
    ).json()
    sub = client.post(
        f"/api/v1/activity-master/activities/{a['id']}/sub-activities",
        json={"name": name, "benchmark_type": "TASK_BASED"},
        headers=pm_header,
    ).json()
    client.patch(
        f"/api/v1/activity-master/sub-activities/{sub['id']}",
        json={"benchmark_period_days": period}, headers=pm_header,
    )
    return a, sub


def _task(project_id, sub_id=None, **counts):
    body = {"project_id": str(project_id), "description": "work"}
    if sub_id:
        body["sub_activity_id"] = sub_id
    body.update(counts)
    return body


def _create_submit(client, header, payload):
    created = client.post(BASE, headers=header, json=payload)
    assert created.status_code == 201, created.text
    res = client.post(f"{BASE}/{created.json()['id']}/submit", headers=header)
    assert res.status_code == 200, res.text
    return created.json()


def _submit_split(client, header, first, second, report_date=TODAY):
    return _create_submit(client, header, {
        "report_date": report_date.isoformat(),
        "report_mode": "split_day",
        "periods": [
            {"day_part": "first_half", **first},
            {"day_part": "second_half", **second},
        ],
    })


def _rows(db):
    """Detail rows of the export for the CURRENT cycle (tests submit today)."""
    return get_pending_benchmark_export(db, cycle=0, today=TODAY)["rows"]


def _sub_rows(db, sub_name):
    return [r for r in _rows(db) if r["sub_activity"] == sub_name]


WORK = {"period_status": "work_at_office", "location": "chennai"}


# --- Full-Day rows -----------------------------------------------------------

def test_full_day_numeric_row_full_day_part_and_header_remarks(
    client, db, setup_author, pm_header
):
    a = setup_author()
    _, sub = _make_numeric_sub(client, pm_header, value=100, name="FD-TAGS")
    _create_submit(client, a["header"], {
        "report_date": TODAY.isoformat(),
        "day_status": "work_at_office", "location": "chennai",
        "remarks": "full day remark",
        "tasks": [_task(a["project"].id, sub["id"], tags_count=80)],
    })

    rows = _sub_rows(db, "FD-TAGS")
    assert len(rows) == 1
    r = rows[0]
    assert r["day_part"] == FULL
    assert r["day_remarks"] == "full day remark"
    assert float(r["target"]) == 100.0
    assert float(r["actual"]) == 80.0
    assert float(r["pending"]) == 20.0
    assert r["unit"] == "tags"


def test_full_day_target_uses_frozen_snapshot_not_live_master(
    client, db, setup_author, pm_header
):
    """An Activity Master edit AFTER submission must not rewrite history."""
    a = setup_author()
    _, sub = _make_numeric_sub(client, pm_header, value=100, name="FROZEN")
    _create_submit(client, a["header"], {
        "report_date": TODAY.isoformat(),
        "tasks": [_task(a["project"].id, sub["id"], tags_count=80)],
    })
    master = db.get(ActivityMaster, uuid_mod.UUID(sub["id"]))
    master.benchmark_value = Decimal("999")
    db.commit()

    (r,) = _sub_rows(db, "FROZEN")
    assert float(r["target"]) == 100.0
    assert float(r["target_total"]) == 100.0


# --- one working half only ---------------------------------------------------

def test_first_half_only_row_with_first_half_remarks(
    client, db, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    _, sub = _make_numeric_sub(client, pm_header, value=120, name="AM-ONLY")
    _submit_split(
        client, a["header"],
        {**WORK, "remarks": "morning remark",
         "tasks": [_task(a["project"].id, sub["id"], tags_count=50)]},
        {"period_status": "leave", "remarks": "afternoon leave"},
    )

    rows = _sub_rows(db, "AM-ONLY")
    assert len(rows) == 1                     # no empty SECOND HALF row
    r = rows[0]
    assert r["day_part"] == FIRST
    assert r["day_remarks"] == "morning remark"
    assert float(r["target"]) == 60.0         # effective snapshot, not 30
    assert float(r["actual"]) == 50.0         # exactly as entered, never halved
    assert float(r["pending"]) == 10.0


def test_second_half_only_row_with_second_half_remarks(
    client, db, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    _, sub = _make_numeric_sub(client, pm_header, value=120, name="PM-ONLY")
    _submit_split(
        client, a["header"],
        {"period_status": "leave", "remarks": "morning leave"},
        {**WORK, "remarks": "afternoon remark",
         "tasks": [_task(a["project"].id, sub["id"], tags_count=40)]},
    )

    rows = _sub_rows(db, "PM-ONLY")
    assert len(rows) == 1                     # no empty FIRST HALF row
    r = rows[0]
    assert r["day_part"] == SECOND
    assert r["day_remarks"] == "afternoon remark"
    assert float(r["target"]) == 60.0
    assert float(r["actual"]) == 40.0


# --- both halves -------------------------------------------------------------

def test_both_halves_different_activities_map_their_own_remarks(
    client, db, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    _, s1 = _make_numeric_sub(client, pm_header, value=120, name="AM-ACT")
    _, s2 = _make_numeric_sub(client, pm_header, value=80, name="PM-ACT", count_field="docs")
    _submit_split(
        client, a["header"],
        {**WORK, "remarks": "AM remark",
         "tasks": [_task(a["project"].id, s1["id"], tags_count=60)]},
        {**WORK, "remarks": "PM remark",
         "tasks": [_task(a["project"].id, s2["id"], docs_count=40)]},
    )

    (am,) = _sub_rows(db, "AM-ACT")
    (pm,) = _sub_rows(db, "PM-ACT")
    assert (am["day_part"], am["day_remarks"]) == (FIRST, "AM remark")
    assert (pm["day_part"], pm["day_remarks"]) == (SECOND, "PM remark")
    # No leakage in either direction, and never concatenated.
    assert "PM remark" not in am["day_remarks"]
    assert "AM remark" not in pm["day_remarks"]
    assert float(am["target"]) == 60.0
    assert float(pm["target"]) == 40.0


def test_same_activity_both_halves_two_rows_summing_to_one_base(
    client, db, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    _, sub = _make_numeric_sub(client, pm_header, value=120, name="SAME-BOTH")
    _submit_split(
        client, a["header"],
        {**WORK, "remarks": "AM", "tasks": [_task(a["project"].id, sub["id"], tags_count=40)]},
        {**WORK, "remarks": "PM", "tasks": [_task(a["project"].id, sub["id"], tags_count=50)]},
    )

    rows = _sub_rows(db, "SAME-BOTH")
    assert len(rows) == 2                     # never merged, never a third row
    first, second = rows
    # FIRST HALF always sorts before SECOND HALF for one date.
    assert first["day_part"] == FIRST and second["day_part"] == SECOND
    assert first["day_remarks"] == "AM" and second["day_remarks"] == "PM"
    # Two effective halves, NOT two full-day targets and NOT one capped row.
    assert [float(r["target"]) for r in rows] == [60.0, 60.0]
    assert [float(r["actual"]) for r in rows] == [40.0, 50.0]
    assert sum(float(r["target_total"]) for r in rows) == 120.0


def test_multiple_activities_in_first_half_repeat_part_and_remarks(
    client, db, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    _, s1 = _make_numeric_sub(client, pm_header, value=120, name="BOM-IDB", count_field="bom")
    _, s2 = _make_numeric_sub(client, pm_header, value=60, name="FMTL-X")
    _submit_split(
        client, a["header"],
        {**WORK, "remarks": "shared AM remark",
         "tasks": [
             _task(a["project"].id, s1["id"], bom_count=30),
             _task(a["project"].id, s2["id"], tags_count=20),
         ]},
        {"period_status": "leave"},
    )

    for name in ("BOM-IDB", "FMTL-X"):
        (r,) = _sub_rows(db, name)
        assert r["day_part"] == FIRST, name
        assert r["day_remarks"] == "shared AM remark", name


def test_multiple_activities_in_second_half_repeat_part_and_remarks(
    client, db, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    _, s1 = _make_numeric_sub(client, pm_header, value=100, name="PM-A", count_field="pages")
    _, s2 = _make_numeric_sub(client, pm_header, value=40, name="PM-B", count_field="records")
    _submit_split(
        client, a["header"],
        {"period_status": "comp_off"},
        {**WORK, "remarks": "shared PM remark",
         "tasks": [
             _task(a["project"].id, s1["id"], pages_count=45),
             _task(a["project"].id, s2["id"], records_count=40),
         ]},
    )

    for name in ("PM-A", "PM-B"):
        (r,) = _sub_rows(db, name)
        assert r["day_part"] == SECOND, name
        assert r["day_remarks"] == "shared PM remark", name


def test_two_projects_in_one_half_stay_one_row(
    client, db, setup_author, pm_header, day_parts_on,
    make_project, make_project_member,
):
    """Two task rows of the same sub-activity in ONE period (different
    projects) collapse into one detail row — actual summed, target once."""
    a = setup_author()
    p2 = make_project(code="P-2", status=ProjectStatus.active)
    make_project_member(project_id=p2.id, employee_id=a["emp"].id)
    _, sub = _make_numeric_sub(client, pm_header, value=120, name="TWO-PROJ")
    _submit_split(
        client, a["header"],
        {**WORK, "remarks": "AM",
         "tasks": [
             _task(a["project"].id, sub["id"], tags_count=30),
             _task(str(p2.id), sub["id"], tags_count=20),
         ]},
        {"period_status": "leave"},
    )

    rows = _sub_rows(db, "TWO-PROJ")
    assert len(rows) == 1                     # no duplicate rows from the join
    r = rows[0]
    assert float(r["target"]) == 60.0         # target counted ONCE per period
    assert float(r["actual"]) == 50.0
    assert "P-1" in r["project"] and "P-2" in r["project"]


# --- non-working day ---------------------------------------------------------

def test_leave_day_produces_no_benchmark_rows(client, db, setup_author, pm_header):
    a = setup_author()
    _create_submit(client, a["header"], {
        "report_date": TODAY.isoformat(), "day_status": "leave",
        "remarks": "on leave", "tasks": [],
    })
    assert _rows(db) == []


# --- historical half-day -----------------------------------------------------

def test_legacy_half_day_label_and_preserved_effective_target(
    client, db, setup_author, pm_header
):
    a = setup_author()
    _, sub = _make_numeric_sub(client, pm_header, value=120, name="OLD-HALF")
    _create_submit(client, a["header"], {
        "report_date": TODAY.isoformat(), "day_status": "half_day",
        "location": "chennai", "remarks": "historic remark",
        "tasks": [_task(a["project"].id, sub["id"], tags_count=50)],
    })

    (r,) = _sub_rows(db, "OLD-HALF")
    assert r["day_part"] == LEGACY            # never guessed to a half
    assert float(r["target"]) == 60.0         # existing effective target intact
    assert float(r["actual"]) == 50.0
    assert r["day_remarks"] == "historic remark"


def test_legacy_periodless_task_falls_back_by_day_status(
    client, db, setup_author, pm_header
):
    """A deploy-gap row (period_id NULL) is labelled from the report header:
    half_day -> HALF DAY (LEGACY), anything else -> FULL DAY."""
    a = setup_author()
    _, sub = _make_numeric_sub(client, pm_header, value=120, name="GAP-ROW")
    created = _create_submit(client, a["header"], {
        "report_date": TODAY.isoformat(), "day_status": "half_day",
        "location": "chennai",
        "tasks": [_task(a["project"].id, sub["id"], tags_count=50)],
    })
    task = db.execute(
        select(WorkReportTask).where(WorkReportTask.report_id == uuid_mod.UUID(created["id"]))
    ).scalar_one()
    task.period_id = None
    db.commit()

    (r,) = _sub_rows(db, "GAP-ROW")
    assert r["day_part"] == LEGACY
    assert float(r["target"]) == 60.0         # snapshot still authoritative


def test_missing_snapshot_falls_back_to_live_base_times_fraction(
    client, db, setup_author, pm_header
):
    """Rows with no frozen snapshot (pre-snapshot history) fall back to the
    live master value scaled by the period fraction — the pre-change rule."""
    a = setup_author()
    _, sub = _make_numeric_sub(client, pm_header, value=100, name="NO-SNAP")
    created = _create_submit(client, a["header"], {
        "report_date": TODAY.isoformat(),
        "tasks": [_task(a["project"].id, sub["id"], tags_count=70)],
    })
    task = db.execute(
        select(WorkReportTask).where(WorkReportTask.report_id == uuid_mod.UUID(created["id"]))
    ).scalar_one()
    task.benchmark_value_snapshot = None
    task.relevant_count_field_snapshot = None
    db.commit()

    (r,) = _sub_rows(db, "NO-SNAP")
    assert r["day_part"] == FULL
    assert float(r["target"]) == 100.0
    assert float(r["actual"]) == 70.0
    assert r["unit"] == "tags"


# --- TASK_BASED rows ---------------------------------------------------------

def test_task_based_full_day_row_keeps_text_cells_and_gains_day_part(
    client, db, setup_author, pm_header
):
    a = setup_author()
    _, sub = _make_task_sub(client, pm_header, name="LUMPSUM", period=1)
    _create_submit(client, a["header"], {
        "report_date": TODAY.isoformat(), "remarks": "task day remark",
        "tasks": [_task(a["project"].id, sub["id"])],
    })

    (r,) = _sub_rows(db, "LUMPSUM")
    assert r["day_part"] == FULL
    assert r["day_remarks"] == "task day remark"
    # Existing TASK_BASED text behaviour, byte-for-byte.
    assert r["target"] == "FINISH WITHIN A DAY"
    assert r["actual"] == "NOT COMPLETED"
    # Overdue is judged from the cycle end (existing export behaviour).
    days = (compute_week_bounds(TODAY)[1] - TODAY).days
    expected = "PENDING" if days <= 0 else f"{days} DAY{'S' if days != 1 else ''} OVERDUE"
    assert r["pending"] == expected
    assert r["target_total"] is None


def test_task_based_in_a_half_carries_period_part_and_remarks(
    client, db, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    _, sub = _make_task_sub(client, pm_header, name="HALF-TASK", period=1)
    _submit_split(
        client, a["header"],
        {**WORK, "remarks": "AM task remark", "tasks": [_task(a["project"].id, sub["id"])]},
        {"period_status": "leave"},
    )

    (r,) = _sub_rows(db, "HALF-TASK")
    assert r["day_part"] == FIRST
    assert r["day_remarks"] == "AM task remark"
    # Deadline semantics unscaled: same 1-day text as a full-day task.
    assert r["target"] == "FINISH WITHIN A DAY"


# --- ordering ----------------------------------------------------------------

def test_rows_order_first_half_before_second_half_deterministically(
    client, db, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    _, sub = _make_numeric_sub(client, pm_header, value=120, name="ORDERED")
    # Second half listed FIRST in the payload — DB insertion order must not win.
    _create_submit(client, a["header"], {
        "report_date": TODAY.isoformat(),
        "report_mode": "split_day",
        "periods": [
            {"day_part": "second_half", **WORK, "remarks": "PM",
             "tasks": [_task(a["project"].id, sub["id"], tags_count=50)]},
            {"day_part": "first_half", **WORK, "remarks": "AM",
             "tasks": [_task(a["project"].id, sub["id"], tags_count=40)]},
        ],
    })

    rows = _sub_rows(db, "ORDERED")
    assert [r["day_part"] for r in rows] == [FIRST, SECOND]
    assert [r["day_remarks"] for r in rows] == ["AM", "PM"]


# --- workbook: DAY PART column C / REMARKS column H --------------------------

EXPORT_URL = "/api/v1/benchmarks/pending-export.xlsx?week_offset=0"


def _sheet(client, pm_header):
    import openpyxl
    from io import BytesIO

    res = client.get(EXPORT_URL, headers=pm_header)
    assert res.status_code == 200, res.text
    # Reopening through openpyxl IS the validity check for the written file.
    return openpyxl.load_workbook(BytesIO(res.content)).active


DATE_C, DAY_PART_C, REMARKS_C, PROJECT_C = 2, 3, 8, 9


def test_workbook_split_day_repeats_date_and_never_merges_details(
    client, db, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    _, sub = _make_numeric_sub(client, pm_header, value=120, name="WB-BOTH")
    _submit_split(
        client, a["header"],
        {**WORK, "remarks": "AM", "tasks": [_task(a["project"].id, sub["id"], tags_count=40)]},
        {**WORK, "remarks": "PM", "tasks": [_task(a["project"].id, sub["id"], tags_count=50)]},
    )

    ws = _sheet(client, pm_header)
    details = [
        r for r in range(3, ws.max_row + 1)
        if ws.cell(r, DAY_PART_C).value in (FIRST, SECOND)
    ]
    assert len(details) == 2
    first_r, second_r = details
    assert ws.cell(first_r, DAY_PART_C).value == FIRST
    assert ws.cell(second_r, DAY_PART_C).value == SECOND
    # Employee and DATE repeat on both period rows — no vertical merging.
    for r in details:
        assert ws.cell(r, 1).value is not None
        assert ws.cell(r, DATE_C).value is not None
    assert ws.cell(first_r, REMARKS_C).value == "AM"
    assert ws.cell(second_r, REMARKS_C).value == "PM"
    # The only merged ranges are the three group headers — nothing overlaps.
    assert {str(m) for m in ws.merged_cells.ranges} == {"J1:O1", "P1:U1", "V1:AA1"}

    # TOTAL row: blank DATE, DAY PART and REMARKS; aggregates both halves.
    totals = [r for r in range(3, ws.max_row + 1) if ws.cell(r, PROJECT_C).value == "TOTAL"]
    assert len(totals) == 1
    t = totals[0]
    assert ws.cell(t, DATE_C).value is None
    assert ws.cell(t, DAY_PART_C).value is None
    assert ws.cell(t, REMARKS_C).value is None
    assert ws.cell(t, 10).value == 120.0      # J: summed effective targets
    assert ws.cell(t, 16).value == 90.0       # P: summed actuals
    assert ws.cell(t, 22).value == 30.0       # V: netted pending


def test_workbook_legacy_half_day_label_and_no_day_remarks_header(
    client, db, setup_author, pm_header
):
    a = setup_author()
    _, sub = _make_numeric_sub(client, pm_header, value=120, name="WB-LEGACY")
    _create_submit(client, a["header"], {
        "report_date": TODAY.isoformat(), "day_status": "half_day",
        "location": "chennai", "remarks": "old remark",
        "tasks": [_task(a["project"].id, sub["id"], tags_count=50)],
    })

    ws = _sheet(client, pm_header)
    assert ws.cell(2, DAY_PART_C).value == "DAY PART"
    assert ws.cell(2, REMARKS_C).value == "REMARKS"
    # DAY REMARKS exists nowhere in the workbook any more.
    everywhere = {
        ws.cell(r, c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    }
    assert "DAY REMARKS" not in everywhere
    detail = next(
        r for r in range(3, ws.max_row + 1) if ws.cell(r, DATE_C).value is not None
    )
    assert ws.cell(detail, DAY_PART_C).value == LEGACY
    assert ws.cell(detail, REMARKS_C).value == "old remark"
    assert ws.cell(detail, 10).value == 60.0  # preserved effective half target


def test_workbook_single_half_produces_no_empty_counterpart_row(
    client, db, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    _, sub = _make_numeric_sub(client, pm_header, value=120, name="WB-ONEHALF")
    _submit_split(
        client, a["header"],
        {"period_status": "leave", "remarks": "am leave"},
        {**WORK, "remarks": "pm work",
         "tasks": [_task(a["project"].id, sub["id"], tags_count=50)]},
    )

    ws = _sheet(client, pm_header)
    parts = [
        ws.cell(r, DAY_PART_C).value
        for r in range(3, ws.max_row + 1)
        if ws.cell(r, DATE_C).value is not None
    ]
    assert parts == [SECOND]                  # one row; no empty FIRST HALF row
