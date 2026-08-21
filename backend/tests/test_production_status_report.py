"""Production Status - Phase 4 (PM cumulative report + Excel export).

Focused suite for the report only. Nothing here touches the Phase 1-3 recording
workflow; those rules are covered in test_production_status.py and are unchanged.

Covers, in order:

  1. PM can read the report               test_project_manager_can_read_report
  2. Head / Lead / employee get 403       test_non_pm_roles_are_rejected
  3. same 403 over HTTP                   test_report_endpoints_are_pm_only_over_http
  4. every project appears                test_report_spans_every_project
  5. every activity in a project appears  test_report_includes_every_activity
  6. REV-0 and REV-1 both survive         test_both_revisions_appear
  7. only the latest row per combination  test_only_latest_row_per_combination
  8. counts stay four independent values  test_counts_are_four_independent_values
  9. BY is the person, not the role       test_by_is_the_actual_person_name
 10. null completed_on stays null         test_null_completed_on_stays_null
 11. remarks are never truncated          test_full_remarks_are_preserved
 12. zero counts stay numeric zero        test_zero_counts_stay_numeric_zero
 13. PROJECT / PLANT convention           test_project_plant_uses_project_information
 14. empty report                         test_empty_report
 15. workbook == the preview dataset      test_workbook_matches_the_report_dataset
 16. workbook columns + formatting        test_workbook_columns_and_formatting
 17. the download is a valid .xlsx        test_download_produces_a_valid_xlsx
 18. the report reads no history          test_report_never_returns_history_rows

Phase 5 adds ONE filter - the month - and section 19 covers it:

 19a. All Months is the old report       test_all_months_is_the_unfiltered_...
      ...and is the default              test_the_default_over_http_is_all_months
      the months on offer                test_the_months_on_offer_come_with_...
 19b. August                             test_august_returns_only_august_records
 19c. September                          test_september_returns_only_september_...
      created_at, never completed_on     test_the_month_is_created_at_never_...
 19d. a month with no records            test_a_month_with_no_records_is_an_empty_...
 19e. preview dataset == download        test_the_workbook_holds_exactly_the_...
 19f. PM-only is unchanged               test_the_month_filter_does_not_widen_...
 19g. history/tab are unchanged          test_the_month_filter_does_not_touch_...

Run:  docker exec wms-backend-1 pytest tests/test_production_status_report.py -q
"""
from datetime import date, datetime, timezone
from io import BytesIO

import openpyxl
import pytest
from sqlalchemy import text

from app.modules.activity_master import service as am_svc
from app.modules.activity_master.schemas import ActivityCreate
from app.modules.plants.models import MaintenancePlant, PlanningPlant
from app.modules.production_status import service as ps_svc
from app.modules.production_status.models import ProjectProductionStatus
from app.modules.production_status.schemas import ProductionStatusCreate
from app.modules.projects import service as proj_svc
from app.modules.projects.models import ProjectStatus
from app.modules.projects.schemas import ActivityMemberCreate
from app.modules.reports_export import export
from app.modules.users.models import UserRole
from app.shared.errors import AppError

REPORT_URL = "/api/v1/production-status/report"
XLSX_URL = "/api/v1/production-status/report.xlsx"

# The workbook's columns, 1-based, so a column move is a one-line change here
# instead of a hunt through magic numbers. There is deliberately no REVISION
# column: the revision lives inside PROJECT / PLANT.
_COLUMNS = [
    "S.NO", "PROJECT / PLANT", "ACTIVITY", "PROJECT STATUS",
    "TAG", "DOC", "SPARES", "CRS",
    "COMPLETED ON", "REMARKS", "BY",
]
_COL = {label: idx for idx, label in enumerate(_COLUMNS, start=1)}

# The sheet mirrors the business's own production file: TAG/DOC/SPARES/CRS sit
# under one merged COUNT banner, so the header is TWO rows and the first data
# row is row 3.
_HEADER_ROWS = 2
_FIRST_DATA_ROW = _HEADER_ROWS + 1

# The colours the business's own file uses for the PROJECT STATUS cell.
_CLOSED_COLOUR = "FF00B050"        # green
_IN_PROGRESS_COLOUR = "FF0070C0"   # blue


def _header_labels(ws):
    """Each column's own name, wherever the two header rows put it.

    A column under the COUNT banner carries its name on row 2; every other
    column is merged across both rows and carries it on row 1. This flattens
    that back to one list, so a test can assert the columns without caring how
    the header is laid out.
    """
    return [
        ws.cell(row=2, column=c).value or ws.cell(row=1, column=c).value
        for c in range(1, len(_COLUMNS) + 1)
    ]


# --- helpers ---------------------------------------------------------------

def _activity(db, code):
    return am_svc.create_activity(db, ActivityCreate(code=code, name=code))


def _payload(activity_id, **over):
    body = dict(
        revision="REV-0",
        activity_id=activity_id,
        status="in_progress",
        tag_count=0,
        doc_count=0,
        spares_count=0,
        crs_count=0,
    )
    body.update(over)
    return ProductionStatusCreate(**body)


def _rows(report):
    return report["rows"]


def _find(report, *, project_code, revision, activity):
    """The one row for a project + revision + activity, or None."""
    for row in _rows(report):
        if (
            row.project_code == project_code
            and row.revision == revision
            and row.activity == activity
        ):
            return row
    return None


@pytest.fixture()
def scene(db, make_user, make_employee, make_project):
    """Two projects, so "cumulative" is actually exercised.

      pm         project_manager - the only role the report serves
      project_a  code GC-A, Head = head_u, activities TAG ESTIMATION + MTL,
                 Planning Plant 2300 (KAHM / KAHN)
      project_b  code GC-B, Head = head_b_u, activity MTL, same Planning Plant
      lead_u     Lead of both activities on project_a
      tag_lead_u Lead of TAG ESTIMATION ONLY, on project_a - the narrow case
      emp_u      plain contributor on project_a

    Plants mirror the real master data: one Planning Plant with two Maintenance
    Plants, plus one belonging to a DIFFERENT Planning Plant so "does not belong
    to this project" can actually be tested.
    """
    class S:
        pass

    s = S()
    s.pm = make_user("pm@x.com", role=UserRole.project_manager)

    # --- plant master data ------------------------------------------------
    s.planning = PlanningPlant(code="2300", description="Dukhan Planning Plant")
    s.other_planning = PlanningPlant(code="1200", description="Doha Planning Plant")
    db.add_all([s.planning, s.other_planning])
    db.flush()
    s.kahm = MaintenancePlant(
        code="KAHM", description="Kahma Main", planning_plant_id=s.planning.id
    )
    s.kahn = MaintenancePlant(
        code="KAHN", description="Kahma North", planning_plant_id=s.planning.id
    )
    # Belongs to another Planning Plant, so this project may never use it.
    s.foreign_plant = MaintenancePlant(
        code="GSMD", description="General Services",
        planning_plant_id=s.other_planning.id,
    )
    db.add_all([s.kahm, s.kahn, s.foreign_plant])
    db.flush()

    s.project_a = make_project(code="GC-A", name="Alpha", status=ProjectStatus.active)
    s.project_b = make_project(code="GC-B", name="Bravo", status=ProjectStatus.active)
    s.project_a.planning_plant_id = s.planning.id
    s.project_b.planning_plant_id = s.planning.id
    db.commit()

    s.head_u = make_user("head@x.com", role=UserRole.employee)
    s.head_e = make_employee(employee_code="H-1", first_name="Hema",
                             last_name="Rao", user_id=s.head_u.id)
    proj_svc.set_project_head(db, s.pm, s.project_a.id, s.head_e.id)

    s.head_b_u = make_user("headb@x.com", role=UserRole.employee)
    s.head_b_e = make_employee(employee_code="H-2", first_name="Bala",
                               last_name="Nathan", user_id=s.head_b_u.id)
    proj_svc.set_project_head(db, s.pm, s.project_b.id, s.head_b_e.id)

    s.tag_est = _activity(db, "TAG ESTIMATION")
    s.mtl = _activity(db, "MTL")

    s.lead_u = make_user("lead@x.com", role=UserRole.employee)
    s.lead_e = make_employee(employee_code="L-1", first_name="Santhosh",
                             last_name="Kumar", user_id=s.lead_u.id)

    for activity in (s.tag_est, s.mtl):
        proj_svc.assign_activity_member(
            db, s.pm, s.project_a.id, activity.id,
            ActivityMemberCreate(employee_id=s.lead_e.id, role="lead"),
        )
    proj_svc.assign_activity_member(
        db, s.pm, s.project_b.id, s.mtl.id,
        ActivityMemberCreate(employee_id=s.lead_e.id, role="lead"),
    )

    s.emp_u = make_user("emp@x.com", role=UserRole.employee)
    s.emp_e = make_employee(employee_code="C-1", user_id=s.emp_u.id)
    proj_svc.assign_activity_member(
        db, s.pm, s.project_a.id, s.tag_est.id,
        ActivityMemberCreate(employee_id=s.emp_e.id, role="contributor"),
    )

    # A contributor on BOTH activities who leads neither of them at first, then
    # takes the Lead of TAG ESTIMATION alone - the narrow-authority case the
    # Head must NOT be reduced to. (lead_u above already holds both Leads, so a
    # separate person is needed to test "one activity only".)
    s.tag_lead_u = make_user("taglead@x.com", role=UserRole.employee)
    s.tag_lead_e = make_employee(
        employee_code="L-9", first_name="Divya", last_name="Raman",
        user_id=s.tag_lead_u.id,
    )
    # Hand the TAG ESTIMATION lead over: an activity has exactly one Lead.
    proj_svc.remove_activity_member(
        db, s.pm, s.project_a.id, s.tag_est.id, s.lead_e.id
    )
    proj_svc.assign_activity_member(
        db, s.pm, s.project_a.id, s.tag_est.id,
        ActivityMemberCreate(employee_id=s.tag_lead_e.id, role="lead"),
    )
    return s


# --- 1. PM can read the report ----------------------------------------------

def test_project_manager_can_read_report(db, scene):
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, tag_count=225, status="closed",
                 completed_on=date(2025, 12, 5), remarks="Issued to client."),
    )
    report = ps_svc.cumulative_report(db, scene.pm)

    assert report["row_count"] == 1
    assert report["generated_at"] is not None
    row = _rows(report)[0]
    assert row.serial == 1
    assert row.project_code == "GC-A"
    assert row.revision == "REV-0"
    assert row.activity == "TAG ESTIMATION"
    assert row.status == "closed"
    assert row.status_label == "CLOSED"
    assert row.tag_count == 225
    assert row.completed_on == date(2025, 12, 5)
    assert row.remarks == "Issued to client."


# --- 2. Head / Lead / employee get 403 --------------------------------------

def test_non_pm_roles_are_rejected(db, scene):
    """Project-scoped read authority is deliberately NOT report authority."""
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id, _payload(scene.tag_est.id, tag_count=10)
    )

    for actor in (scene.head_u, scene.lead_u, scene.emp_u):
        with pytest.raises(AppError) as ei:
            ps_svc.cumulative_report(db, actor)
        assert ei.value.status_code == 403

    # ...while their existing project-level access is untouched. The Head still
    # reads their own project's tab, and the Lead still reads theirs.
    assert len(ps_svc.list_latest(db, scene.head_u, scene.project_a.id)) == 1
    assert len(ps_svc.list_latest(db, scene.lead_u, scene.project_a.id)) == 1


# --- 3. the same 403 over HTTP ----------------------------------------------

def test_report_endpoints_are_pm_only_over_http(db, scene, client, login):
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id, _payload(scene.tag_est.id, tag_count=10)
    )

    for email in ("head@x.com", "lead@x.com", "emp@x.com"):
        hdr = login(email)
        assert client.get(REPORT_URL, headers=hdr).status_code == 403
        assert client.get(XLSX_URL, headers=hdr).status_code == 403

    pm_hdr = login("pm@x.com")
    assert client.get(REPORT_URL, headers=pm_hdr).status_code == 200
    assert client.get(XLSX_URL, headers=pm_hdr).status_code == 200


# --- 4. multiple projects ----------------------------------------------------

def test_report_spans_every_project(db, scene):
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id, _payload(scene.tag_est.id, tag_count=225)
    )
    ps_svc.create_production_status(
        db, scene.head_b_u, scene.project_b.id, _payload(scene.mtl.id, doc_count=40)
    )

    report = ps_svc.cumulative_report(db, scene.pm)
    assert report["row_count"] == 2
    assert {r.project_code for r in _rows(report)} == {"GC-A", "GC-B"}
    # Ordered by project code, and S.NO follows that one ordering.
    assert [r.serial for r in _rows(report)] == [1, 2]
    assert [r.project_code for r in _rows(report)] == ["GC-A", "GC-B"]


# --- 5. multiple activities in one project -----------------------------------

def test_report_includes_every_activity(db, scene):
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id, _payload(scene.tag_est.id, tag_count=225)
    )
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id, _payload(scene.mtl.id, doc_count=12)
    )

    report = ps_svc.cumulative_report(db, scene.pm)
    assert report["row_count"] == 2
    assert {r.activity for r in _rows(report)} == {"TAG ESTIMATION", "MTL"}


# --- 6. REV-0 and REV-1 are separate rows ------------------------------------

def test_both_revisions_appear(db, scene):
    """A revision is part of the row's identity, not a version of it."""
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.mtl.id, revision="REV-0", tag_count=100, status="closed"),
    )
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.mtl.id, revision="REV-1", tag_count=15),
    )

    report = ps_svc.cumulative_report(db, scene.pm)
    assert report["row_count"] == 2
    assert {r.revision for r in _rows(report)} == {"REV-0", "REV-1"}
    # REV-0 is NOT discarded because a newer revision exists.
    rev0 = _find(report, project_code="GC-A", revision="REV-0", activity="MTL")
    rev1 = _find(report, project_code="GC-A", revision="REV-1", activity="MTL")
    assert rev0.tag_count == 100 and rev0.status_label == "CLOSED"
    assert rev1.tag_count == 15 and rev1.status_label == "IN PROGRESS"


# --- 7. only the latest row per project + revision + activity ----------------

def test_only_latest_row_per_combination(db, scene):
    for tag, status in ((180, "in_progress"), (210, "in_progress"), (225, "closed")):
        ps_svc.create_production_status(
            db, scene.head_u, scene.project_a.id,
            _payload(scene.tag_est.id, tag_count=tag, status=status,
                     completed_on=date(2025, 12, 5) if status == "closed" else None),
        )

    report = ps_svc.cumulative_report(db, scene.pm)
    assert report["row_count"] == 1
    row = _rows(report)[0]
    assert row.tag_count == 225
    assert row.status_label == "CLOSED"

    # The superseded updates are still in the trail - the report passed over
    # them, it did not remove them.
    assert len(ps_svc.list_history(db, scene.pm, scene.project_a.id)) == 3

    # And it is the SAME "latest" the project tab shows.
    latest = ps_svc.list_latest(db, scene.pm, scene.project_a.id)
    assert [(r.revision, r.activity_id, r.tag_count) for r in latest] == [
        (row.revision, row.activity_id, row.tag_count)
    ]


# --- 8. four independent counts ----------------------------------------------

def test_counts_are_four_independent_values(db, scene):
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, tag_count=225, doc_count=12,
                 spares_count=7, crs_count=3),
    )
    row = _rows(ps_svc.cumulative_report(db, scene.pm))[0]
    assert (row.tag_count, row.doc_count, row.spares_count, row.crs_count) == (
        225, 12, 7, 3,
    )
    # No combined total is exposed anywhere on the row.
    assert not [f for f in row.model_dump() if "total" in f]


# --- 9. BY is the actual person ----------------------------------------------

def test_by_is_the_actual_person_name(db, scene):
    """Never "PM", "Project Head" or "Activity Lead" - always the person."""
    # Recorded by the TAG ESTIMATION Lead and by the project Head respectively,
    # so two different real people appear in one report.
    ps_svc.create_production_status(
        db, scene.tag_lead_u, scene.project_a.id,
        _payload(scene.tag_est.id, tag_count=225),
    )
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id, _payload(scene.mtl.id, doc_count=4)
    )

    report = ps_svc.cumulative_report(db, scene.pm)
    by = {r.activity: r.by for r in _rows(report)}
    assert by["TAG ESTIMATION"] == "Divya Raman"
    assert by["MTL"] == "Hema Rao"
    assert not {"PM", "Project Head", "Activity Lead"} & set(by.values())


# --- 10. null completed_on stays null ----------------------------------------

def test_null_completed_on_stays_null(db, scene):
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id, _payload(scene.tag_est.id, tag_count=180)
    )
    row = _rows(ps_svc.cumulative_report(db, scene.pm))[0]
    assert row.completed_on is None

    # ...and the workbook leaves the cell genuinely empty, not "-" and not a
    # date somebody guessed.
    ws = _sheet(ps_svc.cumulative_report(db, scene.pm))
    assert ws.cell(row=_FIRST_DATA_ROW, column=_COL["COMPLETED ON"]).value is None


# --- 11. remarks are preserved in full ---------------------------------------

def test_full_remarks_are_preserved(db, scene):
    remark = (
        "Tag estimation completed for all 225 tags across the three units.\n"
        "Punch list received from the client on 02-Dec and closed on 04-Dec.\n"
        "Awaiting formal sign-off before the revision is frozen."
    )
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, tag_count=225, remarks=remark),
    )
    report = ps_svc.cumulative_report(db, scene.pm)
    assert _rows(report)[0].remarks == remark

    # Untruncated in the workbook too - the column is wrapped, not clipped.
    ws = _sheet(report)
    assert ws.cell(row=_FIRST_DATA_ROW, column=_COL["REMARKS"]).value == remark
    assert ws.cell(row=_FIRST_DATA_ROW, column=_COL["REMARKS"]).alignment.wrap_text is True


# --- 12. zero counts stay numeric zero ---------------------------------------

def test_zero_counts_stay_numeric_zero(db, scene):
    """A 0 must reach Excel as the NUMBER 0, never as "-".

    A text placeholder in one cell would break sorting, filtering and SUM() for
    the whole column - which is the entire point of exporting a spreadsheet.
    """
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, tag_count=225, doc_count=0,
                 spares_count=0, crs_count=0),
    )
    report = ps_svc.cumulative_report(db, scene.pm)
    row = _rows(report)[0]
    assert (row.doc_count, row.spares_count, row.crs_count) == (0, 0, 0)

    ws = _sheet(report)
    counts = [_COL[u] for u in ("TAG", "DOC", "SPARES", "CRS")]
    for column in counts:
        value = ws.cell(row=_FIRST_DATA_ROW, column=column).value
        assert isinstance(value, int), f"column {column} is {value!r}, not an int"
    assert [ws.cell(row=_FIRST_DATA_ROW, column=c).value for c in counts] == [225, 0, 0, 0]


# --- 13. PROJECT / PLANT -----------------------------------------------------

def test_report_display_combines_project_plant_and_revision(db, scene):
    """<PROJECT> - <MAINTENANCE PLANT> <REVISION>, from the record's own plant."""
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, revision="REV-0",
                 maintenance_plant_id=scene.kahm.id, tag_count=1),
    )
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, revision="REV-1",
                 maintenance_plant_id=scene.kahn.id, tag_count=2),
    )

    rows = _rows(ps_svc.cumulative_report(db, scene.pm))
    assert [r.project_plant for r in rows] == [
        "GC-A - KAHM REV-0",
        "GC-A - KAHN REV-1",
    ]
    # The parts are still individually available, and the revision is still in
    # the dataset even though the workbook has no column for it.
    assert [r.maintenance_plant_code for r in rows] == ["KAHM", "KAHN"]
    assert [r.revision for r in rows] == ["REV-0", "REV-1"]


def test_report_display_omits_the_plant_when_there_is_none(db, scene):
    """No plant produces a clean project value - never "/ null" or a stray dash.

    The revision is kept: it is NOT NULL on the table and is the only thing
    telling two rows of the same project and activity apart.
    """
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, revision="REV-0", tag_count=1),
    )
    display = _rows(ps_svc.cumulative_report(db, scene.pm))[0].project_plant

    assert display == "GC-A REV-0"
    # No placeholder text and no dangling separator where the plant would have
    # been. (A bare "-" is not junk on its own - project codes contain one.)
    for junk in ("null", "None", "undefined", "/", " - ", "  "):
        assert junk not in display, f"{junk!r} leaked into {display!r}"


def test_report_display_is_the_project_alone_with_no_plant_and_no_revision(db, scene):
    """The pure project case, asserted on the formatter directly.

    A stored row always has a revision (NOT NULL + non-blank CHECK), so this
    combination cannot be produced through the API - the formatter still has to
    handle it cleanly.
    """
    assert (
        ps_svc._project_plant_display(scene.project_a, None, None) == "GC-A"
    )
    assert ps_svc._project_plant_display(scene.project_a, "KAHM", None) == "GC-A - KAHM"
    assert ps_svc._project_plant_display(scene.project_a, None, "REV-0") == "GC-A REV-0"
    # Blank strings are treated as absent, not printed as empty segments.
    assert ps_svc._project_plant_display(scene.project_a, "  ", "  ") == "GC-A"


def test_report_returns_the_selected_maintenance_plant(db, scene):
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, maintenance_plant_id=scene.kahm.id),
    )
    row = _rows(ps_svc.cumulative_report(db, scene.pm))[0]
    assert row.maintenance_plant_id == scene.kahm.id
    assert row.maintenance_plant_code == "KAHM"


def test_project_name_is_returned(db, scene):
    """The project's own identifier, straight off the project row."""
    out = ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id, _payload(scene.tag_est.id)
    )
    assert out.project_code == "GC-A"
    assert out.project_name == "Alpha"
    assert _rows(ps_svc.cumulative_report(db, scene.pm))[0].project_code == "GC-A"


# --- 13b. maintenance plant persistence + validation -------------------------

def test_selected_maintenance_plant_is_persisted(db, scene):
    out = ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, maintenance_plant_id=scene.kahm.id),
    )
    assert out.maintenance_plant_id == scene.kahm.id
    assert out.maintenance_plant_code == "KAHM"
    assert out.maintenance_plant_description == "Kahma Main"

    # Stored on the row itself, not derived at read time from the project.
    stored = db.execute(
        text("SELECT maintenance_plant_id FROM project_production_statuses WHERE id = :i"),
        {"i": out.id},
    ).scalar_one()
    assert stored == scene.kahm.id

    # And it survives a re-read through the project tab's own endpoint.
    latest = ps_svc.list_latest(db, scene.pm, scene.project_a.id)
    assert latest[0].maintenance_plant_code == "KAHM"


def test_maintenance_plant_is_optional(db, scene):
    """A project with no plant chosen records production status perfectly well."""
    out = ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id, _payload(scene.tag_est.id)
    )
    assert out.maintenance_plant_id is None
    assert out.maintenance_plant_code is None
    # Nothing is back-filled from the Planning Plant.
    assert out.planning_plant_code == "2300"
    assert out.maintenance_plant_code != out.planning_plant_code


def test_unrelated_maintenance_plant_is_rejected(db, scene):
    """A plant from another Planning Plant is a 422, never silently dropped."""
    with pytest.raises(AppError) as ei:
        ps_svc.create_production_status(
            db, scene.head_u, scene.project_a.id,
            _payload(scene.tag_est.id, maintenance_plant_id=scene.foreign_plant.id),
        )
    assert ei.value.status_code == 422
    assert "does not belong to this project" in ei.value.message

    # An id that is not a plant at all is refused the same way.
    import uuid as _uuid
    with pytest.raises(AppError) as ei:
        ps_svc.create_production_status(
            db, scene.head_u, scene.project_a.id,
            _payload(scene.tag_est.id, maintenance_plant_id=_uuid.uuid4()),
        )
    assert ei.value.status_code == 422

    # Nothing was written.
    assert db.query(ProjectProductionStatus).count() == 0


def test_inactive_maintenance_plant_cannot_be_chosen(db, scene):
    scene.kahm.is_active = False
    db.commit()
    with pytest.raises(AppError) as ei:
        ps_svc.create_production_status(
            db, scene.head_u, scene.project_a.id,
            _payload(scene.tag_est.id, maintenance_plant_id=scene.kahm.id),
        )
    assert ei.value.status_code == 422


def test_project_without_planning_plant_cannot_select_a_plant(db, scene):
    """No Planning Plant means no plants to offer - and none may be forced in."""
    scene.project_b.planning_plant_id = None
    db.commit()
    with pytest.raises(AppError) as ei:
        ps_svc.create_production_status(
            db, scene.head_b_u, scene.project_b.id,
            _payload(scene.mtl.id, maintenance_plant_id=scene.kahm.id),
        )
    assert ei.value.status_code == 422

    # ...but recording WITHOUT a plant still works.
    out = ps_svc.create_production_status(
        db, scene.head_b_u, scene.project_b.id, _payload(scene.mtl.id)
    )
    assert out.maintenance_plant_id is None


def test_rows_recorded_before_the_plant_column_still_work(db, scene):
    """Migration 0071 left old rows NULL; they must read and report normally."""
    out = ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, revision="REV-0", tag_count=225),
    )
    # Force the pre-0071 shape explicitly, as an upgraded database holds it.
    db.execute(
        text("UPDATE project_production_statuses SET maintenance_plant_id = NULL "
             "WHERE id = :i"),
        {"i": out.id},
    )
    db.commit()

    latest = ps_svc.list_latest(db, scene.pm, scene.project_a.id)
    assert latest[0].maintenance_plant_id is None
    assert latest[0].tag_count == 225

    report = _rows(ps_svc.cumulative_report(db, scene.pm))
    assert report[0].maintenance_plant_code is None
    assert report[0].project_plant == "GC-A REV-0"


def test_maintenance_plant_does_not_change_record_identity(db, scene):
    """Two updates differing only by plant are still ONE current row.

    Identity stays project + revision + activity. The newer update supersedes
    the older exactly as it did before the plant column existed, and the older
    one stays in the history.
    """
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, revision="REV-0",
                 maintenance_plant_id=scene.kahm.id, tag_count=100),
    )
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, revision="REV-0",
                 maintenance_plant_id=scene.kahn.id, tag_count=200),
    )

    report = _rows(ps_svc.cumulative_report(db, scene.pm))
    assert len(report) == 1
    assert report[0].tag_count == 200
    assert report[0].maintenance_plant_code == "KAHN"

    # The superseded update - and its plant - remain in the trail.
    history = ps_svc.list_history(db, scene.pm, scene.project_a.id)
    assert [(h.tag_count, h.maintenance_plant_code) for h in history] == [
        (200, "KAHN"),
        (100, "KAHM"),
    ]


# --- 13c. Head vs Lead activity authority ------------------------------------

def test_head_may_record_for_every_activity_on_the_project(db, scene):
    """A Head's authority is project-wide, not narrowed to activities they lead.

    This is the backend rule the Activity dropdown must mirror: for each of the
    project's activities `activity_staffing_authority` answers "full" for the
    Head, so all of them are offerable.
    """
    from app.core import authz

    for activity in (scene.tag_est, scene.mtl):
        assert (
            authz.activity_staffing_authority(
                db, scene.head_u, scene.project_a, activity.id
            )
            == "full"
        )
        ps_svc.create_production_status(
            db, scene.head_u, scene.project_a.id, _payload(activity.id, tag_count=1)
        )

    recorded = {r.activity for r in _rows(ps_svc.cumulative_report(db, scene.pm))}
    assert recorded == {"TAG ESTIMATION", "MTL"}


def test_lead_may_record_only_for_their_own_activity(db, scene):
    """The Lead's authority stays narrow - unchanged by this patch."""
    from app.core import authz

    # tag_lead leads TAG ESTIMATION only.
    assert (
        authz.activity_staffing_authority(
            db, scene.tag_lead_u, scene.project_a, scene.tag_est.id
        )
        == "lead"
    )
    assert (
        authz.activity_staffing_authority(
            db, scene.tag_lead_u, scene.project_a, scene.mtl.id
        )
        is None
    )

    ps_svc.create_production_status(
        db, scene.tag_lead_u, scene.project_a.id, _payload(scene.tag_est.id)
    )
    with pytest.raises(AppError) as ei:
        ps_svc.create_production_status(
            db, scene.tag_lead_u, scene.project_a.id, _payload(scene.mtl.id)
        )
    assert ei.value.status_code == 403


# --- 14. empty report --------------------------------------------------------

def test_empty_report(db, scene, client, login):
    """No records is a valid report, not an error."""
    report = ps_svc.cumulative_report(db, scene.pm)
    assert report["row_count"] == 0
    assert report["rows"] == []

    res = client.get(REPORT_URL, headers=login("pm@x.com"))
    assert res.status_code == 200
    assert res.json()["rows"] == []
    assert res.json()["row_count"] == 0

    # The workbook is still a valid file with its header row.
    ws = _sheet(report)
    assert ws.max_row == _HEADER_ROWS
    assert ws.cell(row=1, column=1).value == "S.NO"


# --- 15. the workbook IS the preview dataset ---------------------------------

def _sheet(report):
    """Render the report dict to a workbook and hand back the sheet."""
    payload = {**report, "rows": [r.model_dump() for r in report["rows"]]}
    buf = export.build_production_status_report_workbook(payload)
    return openpyxl.load_workbook(buf).active


def test_workbook_matches_the_report_dataset(db, scene):
    """Every preview row, in the same order, with the same values."""
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, tag_count=225, status="closed",
                 completed_on=date(2025, 12, 5), remarks="Closed."),
    )
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.mtl.id, revision="REV-1", doc_count=12),
    )
    ps_svc.create_production_status(
        db, scene.head_b_u, scene.project_b.id, _payload(scene.mtl.id, spares_count=7)
    )

    report = ps_svc.cumulative_report(db, scene.pm)
    ws = _sheet(report)

    def cell(r, label):
        return ws.cell(row=r, column=_COL[label]).value

    assert ws.max_row == report["row_count"] + _HEADER_ROWS   # + the header rows
    for offset, row in enumerate(_rows(report)):
        excel_row = offset + _FIRST_DATA_ROW
        assert cell(excel_row, "S.NO") == row.serial
        assert cell(excel_row, "PROJECT / PLANT") == row.project_plant
        assert cell(excel_row, "ACTIVITY") == row.activity
        assert cell(excel_row, "PROJECT STATUS") == row.status_label
        assert cell(excel_row, "TAG") == row.tag_count
        assert cell(excel_row, "DOC") == row.doc_count
        assert cell(excel_row, "SPARES") == row.spares_count
        assert cell(excel_row, "CRS") == row.crs_count
        completed = cell(excel_row, "COMPLETED ON")
        assert (completed.date() if completed else None) == row.completed_on
        assert (cell(excel_row, "REMARKS") or None) == row.remarks
        assert cell(excel_row, "BY") == row.by
        # The revision is not its own cell - it rides inside PROJECT / PLANT.
        assert row.revision in cell(excel_row, "PROJECT / PLANT")

    # Sequential S.NO with no gaps.
    assert [
        ws.cell(row=r, column=1).value
        for r in range(_FIRST_DATA_ROW, ws.max_row + 1)
    ] == [1, 2, 3]


# --- 16. workbook columns + formatting ---------------------------------------

def test_workbook_columns_and_formatting(db, scene):
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, tag_count=225, status="closed",
                 completed_on=date(2025, 12, 5), remarks="Issued."),
    )
    report = ps_svc.cumulative_report(db, scene.pm)
    payload = {**report, "rows": [r.model_dump() for r in report["rows"]]}
    wb = openpyxl.load_workbook(export.build_production_status_report_workbook(payload))
    ws = wb.active

    assert ws.title == "Production Status"
    assert _header_labels(ws) == _COLUMNS
    assert ws.max_column == 11
    # No REVISION column - the revision is inside PROJECT / PLANT now.
    assert "REVISION" not in _COLUMNS

    # Bold header, frozen below BOTH header rows, autofilter over the row that
    # carries every column's own name.
    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.freeze_panes == "A3"
    assert ws.auto_filter.ref == "A2:K3"

    # Reasonable, explicitly-set column widths.
    assert ws.column_dimensions["B"].width >= 30      # PROJECT / PLANT
    assert ws.column_dimensions["J"].width >= 40      # REMARKS

    # DD-MMM-YYYY on a real date cell, and wrapped remarks.
    completed = ws.cell(row=_FIRST_DATA_ROW, column=_COL["COMPLETED ON"])
    assert completed.number_format == "DD-MMM-YYYY"
    assert completed.value.date() == date(2025, 12, 5)
    assert ws.cell(row=_FIRST_DATA_ROW, column=_COL["REMARKS"]).alignment.wrap_text is True


# --- 16b. the COUNT banner and the status colours ----------------------------

def test_the_four_units_sit_under_one_merged_count_banner(db, scene):
    """The business's own layout: COUNT over TAG / DOC / SPARES / CRS.

    A HEADING, not a total - there is still no combined column, and nothing is
    summed across the four.
    """
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, tag_count=225, doc_count=17),
    )
    ws = _sheet(ps_svc.cumulative_report(db, scene.pm))

    merged = {str(r) for r in ws.merged_cells.ranges}
    # COUNT spans the four unit columns on the first header row...
    assert "E1:H1" in merged
    assert ws.cell(row=1, column=_COL["TAG"]).value == "COUNT"
    # ...and each unit keeps its own name on the second.
    assert [ws.cell(row=2, column=_COL[u]).value
            for u in ("TAG", "DOC", "SPARES", "CRS")] == ["TAG", "DOC", "SPARES", "CRS"]

    # Every other column is merged vertically across both header rows instead.
    for column in ("A", "B", "C", "D", "I", "J", "K"):
        assert f"{column}1:{column}2" in merged

    # Still four independent values, and no total column anywhere.
    assert ws.max_column == len(_COLUMNS)
    assert "TOTAL" not in _header_labels(ws)
    assert ws.cell(row=_FIRST_DATA_ROW, column=_COL["TAG"]).value == 225
    assert ws.cell(row=_FIRST_DATA_ROW, column=_COL["DOC"]).value == 17


def test_project_status_is_coloured_the_way_the_business_sheet_colours_it(db, scene):
    """CLOSED bold green, IN PROGRESS bold blue - from the reference file."""
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, status="closed", completed_on=date(2025, 12, 5)),
    )
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.mtl.id, status="in_progress"),
    )
    report = ps_svc.cumulative_report(db, scene.pm)
    ws = _sheet(report)

    by_status = {}
    for offset, row in enumerate(_rows(report)):
        cell = ws.cell(row=offset + _FIRST_DATA_ROW, column=_COL["PROJECT STATUS"])
        by_status[row.status] = cell

    assert by_status["closed"].value == "CLOSED"
    assert by_status["closed"].font.color.rgb == _CLOSED_COLOUR
    assert by_status["closed"].font.bold is True

    assert by_status["in_progress"].value == "IN PROGRESS"
    assert by_status["in_progress"].font.color.rgb == _IN_PROGRESS_COLOUR
    assert by_status["in_progress"].font.bold is True

    # The colour is keyed on the STORED status, so it cannot drift from the
    # label; and no other cell is coloured.
    plant = ws.cell(row=_FIRST_DATA_ROW, column=_COL["PROJECT / PLANT"])
    assert plant.font.bold is not True


# --- 17. the download is a valid .xlsx ---------------------------------------

def test_download_produces_a_valid_xlsx(db, scene, client, login):
    ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, tag_count=225, status="closed",
                 completed_on=date(2025, 12, 5)),
    )
    res = client.get(XLSX_URL, headers=login("pm@x.com"))

    assert res.status_code == 200
    assert res.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in res.headers["content-disposition"]
    assert res.headers["content-disposition"].endswith('.xlsx"')
    # A real ZIP-based OOXML file, openable by openpyxl.
    assert res.content[:2] == b"PK"
    ws = openpyxl.load_workbook(BytesIO(res.content)).active
    assert ws.title == "Production Status"
    assert ws.cell(row=_FIRST_DATA_ROW, column=_COL["TAG"]).value == 225

    # The downloaded file holds exactly the rows the preview endpoint served.
    preview = client.get(REPORT_URL, headers=login("pm@x.com")).json()
    assert ws.max_row == preview["row_count"] + _HEADER_ROWS
    assert (
        ws.cell(row=_FIRST_DATA_ROW, column=_COL["PROJECT / PLANT"]).value
        == preview["rows"][0]["project_plant"]
    )


# --- 18. the report is not a history export ----------------------------------

def test_report_never_returns_history_rows(db, scene):
    """Cumulative means "one current row each", not "the whole trail"."""
    for tag in (10, 20, 30, 40):
        ps_svc.create_production_status(
            db, scene.head_u, scene.project_a.id, _payload(scene.tag_est.id, tag_count=tag)
        )

    report = ps_svc.cumulative_report(db, scene.pm)
    assert report["row_count"] == 1
    assert _rows(report)[0].tag_count == 40

    # History is untouched and still complete inside the project.
    history = ps_svc.list_history(db, scene.pm, scene.project_a.id)
    assert [h.tag_count for h in history] == [40, 30, 20, 10]

# --- 19. the month filter (Phase 5) ------------------------------------------
#
# The month is the record's created_at month, NEVER completed_on: an IN PROGRESS
# record has no completion date, and filtering on one would drop exactly the rows
# a PM opens a month to look at.
#
# It is the only filter the report accepts. Everything else about it is
# unchanged - still all-project, still cumulative, still PM-only.

AUG = datetime(2026, 8, 14, 9, 0, tzinfo=timezone.utc)
SEP = datetime(2026, 9, 3, 9, 0, tzinfo=timezone.utc)


def _recorded_at(db, row, when):
    """Backdate one record so a month can actually be exercised.

    created_at is a server default (now()), so a test that does not move it can
    only ever produce rows in the current month.
    """
    db.execute(
        text("UPDATE project_production_statuses SET created_at = :ts WHERE id = :i"),
        {"ts": when, "i": row.id},
    )
    db.commit()
    return row


def _seed_two_months(db, scene):
    """Two months of records, with one combination appearing in BOTH.

      August      GC-A REV-0 TAG ESTIMATION  tag=100   (in progress)
                  GC-B REV-0 MTL             doc=5
      September   GC-A REV-0 TAG ESTIMATION  tag=200   supersedes the August one
                  GC-A REV-0 MTL             spares=9

    So: 3 rows cumulatively, 2 in August, 2 in September - and TAG ESTIMATION
    reads 100 in August and 200 in September, which is what pins "latest within
    the month" rather than "latest overall, then discard".
    """
    _recorded_at(db, ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, tag_count=100),
    ), AUG)
    _recorded_at(db, ps_svc.create_production_status(
        db, scene.head_b_u, scene.project_b.id, _payload(scene.mtl.id, doc_count=5),
    ), AUG)
    _recorded_at(db, ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, tag_count=200),
    ), SEP)
    _recorded_at(db, ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id, _payload(scene.mtl.id, spares_count=9),
    ), SEP)


def _labels(report):
    """(project, activity, the count that identifies the seeded row)."""
    return {(r.project_code, r.activity) for r in _rows(report)}


# --- 19a. All Months ---------------------------------------------------------

def test_all_months_is_the_unfiltered_cumulative_report(db, scene):
    """No month = exactly the report Phase 4 already produced."""
    _seed_two_months(db, scene)

    report = ps_svc.cumulative_report(db, scene.pm)

    # Every project, every activity, the latest row of each - across all months.
    assert report["month"] is None
    assert report["row_count"] == 3
    assert _labels(report) == {
        ("GC-A", "MTL"), ("GC-A", "TAG ESTIMATION"), ("GC-B", "MTL"),
    }
    # The September update is what TAG ESTIMATION currently reads.
    assert _find(report, project_code="GC-A", revision="REV-0",
                 activity="TAG ESTIMATION").tag_count == 200
    # Ordering and S.NO are untouched by the new field.
    assert [r.serial for r in _rows(report)] == [1, 2, 3]


def test_the_default_over_http_is_all_months(db, scene, client, login):
    """Omitting the parameter is All Months - the PM opens the dialog on the
    cumulative report, exactly as before this phase."""
    _seed_two_months(db, scene)

    body = client.get(REPORT_URL, headers=login("pm@x.com")).json()
    assert body["month"] is None
    assert body["row_count"] == 3
    assert len(body["rows"]) == 3


def test_the_months_on_offer_come_with_the_report_and_never_shrink(db, scene):
    """The dropdown's options ride on the report - no second endpoint.

    Ascending, drawn from EVERY record (so August is offered even though its
    update was superseded in September), and identical whatever month is
    currently selected - picking August must not leave August as the only thing
    selectable.
    """
    _seed_two_months(db, scene)

    for month in (None, "2026-08", "2026-09", "2026-10"):
        report = ps_svc.cumulative_report(db, scene.pm, month=month)
        assert report["months"] == ["2026-08", "2026-09"]


# --- 19b. August -------------------------------------------------------------

def test_august_returns_only_august_records(db, scene):
    _seed_two_months(db, scene)

    report = ps_svc.cumulative_report(db, scene.pm, month="2026-08")

    assert report["month"] == "2026-08"
    assert report["row_count"] == 2
    assert _labels(report) == {("GC-A", "TAG ESTIMATION"), ("GC-B", "MTL")}
    # The September-only row is absent...
    assert _find(report, project_code="GC-A", revision="REV-0", activity="MTL") is None
    # ...and the shared combination reads what AUGUST recorded, not what
    # September later replaced it with.
    assert _find(report, project_code="GC-A", revision="REV-0",
                 activity="TAG ESTIMATION").tag_count == 100
    # Serials are re-stamped over the filtered list, with no gaps.
    assert [r.serial for r in _rows(report)] == [1, 2]


# --- 19c. September ----------------------------------------------------------

def test_september_returns_only_september_records(db, scene):
    _seed_two_months(db, scene)

    report = ps_svc.cumulative_report(db, scene.pm, month="2026-09")

    assert report["month"] == "2026-09"
    assert report["row_count"] == 2
    assert _labels(report) == {("GC-A", "TAG ESTIMATION"), ("GC-A", "MTL")}
    assert _find(report, project_code="GC-A", revision="REV-0",
                 activity="TAG ESTIMATION").tag_count == 200
    # August's other project is not in September's report.
    assert _find(report, project_code="GC-B", revision="REV-0",
                 activity="MTL") is None


def test_the_month_is_created_at_never_completed_on(db, scene):
    """An IN PROGRESS record has no completion date and must still be findable.

    The August record is recorded in August and completed in December; it belongs
    to August. The September record has no completed_on at all and must not
    disappear because of it.
    """
    _recorded_at(db, ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.tag_est.id, tag_count=100, status="closed",
                 completed_on=date(2026, 12, 5)),
    ), AUG)
    _recorded_at(db, ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id,
        _payload(scene.mtl.id, status="in_progress", completed_on=None),
    ), SEP)

    august = ps_svc.cumulative_report(db, scene.pm, month="2026-08")
    assert august["row_count"] == 1
    assert august["rows"][0].activity == "TAG ESTIMATION"
    assert august["rows"][0].completed_on == date(2026, 12, 5)

    # December - the completion month - holds nothing, and is not even offered.
    assert ps_svc.cumulative_report(db, scene.pm, month="2026-12")["row_count"] == 0
    assert august["months"] == ["2026-08", "2026-09"]

    # The record with no completion date is present in the month it was recorded.
    september = ps_svc.cumulative_report(db, scene.pm, month="2026-09")
    assert september["row_count"] == 1
    assert september["rows"][0].activity == "MTL"
    assert september["rows"][0].completed_on is None


def test_month_boundaries_are_half_open(db, scene):
    """The last instant of a month is in it; the first of the next is not."""
    _recorded_at(db, ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id, _payload(scene.tag_est.id, tag_count=1),
    ), datetime(2026, 8, 31, 23, 59, 59, tzinfo=timezone.utc))
    _recorded_at(db, ps_svc.create_production_status(
        db, scene.head_u, scene.project_a.id, _payload(scene.mtl.id, tag_count=2),
    ), datetime(2026, 9, 1, 0, 0, 0, tzinfo=timezone.utc))

    august = ps_svc.cumulative_report(db, scene.pm, month="2026-08")
    september = ps_svc.cumulative_report(db, scene.pm, month="2026-09")
    assert [r.activity for r in _rows(august)] == ["TAG ESTIMATION"]
    assert [r.activity for r in _rows(september)] == ["MTL"]


# --- 19d. a month with no records --------------------------------------------

def test_a_month_with_no_records_is_an_empty_report(db, scene, client, login):
    """Empty is a valid answer, not an error - the same way the whole report
    already treats "nothing recorded yet"."""
    _seed_two_months(db, scene)

    report = ps_svc.cumulative_report(db, scene.pm, month="2026-10")
    assert report["month"] == "2026-10"
    assert report["row_count"] == 0
    assert report["rows"] == []
    # The choice is still on offer - an empty month does not empty the dropdown.
    assert report["months"] == ["2026-08", "2026-09"]

    res = client.get(f"{REPORT_URL}?month=2026-10", headers=login("pm@x.com"))
    assert res.status_code == 200
    assert res.json()["row_count"] == 0

    # And the download is still a real, openable workbook with its header row.
    res = client.get(f"{XLSX_URL}?month=2026-10", headers=login("pm@x.com"))
    assert res.status_code == 200
    ws = openpyxl.load_workbook(BytesIO(res.content)).active
    assert ws.max_row == _HEADER_ROWS
    assert ws.cell(row=1, column=1).value == "S.NO"


def test_a_malformed_month_is_rejected(db, scene):
    """A bad month is refused rather than silently ignored - a client sending
    junk must fail loudly instead of receiving an unfiltered report it will
    present as filtered."""
    _seed_two_months(db, scene)

    for bad in ("2026", "august", "2026-13", "2026-00", "26-08", "2026-08-14"):
        with pytest.raises(AppError) as ei:
            ps_svc.cumulative_report(db, scene.pm, month=bad)
        assert ei.value.status_code == 422

    # A blank month is not junk - it is "no filter", i.e. All Months.
    assert ps_svc.cumulative_report(db, scene.pm, month="  ")["month"] is None


# --- 19e. the preview and the download are ONE dataset -----------------------

@pytest.mark.parametrize("month", [None, "2026-08", "2026-09", "2026-10"])
def test_the_workbook_holds_exactly_the_filtered_preview_rows(db, scene, month):
    """Whatever the filter, the file is the screen.

    Both renderers read the SAME service call with the SAME month, so this is
    structural - there is no second query to drift.
    """
    _seed_two_months(db, scene)

    report = ps_svc.cumulative_report(db, scene.pm, month=month)
    ws = _sheet(report)

    assert ws.max_row == report["row_count"] + _HEADER_ROWS   # + the header rows
    for offset, row in enumerate(_rows(report)):
        excel_row = offset + _FIRST_DATA_ROW
        assert ws.cell(row=excel_row, column=_COL["S.NO"]).value == row.serial
        assert ws.cell(row=excel_row, column=_COL["PROJECT / PLANT"]).value == (
            row.project_plant
        )
        assert ws.cell(row=excel_row, column=_COL["ACTIVITY"]).value == row.activity
        assert ws.cell(row=excel_row, column=_COL["TAG"]).value == row.tag_count
        assert ws.cell(row=excel_row, column=_COL["DOC"]).value == row.doc_count
        assert ws.cell(row=excel_row, column=_COL["SPARES"]).value == row.spares_count
        assert ws.cell(row=excel_row, column=_COL["CRS"]).value == row.crs_count
    # The columns and their order are untouched by the filter.
    assert _header_labels(ws) == _COLUMNS


@pytest.mark.parametrize("query,expected", [("", 3), ("?month=2026-08", 2),
                                            ("?month=2026-09", 2)])
def test_preview_and_download_agree_over_http(db, scene, client, login, query, expected):
    """The two endpoints, same month, same rows - checked end to end."""
    _seed_two_months(db, scene)
    hdr = login("pm@x.com")

    preview = client.get(f"{REPORT_URL}{query}", headers=hdr).json()
    assert preview["row_count"] == expected

    res = client.get(f"{XLSX_URL}{query}", headers=hdr)
    assert res.status_code == 200
    ws = openpyxl.load_workbook(BytesIO(res.content)).active

    assert ws.max_row == preview["row_count"] + _HEADER_ROWS
    for offset, row in enumerate(preview["rows"]):
        excel_row = offset + _FIRST_DATA_ROW
        assert ws.cell(row=excel_row, column=_COL["PROJECT / PLANT"]).value == (
            row["project_plant"]
        )
        assert ws.cell(row=excel_row, column=_COL["ACTIVITY"]).value == row["activity"]
        assert ws.cell(row=excel_row, column=_COL["S.NO"]).value == row["serial"]


def test_a_month_filtered_download_is_named_for_that_month(db, scene, client, login):
    _seed_two_months(db, scene)

    res = client.get(f"{XLSX_URL}?month=2026-08", headers=login("pm@x.com"))
    assert "PRODUCTION STATUS [AUG 2026].xlsx" in res.headers["content-disposition"]

    # The unfiltered name is unchanged - still dated for the day it was taken.
    assert ps_svc.report_filename(today=date(2026, 8, 21)) == (
        "PRODUCTION STATUS [21 AUG 2026].xlsx"
    )
    assert ps_svc.report_filename(month="2026-09") == (
        "PRODUCTION STATUS [SEP 2026].xlsx"
    )


# --- 19f. authorization is unchanged -----------------------------------------

def test_the_month_filter_does_not_widen_who_may_read_the_report(db, scene, client, login):
    """The filter narrows what a PM sees. It never changes who may see it."""
    _seed_two_months(db, scene)

    for actor in (scene.head_u, scene.lead_u, scene.emp_u):
        for month in (None, "2026-08", "2026-09", "2026-10"):
            with pytest.raises(AppError) as ei:
                ps_svc.cumulative_report(db, actor, month=month)
            assert ei.value.status_code == 403

    # Over HTTP too, on both routes, with the month on the query string.
    for email in ("head@x.com", "lead@x.com", "emp@x.com"):
        hdr = login(email)
        assert client.get(f"{REPORT_URL}?month=2026-08", headers=hdr).status_code == 403
        assert client.get(f"{XLSX_URL}?month=2026-08", headers=hdr).status_code == 403

    pm_hdr = login("pm@x.com")
    assert client.get(f"{REPORT_URL}?month=2026-08", headers=pm_hdr).status_code == 200
    assert client.get(f"{XLSX_URL}?month=2026-08", headers=pm_hdr).status_code == 200

    # A 403 comes BEFORE the month is even looked at - an unauthorized caller
    # never learns whether a month is valid.
    with pytest.raises(AppError) as ei:
        ps_svc.cumulative_report(db, scene.emp_u, month="not-a-month")
    assert ei.value.status_code == 403


# --- 19g. nothing else changed -----------------------------------------------

def test_the_month_filter_does_not_touch_history_or_the_project_tab(db, scene):
    """The filter belongs to the report. Recording and history are untouched."""
    _seed_two_months(db, scene)

    # The project's own tab still shows its current status across all months.
    latest = ps_svc.list_latest(db, scene.pm, scene.project_a.id)
    assert {(r.activity_name, r.tag_count) for r in latest} == {
        ("TAG ESTIMATION", 200), ("MTL", 0),
    }

    # The append-only trail is complete - both months, nothing filtered away.
    history = ps_svc.list_history(
        db, scene.pm, scene.project_a.id, activity_id=scene.tag_est.id
    )
    assert [h.tag_count for h in history] == [200, 100]
