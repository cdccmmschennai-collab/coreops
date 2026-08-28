"""LS (lumpsum) activities: a free-typed Count plus the unit it belongs to.

A lumpsum sub-activity (TASK_STATUS_ONLY, or the legacy TASK_BASED) configures
no relevant_count_field â€” it is measured by completion within a duration. The
report form now lets the employee record a count against it anyway, naming the
unit beside the value ("Count [25] [Tags]"). No new column and no migration: the
value goes in the existing tags_count/docs_count/... column and the chosen unit
is frozen in the existing work_report_tasks.relevant_count_field_snapshot, which
is NULL and unused on a lumpsum row.

What these tests pin:
  - a lumpsum row now REQUIRES a count field + its value (both halves) —
    neither is optional the way it once was;
  - the unit is validated against activity_master's own VALID_COUNT_FIELDS, so
    an arbitrary field name from a client is rejected, never stored;
  - it survives save -> read -> submit (submit must not overwrite it with the
    master's NULL) and comes back for editing;
  - "Other counts" entered on the same row are neither cleared nor overwritten;
  - the Weekly Activity Report puts the value under the chosen unit's own
    column, and every other unit under its own;
  - TASK_WITH_QUANTITY and the numeric modes are untouched â€” their unit still
    comes from Activity Master and a count_field sent for one is ignored.
"""
from datetime import date
from io import BytesIO

import openpyxl
import pytest

from app.modules.activity_master.models import (
    VALID_COUNT_FIELDS,
    is_lumpsum_unit_row,
)
from app.modules.projects.models import ProjectStatus
from app.modules.users.models import UserRole

BASE = "/api/v1/work-reports"
AM = "/api/v1/activity-master"
ROWS = "/api/v1/reports-export/activity-rows"
XLSX = "/api/v1/reports-export/activity-rows.xlsx"
TODAY = date.today().isoformat()


@pytest.fixture()
def setup_author(make_user, make_employee, make_project, make_project_member, login):
    def _make(*, email="emp@x.com", code="E-1", proj_code="P-1"):
        u = make_user(email, role=UserRole.employee)
        e = make_employee(employee_code=code, user_id=u.id)
        p = make_project(code=proj_code, status=ProjectStatus.active)
        make_project_member(project_id=p.id, employee_id=e.id)
        return {"user": u, "emp": e, "project": p, "header": login(email)}

    return _make


@pytest.fixture()
def pm_header(auth_header):
    return auth_header(email="pm@x.com", role=UserRole.project_manager)


def _make_sub(client, admin_header, *, activity_name="MTL", name="Sub", **body):
    a = client.post(
        f"{AM}/activities", json={"name": activity_name}, headers=admin_header
    ).json()
    res = client.post(
        f"{AM}/activities/{a['id']}/sub-activities",
        json={"name": name, **body},
        headers=admin_header,
    )
    assert res.status_code in (200, 201), res.text
    return a, res.json()


def _lumpsum_sub(client, admin_header, *, name="LS PUNCH LIST", days=1):
    """A real lumpsum: a task duration, no quantity, no configured unit."""
    return _make_sub(
        client, admin_header, name=name,
        benchmark_type="TASK_STATUS_ONLY", benchmark_period_days=days,
    )


def _create(client, header, tasks, report_date=TODAY):
    res = client.post(
        BASE, headers=header, json={"report_date": report_date, "tasks": tasks}
    )
    assert res.status_code in (200, 201), res.text
    return res.json()


def _submit(client, header, report_id):
    res = client.post(f"{BASE}/{report_id}/submit", headers=header)
    assert res.status_code == 200, res.text
    return res.json()


def _task_row(project_id, sub_id, **counts):
    return {
        "project_id": str(project_id),
        "description": "w",
        "sub_activity_id": sub_id,
        **counts,
    }


# â”€â”€ the shared predicate â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


def test_only_a_unitless_task_mode_takes_an_employee_chosen_unit():
    """The one rule the backend and the form both resolve through. A TASK mode
    WITH a configured unit (TASK_WITH_QUANTITY) is deliberately not a lumpsum
    row here: its unit belongs to its benchmark."""
    assert is_lumpsum_unit_row("TASK_STATUS_ONLY", None) is True
    assert is_lumpsum_unit_row("TASK_BASED", None) is True       # legacy value
    assert is_lumpsum_unit_row("TASK_WITH_QUANTITY", "spares") is False
    assert is_lumpsum_unit_row("NUMERIC_DAILY", "tags") is False
    assert is_lumpsum_unit_row("NUMERIC", "docs") is False
    assert is_lumpsum_unit_row(None, None) is False              # no benchmark


def test_the_allowed_units_are_activity_masters_own_list():
    assert VALID_COUNT_FIELDS == {
        "tags", "docs", "bom", "spares", "pages", "records"
    }


# â”€â”€ storing the pick â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


def test_ls_count_is_stored_under_the_chosen_unit(client, setup_author, pm_header):
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)

    created = _create(
        client, a["header"],
        [_task_row(a["project"].id, sub["id"], tags_count=25, count_field="tags")],
    )

    task = created["tasks"][0]
    assert task["tags_count"] == 25
    assert task["relevant_count_field_snapshot"] == "tags"
    # The value lands in ONE column; the other five stay at zero.
    assert task["docs_count"] == 0
    assert task["bom_count"] == 0


def test_a_different_unit_stores_in_its_own_column(client, setup_author, pm_header):
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)

    created = _create(
        client, a["header"],
        [_task_row(a["project"].id, sub["id"], docs_count=50, count_field="docs")],
    )

    task = created["tasks"][0]
    assert task["docs_count"] == 50
    assert task["relevant_count_field_snapshot"] == "docs"
    assert task["tags_count"] == 0


@pytest.mark.parametrize("unit", ["tags", "docs", "bom", "spares", "pages", "records"])
def test_every_supported_unit_round_trips(client, setup_author, pm_header, unit):
    a = setup_author(email=f"e-{unit}@x.com", code=f"E-{unit}", proj_code=f"P-{unit}")
    _, sub = _lumpsum_sub(client, pm_header, name=f"LS {unit}")

    created = _create(
        client, a["header"],
        [_task_row(
            a["project"].id, sub["id"], **{f"{unit}_count": 7}, count_field=unit
        )],
    )

    task = created["tasks"][0]
    assert task["relevant_count_field_snapshot"] == unit
    assert task[f"{unit}_count"] == 7


def test_other_counts_are_preserved_beside_the_ls_count(client, setup_author, pm_header):
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)

    created = _create(
        client, a["header"],
        [_task_row(
            a["project"].id, sub["id"],
            tags_count=25,      # the LS Count
            count_field="tags",
            docs_count=10,      # entered under "Other counts"
            bom_count=5,
            spares_count=3,
        )],
    )

    task = created["tasks"][0]
    assert (task["tags_count"], task["docs_count"], task["bom_count"],
            task["spares_count"]) == (25, 10, 5, 3)
    assert task["relevant_count_field_snapshot"] == "tags"


# â”€â”€ validation â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


def test_an_unknown_count_field_is_rejected(client, setup_author, pm_header):
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)

    res = client.post(
        BASE, headers=a["header"],
        json={
            "report_date": TODAY,
            "tasks": [_task_row(
                a["project"].id, sub["id"], tags_count=25, count_field="hours"
            )],
        },
    )
    assert res.status_code == 422, res.text


def test_a_column_name_is_not_a_unit(client, setup_author, pm_header):
    """'tags_count' is the column; 'tags' is the unit. Sending the column name
    must not quietly work â€” it is not in the source-of-truth list."""
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)

    res = client.post(
        BASE, headers=a["header"],
        json={
            "report_date": TODAY,
            "tasks": [_task_row(
                a["project"].id, sub["id"], tags_count=25, count_field="tags_count"
            )],
        },
    )
    assert res.status_code == 422, res.text


# ── the mandatory requirement: an LS row must name its field AND its value ───
#
# A lumpsum sub-activity is measured by completion, not quantity, so nothing
# else on the row forces a number to be entered. Once the count row exists at
# all it must be filled in both halves — a field with nothing counted, or a
# count with nothing to attribute it to, are both refused rather than saved
# silently short. TASK_WITH_QUANTITY and the numeric modes are untouched: the
# gate below only ever fires for `is_lumpsum_task`, computed the same way the
# rest of the module already resolves it (`is_lumpsum_unit_row`).


def test_a_count_with_its_field_saves(client, setup_author, pm_header):
    """1. LS with count field + value → allowed."""
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)

    created = _create(
        client, a["header"],
        [_task_row(
            a["project"].id, sub["id"], count_field="tags", count_value=25
        )],
    )

    task = created["tasks"][0]
    assert task["relevant_count_field_snapshot"] == "tags"
    assert task["tags_count"] == 25


def test_no_count_and_no_field_is_rejected(client, setup_author, pm_header):
    """2. LS without count field → rejected. Nothing at all on the row — the
    server refuses rather than saving a lumpsum activity with nothing
    counted."""
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)

    res = client.post(
        BASE, headers=a["header"],
        json={
            "report_date": TODAY,
            "tasks": [_task_row(a["project"].id, sub["id"])],
        },
    )
    assert res.status_code == 422, res.text
    assert "count" in res.text.lower()


def test_other_counts_alone_do_not_satisfy_the_requirement(
    client, setup_author, pm_header
):
    """2b. "Other counts" are attributed by construction (each is typed into
    its own named input), so they never stand in for the required field — an
    LS row carrying only those is still rejected."""
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)

    res = client.post(
        BASE, headers=a["header"],
        json={
            "report_date": TODAY,
            "tasks": [_task_row(
                a["project"].id, sub["id"], docs_count=10, bom_count=5
            )],
        },
    )
    assert res.status_code == 422, res.text


def test_a_count_with_no_field_is_rejected(client, setup_author, pm_header):
    """3. LS with count field but missing value → rejected. 25 typed with the
    picker left on "Select field" — the server will not guess a column, so it
    refuses rather than filing the number somewhere."""
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)

    res = client.post(
        BASE, headers=a["header"],
        json={
            "report_date": TODAY,
            "tasks": [_task_row(a["project"].id, sub["id"], count_value=25)],
        },
    )
    assert res.status_code == 422, res.text
    assert "field" in res.text.lower()


def test_a_field_with_no_value_is_rejected(client, setup_author, pm_header):
    """3b. The field is picked, but nothing is typed against it (the column
    stays at its untouched 0) — rejected, not silently dropped."""
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)

    res = client.post(
        BASE, headers=a["header"],
        json={
            "report_date": TODAY,
            "tasks": [_task_row(a["project"].id, sub["id"], count_field="tags")],
        },
    )
    assert res.status_code == 422, res.text


def test_a_field_with_an_explicit_zero_is_rejected(client, setup_author, pm_header):
    """3c. An explicit 0 is the same as no value — it is what an untouched
    unit already reads, so it is not a count."""
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)

    res = client.post(
        BASE, headers=a["header"],
        json={
            "report_date": TODAY,
            "tasks": [_task_row(
                a["project"].id, sub["id"], count_field="tags", count_value=0
            )],
        },
    )
    assert res.status_code == 422, res.text


def test_editing_an_ls_row_to_drop_the_count_is_rejected(
    client, setup_author, pm_header
):
    """The rule applies on update too, not just create: a draft that already
    has a valid count cannot be re-saved with the count row cleared."""
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)
    created = _create(
        client, a["header"],
        [_task_row(a["project"].id, sub["id"], count_field="tags", count_value=25)],
    )

    res = client.patch(
        f"{BASE}/{created['id']}", headers=a["header"],
        json={"tasks": [_task_row(a["project"].id, sub["id"])]},
    )
    assert res.status_code == 422, res.text


def test_a_quantity_row_ignores_a_count_field_it_never_asked_for(
    client, setup_author, pm_header
):
    """A benchmarked row's unit belongs to its benchmark. A count_field riding
    along on one (an edited row whose mode changed, say) is CLEARED, so it can
    never repoint the actual the target is measured against."""
    a = setup_author()
    _, sub = _make_sub(
        client, pm_header, name="NUM PAGES", benchmark_type="NUMERIC_DAILY",
        benchmark_value=500, relevant_count_field="pages",
    )

    created = _create(
        client, a["header"],
        [_task_row(
            a["project"].id, sub["id"], pages_count=400, count_field="tags"
        )],
    )
    # Nothing stored at save; submit freezes the MASTER's unit, as before.
    assert created["tasks"][0]["relevant_count_field_snapshot"] is None

    submitted = _submit(client, a["header"], created["id"])
    assert submitted["tasks"][0]["relevant_count_field_snapshot"] == "pages"


def test_task_with_quantity_keeps_its_configured_unit(client, setup_author, pm_header):
    """TASK_WITH_QUANTITY is a task mode WITH a unit â€” it is not a lumpsum row,
    and the unit shown/stored is whatever Activity Master configured, never one
    hardcoded field."""
    a = setup_author()
    _, sub = _make_sub(
        client, pm_header, name="TWQ SPARES", benchmark_type="TASK_WITH_QUANTITY",
        benchmark_value=300, benchmark_period_days=2, relevant_count_field="spares",
    )

    created = _create(
        client, a["header"],
        [_task_row(
            a["project"].id, sub["id"], spares_count=120, count_field="records"
        )],
    )
    submitted = _submit(client, a["header"], created["id"])

    task = submitted["tasks"][0]
    assert task["relevant_count_field_snapshot"] == "spares"
    assert task["spares_count"] == 120
    assert task["benchmark_type_snapshot"] == "TASK_WITH_QUANTITY"


def test_a_quantity_row_still_saves_and_submits_untouched(
    client, setup_author, pm_header
):
    """The conditional rule must not reach a benchmarked row: its count is
    typed into the unit its benchmark names, so it is attributed already and
    needs no picker."""
    a = setup_author()
    _, sub = _make_sub(
        client, pm_header, name="NUM TAGS", benchmark_type="NUMERIC_DAILY",
        benchmark_value=250, relevant_count_field="tags",
    )

    created = _create(
        client, a["header"],
        [_task_row(a["project"].id, sub["id"], tags_count=200)],
    )
    submitted = _submit(client, a["header"], created["id"])

    task = submitted["tasks"][0]
    assert task["tags_count"] == 200
    assert task["relevant_count_field_snapshot"] == "tags"
    assert task["benchmark_value_snapshot"] == "250.00"


def test_an_activity_with_no_benchmark_still_saves(client, setup_author, pm_header):
    """A pure logging line item (benchmark_type NULL - LEAVE, TRAINING) has no
    Count row at all and must be unaffected in every direction."""
    a = setup_author()
    _, sub = _make_sub(client, pm_header, name="TRAINING")

    created = _create(client, a["header"], [_task_row(a["project"].id, sub["id"])])
    submitted = _submit(client, a["header"], created["id"])

    task = submitted["tasks"][0]
    assert task["relevant_count_field_snapshot"] is None
    assert task["benchmark_type_snapshot"] is None


# â”€â”€ surviving submit and edit â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


def test_submit_does_not_erase_the_ls_unit(client, setup_author, pm_header):
    """The regression this guards: _apply_benchmarks freezes the master's
    relevant_count_field at submit, which is NULL for a lumpsum â€” overwriting
    would drop the employee's pick the moment they filed the report."""
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)

    created = _create(
        client, a["header"],
        [_task_row(
            a["project"].id, sub["id"], tags_count=25, count_field="tags",
            docs_count=10,
        )],
    )
    submitted = _submit(client, a["header"], created["id"])

    task = submitted["tasks"][0]
    assert task["relevant_count_field_snapshot"] == "tags"
    assert task["tags_count"] == 25
    assert task["docs_count"] == 10


def test_reopening_a_draft_returns_the_pick_for_editing(client, setup_author, pm_header):
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)
    created = _create(
        client, a["header"],
        [_task_row(
            a["project"].id, sub["id"], tags_count=25, count_field="tags",
            docs_count=10,
        )],
    )

    fetched = client.get(f"{BASE}/{created['id']}", headers=a["header"]).json()

    task = fetched["tasks"][0]
    assert task["relevant_count_field_snapshot"] == "tags"
    assert task["tags_count"] == 25
    assert task["docs_count"] == 10


def test_editing_can_repoint_the_unit(client, setup_author, pm_header):
    """Changing the picker moves the Count: 25 against Tags becomes 40 against
    Docs, and the Other counts entered beside it are untouched."""
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)
    created = _create(
        client, a["header"],
        [_task_row(
            a["project"].id, sub["id"], tags_count=25, count_field="tags",
            bom_count=5,
        )],
    )

    res = client.patch(
        f"{BASE}/{created['id']}", headers=a["header"],
        json={"tasks": [_task_row(
            a["project"].id, sub["id"], docs_count=40, count_field="docs",
            bom_count=5,
        )]},
    )
    assert res.status_code == 200, res.text

    task = res.json()["tasks"][0]
    assert task["relevant_count_field_snapshot"] == "docs"
    assert task["docs_count"] == 40
    assert task["tags_count"] == 0
    assert task["bom_count"] == 5


# â”€â”€ the export â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


def _first_activity(client, header):
    body = client.get(ROWS, headers=header).json()
    assert body["rows"], "expected at least one report row"
    acts = body["rows"][0]["activities"]
    assert acts, "expected at least one activity on the row"
    return acts[0]


def test_export_rows_carry_the_ls_count_under_its_own_unit(
    client, setup_author, pm_header
):
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)
    created = _create(
        client, a["header"],
        [_task_row(
            a["project"].id, sub["id"],
            tags_count=25, count_field="tags",
            docs_count=10, bom_count=5, spares_count=3,
        )],
    )
    _submit(client, a["header"], created["id"])

    activity = _first_activity(client, pm_header)
    assert activity["tags"] == 25
    assert activity["docs"] == 10
    assert activity["bom"] == 5
    assert activity["spares"] == 3
    # Still a lumpsum in the BENCHMARK cell â€” a recorded count is not a target.
    assert activity["benchmark_type"] == "TASK_STATUS_ONLY"


def test_xlsx_puts_the_ls_count_in_the_chosen_units_column(
    client, setup_author, pm_header
):
    """25 under Tags means the NO. OF TAGS column â€” never a generic Count
    column, and never a hardcoded one. The Other counts keep their own columns
    on the same row."""
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)
    created = _create(
        client, a["header"],
        [_task_row(
            a["project"].id, sub["id"],
            tags_count=25, count_field="tags",
            docs_count=10, bom_count=5, spares_count=3,
        )],
    )
    _submit(client, a["header"], created["id"])

    res = client.get(XLSX, headers=pm_header)
    assert res.status_code == 200, res.text
    ws = openpyxl.load_workbook(BytesIO(res.content)).active

    headers = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    cell = dict(zip(headers, row))
    assert cell["NO. OF TAGS"] == 25
    assert cell["NO. OF DOCS"] == 10
    assert cell["NO. OF BOM HEADER"] == 5
    assert cell["NO. OF SPARES"] == 3
    assert cell["NO. OF PAGES"] == 0
    assert cell["BENCHMARK"] == "LS"


def test_xlsx_follows_the_chosen_unit_when_it_changes(
    client, setup_author, pm_header
):
    """The same workbook, the same activity, a different pick: 50 must land
    under DOCS, and TAGS must stay empty."""
    a = setup_author()
    _, sub = _lumpsum_sub(client, pm_header)
    created = _create(
        client, a["header"],
        [_task_row(
            a["project"].id, sub["id"], docs_count=50, count_field="docs"
        )],
    )
    _submit(client, a["header"], created["id"])

    res = client.get(XLSX, headers=pm_header)
    ws = openpyxl.load_workbook(BytesIO(res.content)).active
    cell = dict(zip([c.value for c in ws[1]], [c.value for c in ws[2]]))

    assert cell["NO. OF DOCS"] == 50
    assert cell["NO. OF TAGS"] == 0
