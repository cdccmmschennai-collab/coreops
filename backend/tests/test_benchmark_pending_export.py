"""Full-cycle Benchmark XLSX export (GET /benchmarks/pending-export.xlsx) and
the Fri..Thu cycle bounds behind it.

Layout and styling are matched cell-for-cell to the company reference workbook
(BENCHMARK REPORT 03 JUL - 09 JUL) plus the DAY PART column: 29 columns A..AC
(six units per group: TAGS DOCS BOM SPARES PAGES RECORDS), a two-level yellow
header,
white body rows, bold white TOTAL rows, and a red/green shade on the
DIFFERENCE % cell ONLY. Each NUMERIC sub-activity gets its own TOTAL row
(per-unit target/actual/net-pending + its own uncapped achievement % and the
absolute difference from 100%). A purely TEXTUAL task sub-activity shows detail
rows only — no total, no percentages, no shading. Compensation is scoped to one
employee + cycle + sub-activity + unit and never crosses those bounds."""
from datetime import date, datetime, timedelta
from io import BytesIO

import openpyxl
import pytest
from openpyxl.utils import get_column_letter

from app.modules.activity_master.service import compute_week_bounds
from app.modules.projects.models import ProjectStatus
from app.modules.reports_export.export import (
    _BORDER,
    _DATA_FONT,
    _PB_HEADER_FILL,
    _PB_HEADER_FONT,
    _PB_TOTAL_FONT,
    PENDING_SHEET_NAME,
    date_range_label,
)
from app.modules.users.models import UserRole

BASE = "/api/v1/work-reports"
EXPORT_URL = "/api/v1/benchmarks/pending-export.xlsx"
TODAY_D = date.today()


@pytest.fixture()
def setup_author(make_user, make_employee, make_project, make_project_member, login):
    def _make(*, email="emp@x.com", code="E-1", proj_code="P-1", first_name="Test", last_name="User"):
        u = make_user(email, role=UserRole.employee)
        e = make_employee(employee_code=code, user_id=u.id, first_name=first_name, last_name=last_name)
        p = make_project(code=proj_code, status=ProjectStatus.active)
        make_project_member(project_id=p.id, employee_id=e.id)
        return {"user": u, "emp": e, "project": p, "header": login(email)}

    return _make


@pytest.fixture()
def activity_admin(auth_header):
    return auth_header(email="pm@x.com", role=UserRole.project_manager)


def _make_sub_activity(client, admin_header, *, benchmark_value, name="Sub", count_field="tags"):
    a = client.post(
        "/api/v1/activity-master/activities", json={"name": f"Activity for {name}"}, headers=admin_header
    ).json()
    sub = client.post(
        f"/api/v1/activity-master/activities/{a['id']}/sub-activities",
        json={
            "name": name, "benchmark_type": "NUMERIC",
            "benchmark_value": benchmark_value, "relevant_count_field": count_field,
        },
        headers=admin_header,
    ).json()
    return a, sub


def _make_task_sub(client, admin_header, *, name, period=1):
    """TASK_BASED (lumpsum) sub-activity with a `period`-day completion window."""
    a = client.post(
        "/api/v1/activity-master/activities", json={"name": f"Activity for {name}"}, headers=admin_header
    ).json()
    sub = client.post(
        f"/api/v1/activity-master/activities/{a['id']}/sub-activities",
        json={"name": name, "benchmark_type": "TASK_BASED"},
        headers=admin_header,
    ).json()
    client.patch(
        f"/api/v1/activity-master/sub-activities/{sub['id']}",
        json={"benchmark_period_days": period},
        headers=admin_header,
    )
    return a, sub


def _submit(client, header, project_id, sub_id, report_date, qty, count_field="tags"):
    payload = {
        "report_date": report_date.isoformat(),
        "tasks": [{
            "project_id": str(project_id), "description": "work",
            "sub_activity_id": sub_id, f"{count_field}_count": qty,
        }],
    }
    created = client.post(BASE, headers=header, json=payload)
    assert created.status_code == 201, created.text
    res = client.post(f"{BASE}/{created.json()['id']}/submit", headers=header)
    assert res.status_code == 200, res.text


def _submit_task(client, header, project_id, sub_id, report_date):
    """Submit a plain task row (no counts) and return the created task id."""
    payload = {
        "report_date": report_date.isoformat(),
        "tasks": [{"project_id": str(project_id), "description": "x", "sub_activity_id": sub_id}],
    }
    body = client.post(BASE, headers=header, json=payload).json()
    assert client.post(f"{BASE}/{body['id']}/submit", headers=header).status_code == 200
    return body["tasks"][0]["id"]


def _prev_cycle() -> tuple[date, date]:
    start, end = compute_week_bounds(TODAY_D)
    return start - timedelta(days=7), end - timedelta(days=7)


def _cell_date(value):
    return value.date() if isinstance(value, datetime) else value


def _load_sheet(content: bytes):
    return openpyxl.load_workbook(BytesIO(content)).active


def _fill(cell):
    """The cell's solid fill RGB, or None when it has no visible fill."""
    return cell.fill.fgColor.rgb if cell.fill and cell.fill.fill_type else None


# --- column map (29 cols, A..AC; ACHIEVEMENT % = E, DIFFERENCE % = F) --------
# Six units per group: TAGS DOCS BOM SPARES PAGES RECORDS.
# DAY PART (C) sits directly after DATE; REMARKS (H) sits between SUB ACTIVITY
# and PROJECT; PROJECT (I) carries the TOTAL marker.
EMP, DATE_C, DAY_PART, ACTIVITY, ACH, DIFF, SUB, REMARKS, PROJECT = (
    1, 2, 3, 4, 5, 6, 7, 8, 9
)
TGT_TAGS, TGT_DOCS, TGT_BOM, TGT_SPARES, TGT_PAGES, TGT_RECORDS = 10, 11, 12, 13, 14, 15
ACT_TAGS, ACT_DOCS, ACT_BOM, ACT_SPARES, ACT_PAGES, ACT_RECORDS = 16, 17, 18, 19, 20, 21
PEN_TAGS, PEN_DOCS, PEN_BOM, PEN_SPARES, PEN_PAGES, PEN_RECORDS = 22, 23, 24, 25, 26, 27
CYC_START, CYC_END = 28, 29
LAST_COL = 29

# The only two body colours, and they land on the DIFFERENCE % cell alone.
GREEN = "FFC6EFCE"   # achievement > 100%
RED = "FFFFC7CE"     # achievement < 95%
HEADER_YELLOW = "FFFFFF00"

# Exact widths from the reference workbook, plus the new DAY PART column.
# Widths travel with the SEMANTIC field, not the old column letter: SUB ACTIVITY
# keeps 118.140625 (now G) and PROJECT keeps 86.0 (now I).
EXPECTED_WIDTHS = {
    "A": 26.0,           # EMP CODE & NAME
    "B": 12.0,           # DATE
    "C": 18.0,           # DAY PART — fits "HALF DAY (LEGACY)"
    "D": 22.0,           # ACTIVITY
    "E": 18.85546875,    # ACHIEVEMENT %
    "F": 15.0,           # DIFFERENCE %
    "G": 118.140625,     # SUB ACTIVITY
    "H": 50.0,           # REMARKS
    "I": 86.0,           # PROJECT CODE & TITLE
    # BENCHMARK TARGET J:O — the leading column carries the merged group label.
    "J": 21.42578125, "K": 12.0, "L": 12.0, "M": 12.0, "N": 12.0, "O": 12.0,
    # ACTUAL COMPLETED P:U
    "P": 16.85546875, "Q": 12.0, "R": 12.0, "S": 12.0, "T": 12.0, "U": 12.0,
    # PENDING BENCHMARK V:AA
    "V": 17.7109375, "W": 12.0, "X": 12.0, "Y": 12.0, "Z": 12.0, "AA": 12.0,
    "AB": 13.0, "AC": 13.0,
}
EXPECTED_MERGES = {"J1:O1", "P1:U1", "V1:AA1"}


def _detail_row(ws, label, sub, on_date):
    """Row index of one employee/sub-activity/date DETAIL row (total rows carry
    no DATE, so they never match)."""
    for r in range(3, ws.max_row + 1):
        if (
            ws.cell(r, EMP).value == label
            and ws.cell(r, SUB).value == sub
            and _cell_date(ws.cell(r, DATE_C).value) == on_date
        ):
            return r
    raise AssertionError(f"detail row not found: {label} / {sub} / {on_date}")


def _sub_total_rows(ws, label, sub):
    """Every TOTAL row for an employee+sub-activity (a total row carries "TOTAL"
    in the PROJECT column)."""
    return [
        r for r in range(3, ws.max_row + 1)
        if ws.cell(r, EMP).value == label
        and ws.cell(r, SUB).value == sub
        and ws.cell(r, PROJECT).value == "TOTAL"
    ]


def _sub_total_row(ws, label, sub):
    rows = _sub_total_rows(ws, label, sub)
    assert rows, f"sub-activity total not found: {label} / {sub}"
    return rows[0]


def _all_total_rows(ws):
    """Every TOTAL row in the sheet, by PROJECT == 'TOTAL'."""
    return [r for r in range(3, ws.max_row + 1) if ws.cell(r, PROJECT).value == "TOTAL"]


def _assert_only_diff_cell_shaded(ws, row):
    """No cell on `row` carries a fill except (possibly) DIFFERENCE %."""
    shaded = [
        ws.cell(row, c).coordinate
        for c in range(1, LAST_COL + 1)
        if c != DIFF and _fill(ws.cell(row, c)) is not None
    ]
    assert shaded == [], f"row {row} must only ever shade its E cell; shaded: {shaded}"


# --- cycle bounds -----------------------------------------------------------

def test_week_bounds_are_friday_to_thursday():
    friday = date(2026, 7, 3)  # a Friday
    thursday = date(2026, 7, 9)
    for offset in range(7):
        assert compute_week_bounds(friday + timedelta(days=offset)) == (friday, thursday)
    assert compute_week_bounds(date(2026, 7, 10)) == (date(2026, 7, 10), date(2026, 7, 16))


# --- endpoint guards --------------------------------------------------------

def test_export_requires_project_manager(client, setup_author):
    a = setup_author()
    assert client.get(EXPORT_URL, headers=a["header"]).status_code == 403


def test_export_rejects_unknown_cycle(client, activity_admin):
    assert client.get(f"{EXPORT_URL}?cycle=nope", headers=activity_admin).status_code == 422


# --- layout: exact 29-column order, two-level header -------------------------

def test_default_export_is_previous_cycle_with_reference_layout(client, setup_author, activity_admin):
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=250, name="FMTL")
    cycle_start, cycle_end = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 200)  # pending 50
    # A current-cycle shortfall must NOT leak into the default (previous) export.
    _submit(client, a["header"], a["project"].id, sub["id"], TODAY_D, 10)

    res = client.get(EXPORT_URL, headers=activity_admin)
    assert res.status_code == 200
    assert res.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert f"BENCHMARK REPORT {date_range_label(cycle_start, cycle_end)}.xlsx" in (
        res.headers["content-disposition"]
    )

    ws = _load_sheet(res.content)
    assert ws.title == PENDING_SHEET_NAME

    # Row 2 = the real header row, in the exact required order A..AC.
    assert [ws.cell(2, c).value for c in range(1, LAST_COL + 1)] == [
        "EMP CODE & NAME", "DATE", "DAY PART", "ACTIVITY", "ACHIEVEMENT %",
        "DIFFERENCE %", "SUB ACTIVITY", "REMARKS", "PROJECT CODE & TITLE",
        "TAGS", "DOCS", "BOM", "SPARES", "PAGES", "RECORDS",
        "TAGS", "DOCS", "BOM", "SPARES", "PAGES", "RECORDS",
        "TAGS", "DOCS", "BOM", "SPARES", "PAGES", "RECORDS",
        "CYCLE START", "CYCLE END",
    ]
    # None of the withdrawn columns exist anywhere in the header — and the
    # renamed DAY REMARKS is gone for good.
    labels = {ws.cell(2, c).value for c in range(1, LAST_COL + 1)}
    assert labels.isdisjoint(
        {"ROW TYPE", "TARGET TOTAL", "ACTUAL TOTAL", "PENDING TOTAL",
         "EMPLOYEE TOTAL", "DAY REMARKS", "PERIOD REMARKS"}
    )
    # Row 1 holds ONLY the three merged group labels; A1:I1 and AB1:AC1 are blank.
    assert ws.cell(1, TGT_TAGS).value == "BENCHMARK TARGET"
    assert ws.cell(1, ACT_TAGS).value == "ACTUAL COMPLETED"
    assert ws.cell(1, PEN_TAGS).value == "PENDING BENCHMARK"
    for c in list(range(1, PROJECT + 1)) + [CYC_START, CYC_END]:
        assert ws.cell(1, c).value is None
    # Exactly 29 columns: nothing in col 30.
    assert ws.cell(2, LAST_COL + 1).value is None

    assert {str(m) for m in ws.merged_cells.ranges} == EXPECTED_MERGES
    assert ws.auto_filter.ref == "A2:AC4"

    # Detail row, then its sub-activity TOTAL. No employee grand total.
    d = _detail_row(ws, "E-1 - Test User", "FMTL", cycle_start)
    assert ws.cell(d, EMP).value == "E-1 - Test User"
    assert _cell_date(ws.cell(d, DATE_C).value) == cycle_start
    assert ws.cell(d, PROJECT).value == "P-1 - Test Project"
    assert ws.cell(d, ACTIVITY).value == "Activity for FMTL"
    assert ws.cell(d, SUB).value == "FMTL"
    assert ws.cell(d, ACH).value is None             # % only on the TOTAL row
    assert ws.cell(d, DIFF).value is None
    assert ws.cell(d, TGT_TAGS).value == 250
    assert ws.cell(d, TGT_DOCS).value is None
    assert ws.cell(d, ACT_TAGS).value == 200
    assert ws.cell(d, PEN_TAGS).value == 50
    assert _cell_date(ws.cell(d, CYC_START).value) == cycle_start
    assert _cell_date(ws.cell(d, CYC_END).value) == cycle_end

    t = _sub_total_row(ws, "E-1 - Test User", "FMTL")
    assert ws.cell(t, EMP).value == "E-1 - Test User"    # repeated (filterable)
    assert ws.cell(t, DATE_C).value is None              # DATE blank on totals
    assert ws.cell(t, PROJECT).value == "TOTAL"          # marker in PROJECT col
    assert ws.cell(t, ACTIVITY).value == "Activity for FMTL"
    assert ws.cell(t, SUB).value == "FMTL"
    assert ws.cell(t, TGT_TAGS).value == 250
    assert ws.cell(t, ACT_TAGS).value == 200
    assert ws.cell(t, PEN_TAGS).value == 50
    assert ws.cell(t, ACH).value == 0.8
    assert ws.cell(t, ACH).number_format == "0.00%"
    assert ws.cell(t, DIFF).value == pytest.approx(0.2)
    assert ws.cell(t, DIFF).number_format == "0.00%"
    assert _fill(ws.cell(t, DIFF)) == RED
    # No employee grand total row and no current-cycle rows leaked in.
    assert len(_all_total_rows(ws)) == 1
    assert ws.max_row == 4


def test_no_employee_grand_total_is_generated(client, setup_author, activity_admin):
    """One employee with two numeric sub-activities gets one TOTAL per
    sub-activity and NO combined employee total."""
    a = setup_author()
    _, fmtl = _make_sub_activity(client, activity_admin, benchmark_value=250, name="FMTL")
    _, mtl = _make_sub_activity(client, activity_admin, benchmark_value=100, name="MTL")
    cycle_start, _ = _prev_cycle()
    day2 = cycle_start + timedelta(days=1)
    _submit(client, a["header"], a["project"].id, fmtl["id"], cycle_start, 200)
    _submit(client, a["header"], a["project"].id, mtl["id"], day2, 80)

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    totals = _all_total_rows(ws)
    assert len(totals) == 2                                 # FMTL + MTL, nothing else
    # Every TOTAL row names its sub-activity (a grand total would leave it blank).
    assert all(ws.cell(r, SUB).value in ("FMTL", "MTL") for r in totals)


# --- export content ---------------------------------------------------------

def test_cycle_current_exports_the_active_cycle(client, setup_author, activity_admin):
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=120, name="FMTL")
    _submit(client, a["header"], a["project"].id, sub["id"], TODAY_D, 100)  # pending 20

    ws = _load_sheet(client.get(f"{EXPORT_URL}?cycle=current", headers=activity_admin).content)
    label = "E-1 - Test User"
    assert ws.cell(_detail_row(ws, label, "FMTL", TODAY_D), PEN_TAGS).value == 20
    total = _sub_total_row(ws, label, "FMTL")
    assert ws.cell(total, ACH).value == pytest.approx(100 / 120)   # 83.33%, uncapped-ready
    assert _fill(ws.cell(total, DIFF)) == RED


def test_employees_performance_cycle_param_switches_window(client, setup_author, activity_admin):
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=250, name="FMTL")
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 200)

    def rows(query: str = "") -> dict:
        res = client.get(f"/api/v1/benchmarks/employees-performance{query}", headers=activity_admin)
        assert res.status_code == 200
        return {r["employee_code"]: r for r in res.json()["items"]}

    current = rows()
    assert float(current["E-1"]["pending"]) == 0
    assert float(current["E-1"]["target"]) == 0

    prev = rows("?cycle=previous")
    assert float(prev["E-1"]["pending"]) == 50
    assert float(prev["E-1"]["target"]) == 250
    assert float(prev["E-1"]["actual"]) == 200

    assert client.get(
        "/api/v1/benchmarks/employees-performance?cycle=nope", headers=activity_admin
    ).status_code == 422


# --- numeric grouping / compensation ----------------------------------------

def test_fmtl_and_mtl_get_separate_sub_activity_totals(client, setup_author, activity_admin):
    a = setup_author()
    _, fmtl = _make_sub_activity(client, activity_admin, benchmark_value=250, name="FMTL")
    _, mtl = _make_sub_activity(client, activity_admin, benchmark_value=100, name="MTL")
    cycle_start, _ = _prev_cycle()
    day2 = cycle_start + timedelta(days=1)
    _submit(client, a["header"], a["project"].id, fmtl["id"], cycle_start, 200)  # 80%
    _submit(client, a["header"], a["project"].id, mtl["id"], day2, 80)           # 80%

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    f_total = _sub_total_row(ws, label, "FMTL")
    m_total = _sub_total_row(ws, label, "MTL")
    assert f_total != m_total
    assert ws.cell(f_total, TGT_TAGS).value == 250 and ws.cell(f_total, ACH).value == 0.8
    assert ws.cell(m_total, TGT_TAGS).value == 100 and ws.cell(m_total, ACH).value == 0.8
    assert f_total < m_total  # sub-activities sort ascending: FMTL before MTL


def test_rows_sort_by_activity_then_sub_activity(client, setup_author, activity_admin):
    """Output order is employee -> activity -> sub-activity -> date. Two
    sub-activities whose names sort opposite to their activities prove the
    activity is the outer key."""
    a = setup_author()
    # Activity "A ..." owns the later-sorting sub-activity name, and vice versa.
    act_a = client.post(
        "/api/v1/activity-master/activities", json={"name": "AAA ACTIVITY"}, headers=activity_admin
    ).json()
    act_z = client.post(
        "/api/v1/activity-master/activities", json={"name": "ZZZ ACTIVITY"}, headers=activity_admin
    ).json()

    def sub_of(activity, name):
        return client.post(
            f"/api/v1/activity-master/activities/{activity['id']}/sub-activities",
            json={
                "name": name, "benchmark_type": "NUMERIC",
                "benchmark_value": 100, "relevant_count_field": "tags",
            },
            headers=activity_admin,
        ).json()

    z_named = sub_of(act_a, "ZZZ SUB")   # AAA ACTIVITY / ZZZ SUB
    a_named = sub_of(act_z, "AAA SUB")   # ZZZ ACTIVITY / AAA SUB
    cycle_start, _ = _prev_cycle()
    day2 = cycle_start + timedelta(days=1)
    _submit(client, a["header"], a["project"].id, z_named["id"], cycle_start, 100)
    _submit(client, a["header"], a["project"].id, a_named["id"], day2, 100)

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    # AAA ACTIVITY's rows come first even though its sub-activity name sorts last.
    assert _sub_total_row(ws, label, "ZZZ SUB") < _sub_total_row(ws, label, "AAA SUB")
    assert ws.cell(3, ACTIVITY).value == "AAA ACTIVITY"


def test_two_dates_of_one_sub_activity_combine_into_one_subtotal(client, setup_author, activity_admin):
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=200, name="FMTL")
    cycle_start, _ = _prev_cycle()
    day3 = cycle_start + timedelta(days=2)
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 250)  # day1 250/200
    _submit(client, a["header"], a["project"].id, sub["id"], day3, 120)         # day3 120/200

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    assert len(_sub_total_rows(ws, label, "FMTL")) == 1
    total = _sub_total_row(ws, label, "FMTL")
    assert ws.cell(total, TGT_TAGS).value == 400        # 200 + 200
    assert ws.cell(total, ACT_TAGS).value == 370        # 250 + 120
    assert ws.cell(total, PEN_TAGS).value == 30         # max(0, 400 - 370), NOT 0 + 80
    assert ws.cell(total, ACH).value == pytest.approx(370 / 400)   # 92.50%
    assert ws.cell(_detail_row(ws, label, "FMTL", cycle_start), PEN_TAGS).value == 0
    assert ws.cell(_detail_row(ws, label, "FMTL", day3), PEN_TAGS).value == 80


def test_later_overachievement_offsets_earlier_shortage(client, setup_author, activity_admin):
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=250, name="FMTL")
    cycle_start, _ = _prev_cycle()
    monday = cycle_start + timedelta(days=3)
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 100)  # short 150
    _submit(client, a["header"], a["project"].id, sub["id"], monday, 400)       # surplus 150

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    assert ws.cell(_detail_row(ws, label, "FMTL", cycle_start), PEN_TAGS).value == 150
    assert ws.cell(_detail_row(ws, label, "FMTL", monday), PEN_TAGS).value == 0
    total = _sub_total_row(ws, label, "FMTL")
    assert ws.cell(total, TGT_TAGS).value == 500
    assert ws.cell(total, ACT_TAGS).value == 500
    assert ws.cell(total, PEN_TAGS).value == 0     # NOT 150 (sum of daily shortages)
    assert ws.cell(total, ACH).value == 1.0
    # Exactly 100% is "met", not "ahead" -> no shade at all.
    assert ws.cell(total, DIFF).value == 0.0
    assert _fill(ws.cell(total, DIFF)) is None


def test_earlier_overachievement_offsets_later_shortage(client, setup_author, activity_admin):
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=250, name="FMTL")
    cycle_start, _ = _prev_cycle()
    monday = cycle_start + timedelta(days=3)
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 400)  # surplus 150
    _submit(client, a["header"], a["project"].id, sub["id"], monday, 100)       # short 150

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    assert ws.cell(_detail_row(ws, label, "FMTL", cycle_start), PEN_TAGS).value == 0
    assert ws.cell(_detail_row(ws, label, "FMTL", monday), PEN_TAGS).value == 150
    total = _sub_total_row(ws, label, "FMTL")
    assert ws.cell(total, PEN_TAGS).value == 0
    assert ws.cell(total, ACH).value == 1.0
    assert _fill(ws.cell(total, DIFF)) is None


def test_fmtl_overachievement_does_not_offset_mtl_shortage(client, setup_author, activity_admin):
    """A surplus in FMTL must NOT pay down a shortfall in MTL, even though both
    count TAGS. Each sub-activity nets only itself."""
    a = setup_author()
    _, fmtl = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    _, mtl = _make_sub_activity(client, activity_admin, benchmark_value=100, name="MTL")
    cycle_start, _ = _prev_cycle()
    day2 = cycle_start + timedelta(days=1)
    _submit(client, a["header"], a["project"].id, fmtl["id"], cycle_start, 200)  # +100 FMTL
    _submit(client, a["header"], a["project"].id, mtl["id"], day2, 40)           # short 60 MTL

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    f_total = _sub_total_row(ws, label, "FMTL")
    m_total = _sub_total_row(ws, label, "MTL")
    assert ws.cell(f_total, PEN_TAGS).value == 0 and ws.cell(f_total, ACH).value == 2.0
    assert ws.cell(m_total, PEN_TAGS).value == 60 and ws.cell(m_total, ACH).value == 0.4


def test_tags_excess_does_not_offset_docs_shortage(client, setup_author, activity_admin):
    """Different benchmark units never compensate: a TAGS surplus must NOT offset
    a DOCS shortfall. Each lands in its own subtotal, netted per unit."""
    a = setup_author()
    _, tags_sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="TAGGED")
    _, docs_sub = _make_sub_activity(
        client, activity_admin, benchmark_value=100, name="DOCD", count_field="docs"
    )
    cycle_start, _ = _prev_cycle()
    day2 = cycle_start + timedelta(days=1)  # one work report per date, so split days
    _submit(client, a["header"], a["project"].id, tags_sub["id"], cycle_start, 150)  # +50 tags
    _submit(client, a["header"], a["project"].id, docs_sub["id"], day2, 40, count_field="docs")

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    tagged = _sub_total_row(ws, label, "TAGGED")
    docd = _sub_total_row(ws, label, "DOCD")
    assert ws.cell(tagged, PEN_TAGS).value == 0 and ws.cell(tagged, ACH).value == 1.5
    assert ws.cell(docd, PEN_DOCS).value == 60 and ws.cell(docd, ACH).value == 0.4


def test_compensation_is_scoped_per_employee(client, setup_author, activity_admin):
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    cycle_start, _ = _prev_cycle()
    over = setup_author(email="o@x.com", code="O-1", proj_code="PO", last_name="Over")
    _submit(client, over["header"], over["project"].id, sub["id"], cycle_start, 300)  # +200
    under = setup_author(email="u@x.com", code="U-1", proj_code="PU", last_name="Under")
    _submit(client, under["header"], under["project"].id, sub["id"], cycle_start, 40)  # short 60

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    u = _sub_total_row(ws, "U-1 - Test Under", "FMTL")
    assert ws.cell(u, PEN_TAGS).value == 60 and ws.cell(u, ACH).value == 0.4
    o = _sub_total_row(ws, "O-1 - Test Over", "FMTL")
    assert ws.cell(o, PEN_TAGS).value == 0 and ws.cell(o, ACH).value == 3.0


def test_percentage_uses_the_entire_cycle(client, setup_author, activity_admin):
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    cycle_start, _ = _prev_cycle()
    day2 = cycle_start + timedelta(days=1)
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 120)  # over
    _submit(client, a["header"], a["project"].id, sub["id"], day2, 60)          # short

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    total = _sub_total_row(ws, "E-1 - Test User", "FMTL")
    assert ws.cell(total, TGT_TAGS).value == 200
    assert ws.cell(total, ACT_TAGS).value == 180
    assert ws.cell(total, PEN_TAGS).value == 20
    assert ws.cell(total, ACH).value == 0.9


def test_achievement_uncapped_above_100_percent(client, setup_author, activity_admin):
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 175)  # 175%

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    total = _sub_total_row(ws, "E-1 - Test User", "FMTL")
    assert ws.cell(total, PEN_TAGS).value == 0
    assert ws.cell(total, ACH).value == 1.75           # uncapped
    assert ws.cell(total, DIFF).value == pytest.approx(0.75)   # 220% -> 120%, no cap
    assert _fill(ws.cell(total, DIFF)) == GREEN


def test_employee_with_no_pending_is_still_exported(client, setup_author, activity_admin):
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 100)  # exactly met

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    total = _sub_total_row(ws, "E-1 - Test User", "FMTL")
    assert ws.cell(total, PEN_TAGS).value == 0
    assert ws.cell(total, ACH).value == 1.0
    assert _fill(ws.cell(total, DIFF)) is None


# --- required acceptance cases: the three shade bands ------------------------

def test_acceptance_below_95_shades_difference_cell_red(client, setup_author, activity_admin):
    """Target 65 / actual 60 -> 92.31% achievement, 7.69% difference. The E cell
    alone is red; F and every other cell stay unfilled."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=65, name="FMTL")
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 60)

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    t = _sub_total_row(ws, "E-1 - Test User", "FMTL")
    assert ws.cell(t, ACH).value == pytest.approx(60 / 65)          # 92.31%
    assert ws.cell(t, DIFF).value == pytest.approx(1 - 60 / 65)     # 7.69%
    assert round(ws.cell(t, DIFF).value * 100, 2) == 7.69
    assert _fill(ws.cell(t, DIFF)) == RED
    assert _fill(ws.cell(t, ACH)) is None       # the decider is never shaded
    _assert_only_diff_cell_shaded(ws, t)


def test_acceptance_95_to_100_has_no_shade_anywhere(client, setup_author, activity_admin):
    """Target 100 / actual 97 -> 97.00%, 3.00% difference and NO fill at all."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 97)

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    t = _sub_total_row(ws, "E-1 - Test User", "FMTL")
    assert ws.cell(t, ACH).value == 0.97
    assert ws.cell(t, DIFF).value == pytest.approx(0.03)
    assert _fill(ws.cell(t, DIFF)) is None
    # The entire row carries no fill whatsoever.
    for c in range(1, LAST_COL + 1):
        assert _fill(ws.cell(t, c)) is None


def test_acceptance_above_100_shades_difference_cell_green(client, setup_author, activity_admin):
    """Target 400 / actual 500 -> 125.00% achievement, 25.00% difference. The G
    cell alone is green."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=200, name="FMTL")
    cycle_start, _ = _prev_cycle()
    day2 = cycle_start + timedelta(days=1)
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 250)
    _submit(client, a["header"], a["project"].id, sub["id"], day2, 250)   # 500 / 400

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    t = _sub_total_row(ws, "E-1 - Test User", "FMTL")
    assert ws.cell(t, TGT_TAGS).value == 400 and ws.cell(t, ACT_TAGS).value == 500
    assert ws.cell(t, ACH).value == 1.25
    assert ws.cell(t, DIFF).value == pytest.approx(0.25)
    assert _fill(ws.cell(t, DIFF)) == GREEN
    assert _fill(ws.cell(t, ACH)) is None
    _assert_only_diff_cell_shaded(ws, t)


def test_shade_boundaries_are_strict(client, setup_author, activity_admin):
    """94.99% red, 95.00% none, 100.00% none, 100.01% green — decided by F, worn
    by G."""
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=10000, name="FMTL")
    cycle_start, _ = _prev_cycle()
    cases = [
        ("R", 9499, RED),      # 94.99%
        ("N", 9500, None),     # 95.00% — exactly on the boundary, no shade
        ("M", 10000, None),    # 100.00% — met, no shade
        ("G", 10001, GREEN),   # 100.01%
    ]
    for code, actual, _expected in cases:
        emp = setup_author(
            email=f"{code}@x.com", code=f"{code}-1", proj_code=f"P{code}", last_name=code
        )
        _submit(client, emp["header"], emp["project"].id, sub["id"], cycle_start, actual)

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    for code, actual, expected in cases:
        t = _sub_total_row(ws, f"{code}-1 - Test {code}", "FMTL")
        assert ws.cell(t, ACH).value == pytest.approx(actual / 10000)
        assert _fill(ws.cell(t, DIFF)) == expected, f"{code}: {actual / 100:.2f}%"
        assert _fill(ws.cell(t, ACH)) is None
        _assert_only_diff_cell_shaded(ws, t)


def test_no_full_row_is_ever_coloured(client, setup_author, activity_admin):
    """Across a mixed sheet (red, green, unshaded and textual rows), the ONLY
    filled body cells in the whole workbook are DIFFERENCE % cells."""
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    cycle_start, _ = _prev_cycle()
    for code, actual in (("R", 80), ("G", 150), ("N", 97)):
        emp = setup_author(
            email=f"{code}@x.com", code=f"{code}-1", proj_code=f"P{code}", last_name=code
        )
        _submit(client, emp["header"], emp["project"].id, sub["id"], cycle_start, actual)

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    filled = {
        ws.cell(r, c).coordinate: _fill(ws.cell(r, c))
        for r in range(3, ws.max_row + 1)
        for c in range(1, LAST_COL + 1)
        if _fill(ws.cell(r, c)) is not None
    }
    assert all(coord.startswith("F") for coord in filled), filled
    assert set(filled.values()) <= {RED, GREEN}


# --- sub_activity_id grouping + filter integrity ----------------------------

def test_grouping_keys_on_sub_activity_id_not_name(client, setup_author, activity_admin):
    """Two DISTINCT sub-activities sharing a displayed name stay apart (grouped
    by id) -> two subtotal rows, not one merged total."""
    a = setup_author()
    _, s1 = _make_sub_activity(client, activity_admin, benchmark_value=100, name="SHARED")
    _, s2 = _make_sub_activity(client, activity_admin, benchmark_value=100, name="SHARED")
    assert s1["id"] != s2["id"]
    cycle_start, _ = _prev_cycle()
    day2 = cycle_start + timedelta(days=1)
    _submit(client, a["header"], a["project"].id, s1["id"], cycle_start, 100)
    _submit(client, a["header"], a["project"].id, s2["id"], day2, 40)

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    assert len(_sub_total_rows(ws, "E-1 - Test User", "SHARED")) == 2


def test_every_row_repeats_exact_employee_code_and_name(client, setup_author, activity_admin):
    a = setup_author(code="CDC002", first_name="RAMASRINIVASAMOORTHY", last_name="D")
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="MTL-ASSET PHOTO DATA POPULATION")
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 80)

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    expected = "CDC002 - RAMASRINIVASAMOORTHY D"
    detail = _detail_row(ws, expected, "MTL-ASSET PHOTO DATA POPULATION", cycle_start)
    total = _sub_total_row(ws, expected, "MTL-ASSET PHOTO DATA POPULATION")
    assert ws.cell(detail, EMP).value == expected
    assert ws.cell(total, EMP).value == expected


def test_filter_by_sub_activity_keeps_detail_and_subtotal(client, setup_author, activity_admin):
    """The TOTAL row repeats the exact sub-activity name (not a bare "TOTAL" in
    SUB ACTIVITY), so a sub-activity filter keeps both detail and its total."""
    a = setup_author(code="CDC002", first_name="RAMASRINIVASAMOORTHY", last_name="D")
    sub_name = "MTL-ASSET PHOTO DATA POPULATION"
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name=sub_name)
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 80)

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    kept = [r for r in range(3, ws.max_row + 1) if ws.cell(r, SUB).value == sub_name]
    assert len(kept) == 2  # the detail row and its total row
    projects = {ws.cell(r, PROJECT).value for r in kept}
    assert projects == {"P-1 - Test Project", "TOTAL"}


# --- textual task rows: detail only, no total, no %, no shade ----------------

def test_textual_task_rows_have_no_total_or_percentage_mahesvari(
    client, db, setup_author, activity_admin,
):
    """Exact required case: a purely textual task sub-activity shows ONLY its
    dated detail rows — no subtotal, no percentages, no shading, no numeric
    totals."""
    import uuid as uuid_mod

    from app.modules.work_reports.models import WorkReportTask

    a = setup_author(code="CDCON064", first_name="MAHESVARI", last_name="S")
    sub_name = "BOM IDB-ADDRRESSING SPIR DOC AS PER MDR/PUNCH LIST FOR NOT AVAILABLE SPIR"
    _, task = _make_task_sub(client, activity_admin, name=sub_name)
    cycle_start, _ = _prev_cycle()
    d1, d2 = cycle_start, cycle_start + timedelta(days=1)
    for d in (d1, d2):
        tid = _submit_task(client, a["header"], a["project"].id, task["id"], d)
        row = db.get(WorkReportTask, uuid_mod.UUID(tid))
        row.due_date, row.is_completed, row.completed_date = d, True, d
    db.commit()

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "CDCON064 - MAHESVARI S"

    # Exactly two rows for the sub-activity — both DETAIL, no total.
    sub_rows = [r for r in range(3, ws.max_row + 1) if ws.cell(r, SUB).value == sub_name]
    assert len(sub_rows) == 2
    assert _sub_total_rows(ws, label, sub_name) == []
    for r, d in zip(sub_rows, (d1, d2)):
        assert _cell_date(ws.cell(r, DATE_C).value) == d
        assert ws.cell(r, TGT_TAGS).value == "FINISH WITHIN A DAY"
        assert ws.cell(r, ACT_TAGS).value == "FINISHED"
        assert ws.cell(r, PEN_TAGS).value == "NO PENDING"
        assert ws.cell(r, ACH).value is None            # percentages blank
        assert ws.cell(r, DIFF).value is None
        assert _fill(ws.cell(r, DIFF)) is None          # and never shaded
        _assert_only_diff_cell_shaded(ws, r)
        # Textual rows keep normal detail-row styling.
        assert ws.cell(r, EMP).font.bold is False
        assert ws.cell(r, EMP).font.name == "Arial" and ws.cell(r, EMP).font.sz == 10
        assert ws.cell(r, EMP).border.top.style == "thin"
        # No numeric totals in any target/actual/pending unit cell.
        for c in (TGT_TAGS, ACT_TAGS, PEN_TAGS):
            assert not isinstance(ws.cell(r, c).value, (int, float))


def test_textual_status_variants_stay_textual_and_excluded(client, db, setup_author, activity_admin):
    """FINISH WITHIN 2 DAYS / NOT COMPLETED / N DAYS OVERDUE remain textual and
    never produce a subtotal or percentage."""
    import uuid as uuid_mod

    from app.modules.work_reports.models import WorkReportTask

    a = setup_author()
    _, task = _make_task_sub(client, activity_admin, name="OVERDUE-TASK", period=2)
    cycle_start, cycle_end = _prev_cycle()
    tid = _submit_task(client, a["header"], a["project"].id, task["id"], cycle_start)
    row = db.get(WorkReportTask, uuid_mod.UUID(tid))
    row.due_date = cycle_end - timedelta(days=3)   # 3 days overdue as of cycle end
    db.commit()

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    d = _detail_row(ws, label, "OVERDUE-TASK", cycle_start)
    assert ws.cell(d, TGT_TAGS).value == "FINISH WITHIN 2 DAYS"
    assert ws.cell(d, ACT_TAGS).value == "NOT COMPLETED"
    assert ws.cell(d, PEN_TAGS).value == "3 DAYS OVERDUE"
    assert ws.cell(d, ACH).value is None
    assert ws.cell(d, DIFF).value is None
    assert _sub_total_rows(ws, label, "OVERDUE-TASK") == []


def test_count_based_task_is_numeric_and_gets_a_subtotal(client, db, setup_author, activity_admin):
    """CASE A (count-based lumpsum with real numbers) is treated as numeric: its
    bare numbers feed a subtotal + %. A plain textual task alongside it gets
    none."""
    import uuid as uuid_mod

    from app.modules.activity_master.models import ActivityMaster
    from app.modules.work_reports.models import WorkReportTask

    a = setup_author()
    _, count_sub = _make_task_sub(client, activity_admin, name="COUNT-LUMP")
    _, text_sub = _make_task_sub(client, activity_admin, name="TEXT-TASK")
    master = db.get(ActivityMaster, uuid_mod.UUID(count_sub["id"]))
    master.benchmark_value, master.relevant_count_field = 1000, "tags"
    db.commit()

    cycle_start, cycle_end = _prev_cycle()
    payload = {
        "report_date": cycle_start.isoformat(),
        "tasks": [
            {"project_id": str(a["project"].id), "description": "c",
             "sub_activity_id": count_sub["id"], "tags_count": 500},
            {"project_id": str(a["project"].id), "description": "t",
             "sub_activity_id": text_sub["id"]},
        ],
    }
    body = client.post(BASE, headers=a["header"], json=payload).json()
    assert client.post(f"{BASE}/{body['id']}/submit", headers=a["header"]).status_code == 200
    due = {count_sub["id"]: cycle_end - timedelta(days=5), text_sub["id"]: cycle_end - timedelta(days=3)}
    for t in body["tasks"]:
        row = db.get(WorkReportTask, uuid_mod.UUID(t["id"]))
        row.due_date = due[str(row.sub_activity_id)]
    db.commit()

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    # Count-based lumpsum: text detail cells, numeric subtotal (500/1000 = 50%).
    dc = _detail_row(ws, label, "COUNT-LUMP", cycle_start)
    assert ws.cell(dc, TGT_TAGS).value == "1000 TAGS PER DAY"
    ct = _sub_total_row(ws, label, "COUNT-LUMP")
    assert ws.cell(ct, TGT_TAGS).value == 1000
    assert ws.cell(ct, ACT_TAGS).value == 500
    assert ws.cell(ct, ACH).value == 0.5
    assert ws.cell(ct, DIFF).value == pytest.approx(0.5)
    assert _fill(ws.cell(ct, DIFF)) == RED
    # Plain textual task: detail only, no subtotal.
    assert _sub_total_rows(ws, label, "TEXT-TASK") == []


def test_numeric_and_textual_mixed_employee(client, db, setup_author, activity_admin):
    """One employee with a NUMERIC sub-activity and a textual DONE task: the
    numeric one gets its subtotal, the textual one does not, and the completed
    task never moves the numeric %."""
    import uuid as uuid_mod

    from app.modules.work_reports.models import WorkReportTask

    a = setup_author()
    _, num = _make_sub_activity(client, activity_admin, benchmark_value=100, name="NUM")
    _, done = _make_task_sub(client, activity_admin, name="DONE")
    cycle_start, _ = _prev_cycle()
    day1, day2, day3 = cycle_start, cycle_start + timedelta(days=1), cycle_start + timedelta(days=2)
    _submit(client, a["header"], a["project"].id, num["id"], day1, 200)  # over
    _submit(client, a["header"], a["project"].id, num["id"], day2, 80)   # short -> included
    tid = _submit_task(client, a["header"], a["project"].id, done["id"], day3)
    task = db.get(WorkReportTask, uuid_mod.UUID(tid))
    task.due_date, task.is_completed, task.completed_date = day3, True, day3
    db.commit()

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    num_total = _sub_total_row(ws, label, "NUM")
    assert ws.cell(num_total, ACH).value == 1.4        # (200+80)/(100+100)
    assert _sub_total_rows(ws, label, "DONE") == []    # textual DONE: no total
    d = _detail_row(ws, label, "DONE", day3)
    assert ws.cell(d, TGT_TAGS).value == "FINISH WITHIN A DAY"
    assert ws.cell(d, ACT_TAGS).value == "FINISHED"


def test_no_numeric_target_does_not_divide_by_zero(client, db, setup_author, activity_admin):
    """A plain textual task has no numeric target — export succeeds (no /0) and
    emits no subtotal for it."""
    import uuid as uuid_mod

    from app.modules.work_reports.models import WorkReportTask

    a = setup_author()
    _, task = _make_task_sub(client, activity_admin, name="PLAIN")
    cycle_start, cycle_end = _prev_cycle()
    tid = _submit_task(client, a["header"], a["project"].id, task["id"], cycle_start)
    row = db.get(WorkReportTask, uuid_mod.UUID(tid))
    row.due_date = cycle_end - timedelta(days=1)
    db.commit()

    res = client.get(EXPORT_URL, headers=activity_admin)
    assert res.status_code == 200  # no ZeroDivisionError
    ws = _load_sheet(res.content)
    assert _sub_total_rows(ws, "E-1 - Test User", "PLAIN") == []


# --- Excel style regression (compare actual stored workbook properties) -------

def test_header_style_matches_reference(client, setup_author, activity_admin):
    """Yellow FFFFFF00 Arial 10 bold centred+wrapped with thin borders across the
    WHOLE A1:AB2 block, including the blank cells above A:H and AA:AB."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 80)

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)

    assert ws["A2"].fill.fill_type == "solid"
    assert ws["A2"].fill.fgColor.rgb == HEADER_YELLOW == _PB_HEADER_FILL.fgColor.rgb
    assert ws["A2"].font.name == "Arial" == _PB_HEADER_FONT.name
    assert ws["A2"].font.sz == 10
    assert ws["A2"].font.bold is True
    assert ws["A2"].alignment.horizontal == "center"
    assert ws["A2"].alignment.vertical == "center"

    # Every style-bearing header cell. Inside a merge only the anchor holds the
    # style — Excel paints it across the merged span — so the continuation cells
    # of H1:M1 / N1:S1 / T1:Y1 are skipped, exactly as the reference stores them.
    merged_continuations = {
        c for c in range(TGT_TAGS, PEN_RECORDS + 1)
        if c not in (TGT_TAGS, ACT_TAGS, PEN_TAGS)
    }
    for row in (1, 2):
        for c in range(1, LAST_COL + 1):
            if row == 1 and c in merged_continuations:
                continue
            cell = ws.cell(row, c)
            assert _fill(cell) == HEADER_YELLOW, cell.coordinate
            assert (cell.font.name, cell.font.sz, cell.font.bold) == ("Arial", 10, True)
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.vertical == "center"
            assert cell.alignment.wrap_text is True
            for side in ("top", "bottom", "left", "right"):
                assert getattr(cell.border, side).style == "thin", f"{cell.coordinate}.{side}"


def test_worksheet_configuration_matches_reference(client, setup_author, activity_admin):
    """Sheet name, merges, freeze panes, filter range, row heights, widths."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 80)

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)

    assert ws.title == "Pending Benchmark" == PENDING_SHEET_NAME
    assert {str(m) for m in ws.merged_cells.ranges} == EXPECTED_MERGES
    assert ws.freeze_panes == "A3"
    assert ws.auto_filter.ref.startswith("A2:")
    assert ws.auto_filter.ref == f"A2:AC{ws.max_row}"
    assert ws.row_dimensions[1].height == 15.0
    assert ws.row_dimensions[2].height == 25.5
    assert ws.sheet_format.defaultRowHeight == 15.0
    for col, width in EXPECTED_WIDTHS.items():
        assert ws.column_dimensions[col].width == width, col


def test_detail_and_total_row_style_matches_reference(client, setup_author, activity_admin):
    """Detail rows: Arial 10 regular, no fill, top-aligned, thin borders, ISO
    dates. Total rows: identical but bold, with 0.00% on F and G."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 80)  # 80% -> red

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    d = _detail_row(ws, label, "FMTL", cycle_start)
    t = _sub_total_row(ws, label, "FMTL")

    for row, bold, font in ((d, False, _DATA_FONT), (t, True, _PB_TOTAL_FONT)):
        for c in range(1, LAST_COL + 1):
            cell = ws.cell(row, c)
            assert cell.font.name == font.name == "Arial", cell.coordinate
            assert cell.font.sz == 10
            assert cell.font.bold is bold, cell.coordinate
            assert cell.alignment.vertical == "top"
            # A:H and AA:AB left, the three unit groups (I:Z) centered.
            expected_h = "center" if TGT_TAGS <= c <= PEN_RECORDS else "left"
            assert cell.alignment.horizontal == expected_h, cell.coordinate
            for side in ("top", "bottom", "left", "right"):
                assert getattr(cell.border, side).style == _BORDER.top.style == "thin"

    for c in (DATE_C, CYC_START, CYC_END):
        assert ws.cell(d, c).number_format == "yyyy-mm-dd"
        assert ws.cell(t, c).number_format == "yyyy-mm-dd"
    assert ws.cell(t, ACH).number_format == "0.00%"
    assert ws.cell(t, DIFF).number_format == "0.00%"

    # Detail row is entirely unfilled; the total row shades only its E cell.
    for c in range(1, LAST_COL + 1):
        assert _fill(ws.cell(d, c)) is None, ws.cell(d, c).coordinate
    assert _fill(ws.cell(t, DIFF)) == RED
    _assert_only_diff_cell_shaded(ws, t)


# --- PAGES / RECORDS: the two units added alongside TAGS/DOCS/BOM/SPARES -----
#
# These pin the Phase 4 widening (4 units -> 6). The failure they guard against
# is a page or record count landing in the DOCS columns: three DOC IDB
# sub-activities were migrated DOCS -> RECORDS by migration 0058, and a
# regression there would silently mix records into a documents column.

def _make_daily_sub(client, admin_header, *, benchmark_value, name, count_field):
    """NUMERIC_DAILY sub-activity measured in `count_field` (the mode the three
    DOC IDB activities use after migration 0058)."""
    a = client.post(
        "/api/v1/activity-master/activities", json={"name": f"Activity for {name}"},
        headers=admin_header,
    ).json()
    sub = client.post(
        f"/api/v1/activity-master/activities/{a['id']}/sub-activities",
        json={
            "name": name, "benchmark_type": "NUMERIC_DAILY",
            "benchmark_value": benchmark_value, "relevant_count_field": count_field,
        },
        headers=admin_header,
    ).json()
    return a, sub


def _make_task_qty_sub(client, admin_header, *, benchmark_value, name, count_field, period=1):
    """TASK_WITH_QUANTITY sub-activity (the mode the three MTL activities use):
    carries a quantity AND a completion deadline."""
    a = client.post(
        "/api/v1/activity-master/activities", json={"name": f"Activity for {name}"},
        headers=admin_header,
    ).json()
    sub = client.post(
        f"/api/v1/activity-master/activities/{a['id']}/sub-activities",
        json={
            "name": name, "benchmark_type": "TASK_WITH_QUANTITY",
            "benchmark_value": benchmark_value, "relevant_count_field": count_field,
            "benchmark_period_days": period,
        },
        headers=admin_header,
    ).json()
    return a, sub


def _unit_cells(ws, row):
    """Every unit cell on `row` that carries a value, keyed "GROUP UNIT". Lets a
    test assert where a number landed AND that it landed nowhere else."""
    groups = (("TGT", TGT_TAGS), ("ACT", ACT_TAGS), ("PEN", PEN_TAGS))
    units = ("TAGS", "DOCS", "BOM", "SPARES", "PAGES", "RECORDS")
    return {
        f"{g} {u}": ws.cell(row, start + i).value
        for g, start in groups
        for i, u in enumerate(units)
        if ws.cell(row, start + i).value is not None
    }


def test_pages_below_target_is_red(client, setup_author, activity_admin):
    """Target 500 PAGES, actual 400 -> pending 100, achievement 80%,
    difference 20%, G red."""
    a = setup_author()
    _, sub = _make_daily_sub(
        client, activity_admin, benchmark_value=500, name="MTL-PAGES", count_field="pages"
    )
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 400, count_field="pages")

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    t = _sub_total_row(ws, "E-1 - Test User", "MTL-PAGES")
    assert ws.cell(t, TGT_PAGES).value == 500
    assert ws.cell(t, ACT_PAGES).value == 400
    assert ws.cell(t, PEN_PAGES).value == 100
    assert ws.cell(t, ACH).value == 0.8
    assert ws.cell(t, DIFF).value == pytest.approx(0.2)
    assert ws.cell(t, ACH).number_format == "0.00%"
    assert ws.cell(t, DIFF).number_format == "0.00%"
    assert _fill(ws.cell(t, DIFF)) == RED
    _assert_only_diff_cell_shaded(ws, t)


def test_pages_above_target_is_green_and_uncapped(client, setup_author, activity_admin):
    """Target 500 PAGES, actual 620 -> pending 0, achievement 124% (NOT capped
    at 100), difference 24%, G green."""
    a = setup_author()
    _, sub = _make_daily_sub(
        client, activity_admin, benchmark_value=500, name="MTL-PAGES", count_field="pages"
    )
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 620, count_field="pages")

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    t = _sub_total_row(ws, "E-1 - Test User", "MTL-PAGES")
    assert ws.cell(t, TGT_PAGES).value == 500
    assert ws.cell(t, ACT_PAGES).value == 620
    assert ws.cell(t, PEN_PAGES).value == 0
    assert ws.cell(t, ACH).value == pytest.approx(1.24)
    assert ws.cell(t, DIFF).value == pytest.approx(0.24)
    assert _fill(ws.cell(t, DIFF)) == GREEN
    _assert_only_diff_cell_shaded(ws, t)


def test_records_below_target_is_red(client, setup_author, activity_admin):
    """Target 1000 RECORDS, actual 850 -> pending 150, achievement 85%,
    difference 15%, G red."""
    a = setup_author()
    _, sub = _make_daily_sub(
        client, activity_admin, benchmark_value=1000, name="DOC IDB-QC", count_field="records"
    )
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 850, count_field="records")

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    t = _sub_total_row(ws, "E-1 - Test User", "DOC IDB-QC")
    assert ws.cell(t, TGT_RECORDS).value == 1000
    assert ws.cell(t, ACT_RECORDS).value == 850
    assert ws.cell(t, PEN_RECORDS).value == 150
    assert ws.cell(t, ACH).value == 0.85
    assert ws.cell(t, DIFF).value == pytest.approx(0.15)
    assert _fill(ws.cell(t, DIFF)) == RED
    _assert_only_diff_cell_shaded(ws, t)


def test_pages_values_land_only_in_pages_columns(client, setup_author, activity_admin):
    """A PAGES benchmark must touch the PAGES columns and NOTHING else - in
    particular not DOCS."""
    a = setup_author()
    _, sub = _make_daily_sub(
        client, activity_admin, benchmark_value=500, name="MTL-PAGES", count_field="pages"
    )
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 400, count_field="pages")

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    d = _detail_row(ws, label, "MTL-PAGES", cycle_start)
    t = _sub_total_row(ws, label, "MTL-PAGES")
    assert _unit_cells(ws, d) == {"TGT PAGES": 500, "ACT PAGES": 400, "PEN PAGES": 100}
    assert _unit_cells(ws, t) == {"TGT PAGES": 500, "ACT PAGES": 400, "PEN PAGES": 100}


def test_records_values_land_only_in_records_columns_never_docs(
    client, setup_author, activity_admin
):
    """The DOCS -> RECORDS regression guard: a migrated DOC IDB sub-activity's
    numbers must appear under RECORDS and leave every DOCS column empty."""
    a = setup_author()
    _, sub = _make_daily_sub(
        client, activity_admin, benchmark_value=1000, name="DOC IDB-REWORK", count_field="records"
    )
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 850, count_field="records")

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    d = _detail_row(ws, label, "DOC IDB-REWORK", cycle_start)
    t = _sub_total_row(ws, label, "DOC IDB-REWORK")
    assert _unit_cells(ws, d) == {"TGT RECORDS": 1000, "ACT RECORDS": 850, "PEN RECORDS": 150}
    assert _unit_cells(ws, t) == {"TGT RECORDS": 1000, "ACT RECORDS": 850, "PEN RECORDS": 150}
    for row in (d, t):
        for col in (TGT_DOCS, ACT_DOCS, PEN_DOCS):
            assert ws.cell(row, col).value is None, f"records leaked into DOCS at {col}"


def test_unrelated_docs_activity_still_lands_under_docs(client, setup_author, activity_admin):
    """Widening to six units must not disturb a genuine DOCS activity: it stays
    in the DOCS columns and out of the RECORDS ones."""
    a = setup_author()
    _, sub = _make_sub_activity(
        client, activity_admin, benchmark_value=200, name="DOC-REAL", count_field="docs"
    )
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 150, count_field="docs")

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    t = _sub_total_row(ws, label, "DOC-REAL")
    assert _unit_cells(ws, t) == {"TGT DOCS": 200, "ACT DOCS": 150, "PEN DOCS": 50}
    assert ws.cell(t, ACH).value == 0.75


def test_pages_and_records_coexist_without_compensating(client, setup_author, activity_admin):
    """A PAGES surplus must not offset a RECORDS shortfall: separate
    sub-activities, separate subtotals, separate percentages."""
    a = setup_author()
    _, pages_sub = _make_daily_sub(
        client, activity_admin, benchmark_value=500, name="MTL-PAGES", count_field="pages"
    )
    _, rec_sub = _make_daily_sub(
        client, activity_admin, benchmark_value=1000, name="DOC IDB-QC", count_field="records"
    )
    cycle_start, _ = _prev_cycle()
    day2 = cycle_start + timedelta(days=1)  # one report per date
    _submit(client, a["header"], a["project"].id, pages_sub["id"], cycle_start, 620, count_field="pages")
    _submit(client, a["header"], a["project"].id, rec_sub["id"], day2, 850, count_field="records")

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    p = _sub_total_row(ws, label, "MTL-PAGES")
    r = _sub_total_row(ws, label, "DOC IDB-QC")
    assert ws.cell(p, PEN_PAGES).value == 0 and ws.cell(p, ACH).value == pytest.approx(1.24)
    assert ws.cell(r, PEN_RECORDS).value == 150 and ws.cell(r, ACH).value == 0.85
    assert _fill(ws.cell(p, DIFF)) == GREEN
    assert _fill(ws.cell(r, DIFF)) == RED


def test_task_with_quantity_pages_counted_exactly_once(client, setup_author, activity_admin):
    """TASK_WITH_QUANTITY carries a quantity AND a deadline, so it could be
    picked up by both the daily-quantity ledger and the task lumpsum query. It
    must appear through the lumpsum side ONLY: one subtotal, counted once."""
    a = setup_author()
    _, sub = _make_task_qty_sub(
        client, activity_admin, benchmark_value=500, name="MTL-TASK-PAGES", count_field="pages"
    )
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 400, count_field="pages")

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    # Exactly one subtotal, and exactly one detail row for the sub-activity.
    assert len(_sub_total_rows(ws, label, "MTL-TASK-PAGES")) == 1
    details = [
        r for r in range(3, ws.max_row + 1)
        if ws.cell(r, SUB).value == "MTL-TASK-PAGES"
        and ws.cell(r, PROJECT).value != "TOTAL"
    ]
    assert len(details) == 1
    # Counted once: the target is 500, not 1000.
    t = _sub_total_row(ws, label, "MTL-TASK-PAGES")
    assert ws.cell(t, TGT_PAGES).value == 500
    assert ws.cell(t, ACT_PAGES).value == 400
    assert ws.cell(t, PEN_PAGES).value == 100
    assert ws.cell(t, ACH).value == 0.8
    assert _fill(ws.cell(t, DIFF)) == RED


def test_textual_task_beside_pages_gets_no_subtotal_and_blank_f_g(
    client, setup_author, activity_admin
):
    """A textual "FINISH WITHIN A DAY" task keeps its detail rows only - no
    subtotal, blank ACHIEVEMENT %/DIFFERENCE %, no shading - while a PAGES
    activity for the same employee still totals normally."""
    a = setup_author()
    _, pages_sub = _make_daily_sub(
        client, activity_admin, benchmark_value=500, name="MTL-PAGES", count_field="pages"
    )
    _, text_sub = _make_task_sub(client, activity_admin, name="TEXT-TASK")
    cycle_start, _ = _prev_cycle()
    payload = {
        "report_date": cycle_start.isoformat(),
        "tasks": [
            {"project_id": str(a["project"].id), "description": "p",
             "sub_activity_id": pages_sub["id"], "pages_count": 400},
            {"project_id": str(a["project"].id), "description": "t",
             "sub_activity_id": text_sub["id"]},
        ],
    }
    body = client.post(BASE, headers=a["header"], json=payload).json()
    assert client.post(f"{BASE}/{body['id']}/submit", headers=a["header"]).status_code == 200

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    assert _sub_total_rows(ws, label, "TEXT-TASK") == []
    d = _detail_row(ws, label, "TEXT-TASK", cycle_start)
    assert ws.cell(d, TGT_TAGS).value == "FINISH WITHIN A DAY"
    assert ws.cell(d, ACH).value is None
    assert ws.cell(d, DIFF).value is None
    for c in range(1, LAST_COL + 1):
        assert _fill(ws.cell(d, c)) is None, ws.cell(d, c).coordinate
    # The PAGES activity is unaffected by the textual row sharing the report.
    t = _sub_total_row(ws, label, "MTL-PAGES")
    assert ws.cell(t, ACH).value == 0.8


def test_only_difference_column_is_ever_filled_across_six_units(client, setup_author, activity_admin):
    """Sheet-wide colour audit with PAGES and RECORDS in play: outside the yellow
    header, the ONLY filled cells in the whole workbook are DIFFERENCE % cells (column F),
    and they only ever carry the two approved colours."""
    a = setup_author()
    _, pages_sub = _make_daily_sub(
        client, activity_admin, benchmark_value=500, name="MTL-PAGES", count_field="pages"
    )
    _, rec_sub = _make_daily_sub(
        client, activity_admin, benchmark_value=1000, name="DOC IDB-QC", count_field="records"
    )
    cycle_start, _ = _prev_cycle()
    day2 = cycle_start + timedelta(days=1)
    _submit(client, a["header"], a["project"].id, pages_sub["id"], cycle_start, 620, count_field="pages")
    _submit(client, a["header"], a["project"].id, rec_sub["id"], day2, 850, count_field="records")

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    filled = {
        ws.cell(r, c).coordinate: _fill(ws.cell(r, c))
        for r in range(3, ws.max_row + 1)
        for c in range(1, LAST_COL + 1)
        if _fill(ws.cell(r, c)) is not None
    }
    assert all(coord.startswith("F") for coord in filled), filled
    assert set(filled.values()) <= {RED, GREEN}
    # The header keeps its yellow across all 29 columns. Row 1's merged
    # continuation cells hold no style of their own (Excel paints the anchor's
    # fill across the span), so only the anchors are checked there.
    anchors = (TGT_TAGS, ACT_TAGS, PEN_TAGS)
    for c in range(1, LAST_COL + 1):
        if not (TGT_TAGS <= c <= PEN_RECORDS) or c in anchors:
            assert _fill(ws.cell(1, c)) == HEADER_YELLOW, ws.cell(1, c).coordinate
        assert _fill(ws.cell(2, c)) == HEADER_YELLOW, ws.cell(2, c).coordinate


# --- Phase 4.1: REMARKS column + current-plus-three-cycle selection ----------
#
# REMARKS (H) carries WorkReport.remarks for full-day rows - the employee's own
# remark for that report DATE (split-day rows carry their period remarks; see
# test_pending_export_day_parts.py). It is NOT ActivityMaster.benchmark_remarks
# (guidance TO the employee, e.g. "500 REQUIRED PAGES/DAY"), which must never
# appear here.

def _submit_with_remarks(client, header, project_id, sub_id, report_date, qty,
                         remarks, count_field="tags"):
    """Submit one day's report carrying a day-level remark."""
    payload = {
        "report_date": report_date.isoformat(),
        "remarks": remarks,
        "tasks": [{
            "project_id": str(project_id), "description": "work",
            "sub_activity_id": sub_id, f"{count_field}_count": qty,
        }],
    }
    created = client.post(BASE, headers=header, json=payload)
    assert created.status_code == 201, created.text
    res = client.post(f"{BASE}/{created.json()['id']}/submit", headers=header)
    assert res.status_code == 200, res.text


def test_remarks_is_column_h_and_project_is_column_i(client, setup_author, activity_admin):
    """Header positions the reorder must hold: DAY PART C, ACTIVITY D,
    ACHIEVEMENT E, DIFFERENCE F, SUB ACTIVITY G, REMARKS H, PROJECT I,
    groups from J."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    cycle_start, _ = _prev_cycle()
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 80)

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    assert ws.cell(2, 3).value == "DAY PART"
    assert ws.cell(2, 4).value == "ACTIVITY"
    assert ws.cell(2, 5).value == "ACHIEVEMENT %"
    assert ws.cell(2, 6).value == "DIFFERENCE %"
    assert ws.cell(2, 7).value == "SUB ACTIVITY"
    assert ws.cell(2, 8).value == "REMARKS"
    assert ws.cell(2, 9).value == "PROJECT CODE & TITLE"
    # The unit groups start at J (column 10) and the sheet ends at AC (29).
    assert ws.cell(1, 10).value == "BENCHMARK TARGET"
    assert get_column_letter(ws.max_column) == "AC"
    assert {str(m) for m in ws.merged_cells.ranges} == {"J1:O1", "P1:U1", "V1:AA1"}
    assert ws.freeze_panes == "A3"


def test_day_remark_repeats_on_every_detail_row_of_that_day(
    client, setup_author, activity_admin
):
    """One report, two sub-activities: both detail rows carry the same day
    remark, so filtering to either row alone still shows it."""
    a = setup_author()
    _, s1 = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL-AUDIT")
    _, s2 = _make_sub_activity(
        client, activity_admin, benchmark_value=200, name="MTL-DOC", count_field="docs"
    )
    cycle_start, _ = _prev_cycle()
    payload = {
        "report_date": cycle_start.isoformat(),
        "remarks": "Checked audit queries",
        "tasks": [
            {"project_id": str(a["project"].id), "description": "x",
             "sub_activity_id": s1["id"], "tags_count": 80},
            {"project_id": str(a["project"].id), "description": "y",
             "sub_activity_id": s2["id"], "docs_count": 150},
        ],
    }
    body = client.post(BASE, headers=a["header"], json=payload).json()
    assert client.post(f"{BASE}/{body['id']}/submit", headers=a["header"]).status_code == 200

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    for sub_name in ("FMTL-AUDIT", "MTL-DOC"):
        d = _detail_row(ws, label, sub_name, cycle_start)
        assert ws.cell(d, REMARKS).value == "Checked audit queries", sub_name
    # No merged cells were introduced to achieve the repeat.
    assert {str(m) for m in ws.merged_cells.ranges} == EXPECTED_MERGES


def test_different_days_show_their_own_remarks(client, setup_author, activity_admin):
    """Each report date carries its own remark; one day's text never bleeds
    into another's row."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    cycle_start, _ = _prev_cycle()
    day2 = cycle_start + timedelta(days=1)
    _submit_with_remarks(client, a["header"], a["project"].id, sub["id"],
                         cycle_start, 80, "Day one remark")
    _submit_with_remarks(client, a["header"], a["project"].id, sub["id"],
                         day2, 90, "Day two remark")

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    assert ws.cell(_detail_row(ws, label, "FMTL", cycle_start), REMARKS).value == "Day one remark"
    assert ws.cell(_detail_row(ws, label, "FMTL", day2), REMARKS).value == "Day two remark"


def test_blank_and_whitespace_remarks_stay_blank(client, setup_author, activity_admin):
    """No remark, or a whitespace-only one, leaves the cell empty - never a
    stray space, and never filled in from somewhere else."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    cycle_start, _ = _prev_cycle()
    day2 = cycle_start + timedelta(days=1)
    _submit(client, a["header"], a["project"].id, sub["id"], cycle_start, 80)  # no remarks
    _submit_with_remarks(client, a["header"], a["project"].id, sub["id"], day2, 90, "   ")

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    label = "E-1 - Test User"
    assert ws.cell(_detail_row(ws, label, "FMTL", cycle_start), REMARKS).value is None
    assert ws.cell(_detail_row(ws, label, "FMTL", day2), REMARKS).value is None


def test_total_row_has_blank_remarks_and_total_in_column_i(
    client, setup_author, activity_admin
):
    """A TOTAL row spans the cycle, not one day, so REMARKS is blank and the
    TOTAL marker sits in PROJECT (column I) with the percentages in E and F."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    cycle_start, _ = _prev_cycle()
    _submit_with_remarks(client, a["header"], a["project"].id, sub["id"],
                         cycle_start, 80, "Some day remark")

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    t = _sub_total_row(ws, "E-1 - Test User", "FMTL")
    assert ws.cell(t, EMP).value == "E-1 - Test User"
    assert ws.cell(t, DATE_C).value is None          # B blank
    assert ws.cell(t, ACTIVITY).value == "Activity for FMTL"
    assert ws.cell(t, ACH).value == 0.8              # D
    assert ws.cell(t, DIFF).value == pytest.approx(0.2)   # E
    assert ws.cell(t, SUB).value == "FMTL"           # F
    assert ws.cell(t, REMARKS).value is None         # H blank on a total
    assert ws.cell(t, PROJECT).value == "TOTAL"      # I marker
    assert ws.cell(t, ACH).number_format == "0.00%"
    assert ws.cell(t, DIFF).number_format == "0.00%"
    assert _fill(ws.cell(t, DIFF)) == RED
    _assert_only_diff_cell_shaded(ws, t)


def test_day_remarks_never_carries_benchmark_remarks(client, db, setup_author, activity_admin):
    """Activity Master's benchmark_remarks is guidance TO the employee; it must
    never be substituted into the employee's DAY REMARKS column."""
    import uuid as uuid_mod

    from app.modules.activity_master.models import ActivityMaster

    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=500, name="MTL-PAGES",
                                count_field="pages")
    master = db.get(ActivityMaster, uuid_mod.UUID(sub["id"]))
    master.benchmark_remarks = "500 REQUIRED PAGES/DAY"
    db.commit()

    cycle_start, _ = _prev_cycle()
    _submit_with_remarks(client, a["header"], a["project"].id, sub["id"],
                         cycle_start, 400, "Waiting on vendor drawings", count_field="pages")

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    d = _detail_row(ws, "E-1 - Test User", "MTL-PAGES", cycle_start)
    assert ws.cell(d, REMARKS).value == "Waiting on vendor drawings"
    # The Activity Master guidance appears nowhere in the sheet.
    everywhere = {
        ws.cell(r, c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, LAST_COL + 1)
    }
    assert "500 REQUIRED PAGES/DAY" not in everywhere


def test_day_remarks_cell_wraps_and_is_top_aligned(client, setup_author, activity_admin):
    """Free-text remarks wrap inside column H rather than spilling across the
    unit columns."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    cycle_start, _ = _prev_cycle()
    _submit_with_remarks(client, a["header"], a["project"].id, sub["id"],
                         cycle_start, 80, "A fairly long day remark " * 5)

    ws = _load_sheet(client.get(EXPORT_URL, headers=activity_admin).content)
    d = _detail_row(ws, "E-1 - Test User", "FMTL", cycle_start)
    cell = ws.cell(d, REMARKS)
    assert cell.alignment.wrap_text is True
    assert cell.alignment.vertical == "top"
    assert ws.column_dimensions["H"].width == 50.0


# --- cycle selection: current plus the previous three ------------------------

def test_resolve_week_offset_accepts_offsets_and_legacy_aliases():
    """One resolver, one source of truth: integer offsets 0..3 plus the two
    legacy strings."""
    from app.modules.benchmarks.service import resolve_week_offset

    assert [resolve_week_offset(i) for i in range(4)] == [0, 1, 2, 3]
    assert resolve_week_offset("current") == 0
    assert resolve_week_offset("previous") == 1
    assert resolve_week_offset("2") == 2          # query strings arrive as text
    assert resolve_week_offset(None) == 0


@pytest.mark.parametrize("bad", [-1, 4, 99, "next", "", "1.5"])
def test_resolve_week_offset_rejects_unsupported_values(bad):
    """An unsupported selector is REJECTED, never silently swapped for another
    cycle - that would hand back a period nobody asked for."""
    from app.modules.benchmarks.service import resolve_week_offset

    with pytest.raises(ValueError):
        resolve_week_offset(bad)


@pytest.mark.parametrize("offset", [0, 1, 2, 3])
def test_cycle_window_walks_back_whole_friday_to_thursday_weeks(offset):
    """Offset N is exactly N weeks before the current cycle, and every cycle is
    a 7-day Friday..Thursday span."""
    from app.modules.benchmarks.service import _cycle_window

    current_start, _ = compute_week_bounds(TODAY_D)
    start, end = _cycle_window(offset, TODAY_D)
    assert start == current_start - timedelta(days=7 * offset)
    assert end == start + timedelta(days=6)
    assert (end - start).days + 1 == 7           # exactly seven calendar dates
    assert start.weekday() == 4 and end.weekday() == 3   # Friday .. Thursday


def test_legacy_cycle_strings_match_their_offsets():
    """current == offset 0 and previous == offset 1, so old clients keep
    working against the same windows."""
    from app.modules.benchmarks.service import _cycle_window

    assert _cycle_window("current", TODAY_D) == _cycle_window(0, TODAY_D)
    assert _cycle_window("previous", TODAY_D) == _cycle_window(1, TODAY_D)


def test_export_week_offset_selects_the_right_cycle_and_filename(
    client, setup_author, activity_admin
):
    """One submission per cycle for four cycles; each week_offset exports only
    its own cycle's row, and the filename names the SELECTED cycle."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    current_start, _ = compute_week_bounds(TODAY_D)

    # A distinct actual per cycle makes leakage obvious.
    actuals = {0: 10, 1: 20, 2: 30, 3: 40}
    for offset, qty in actuals.items():
        _submit(client, a["header"], a["project"].id, sub["id"],
                current_start - timedelta(days=7 * offset), qty)

    for offset, qty in actuals.items():
        res = client.get(f"{EXPORT_URL}?week_offset={offset}", headers=activity_admin)
        assert res.status_code == 200, res.text
        start = current_start - timedelta(days=7 * offset)
        end = start + timedelta(days=6)
        # Filename names the selected cycle, not today.
        assert f"BENCHMARK REPORT {date_range_label(start, end)}.xlsx" in (
            res.headers["content-disposition"]
        )
        ws = _load_sheet(res.content)
        # Exactly this cycle's dates, and only this cycle's actual.
        dates = {
            _cell_date(ws.cell(r, DATE_C).value)
            for r in range(3, ws.max_row + 1)
            if ws.cell(r, DATE_C).value is not None
        }
        assert dates == {start}, f"offset {offset} leaked dates {dates}"
        t = _sub_total_row(ws, "E-1 - Test User", "FMTL")
        assert ws.cell(t, ACT_TAGS).value == qty
        assert _cell_date(ws.cell(t, CYC_START).value) == start
        assert _cell_date(ws.cell(t, CYC_END).value) == end


def test_no_rows_leak_between_neighbouring_cycles(client, setup_author, activity_admin):
    """Adjacent cycles stay disjoint: offset 1 never shows offset 0's or
    offset 2's rows, and no pending value carries across."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    current_start, _ = compute_week_bounds(TODAY_D)
    for offset, qty in ((0, 10), (1, 60), (2, 30)):
        _submit(client, a["header"], a["project"].id, sub["id"],
                current_start - timedelta(days=7 * offset), qty)

    ws = _load_sheet(
        client.get(f"{EXPORT_URL}?week_offset=1", headers=activity_admin).content
    )
    t = _sub_total_row(ws, "E-1 - Test User", "FMTL")
    # Only the middle cycle's numbers: 60/100, pending 40 - not 10 or 30.
    assert ws.cell(t, ACT_TAGS).value == 60
    assert ws.cell(t, PEN_TAGS).value == 40
    assert ws.cell(t, ACH).value == 0.6


@pytest.mark.parametrize("bad", ["-1", "4", "next"])
def test_export_rejects_invalid_week_offset(client, activity_admin, bad):
    res = client.get(f"{EXPORT_URL}?week_offset={bad}", headers=activity_admin)
    assert res.status_code == 422, res.text


def test_export_legacy_cycle_param_still_works(client, setup_author, activity_admin):
    """?cycle=current|previous keeps selecting offsets 0 and 1."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    current_start, _ = compute_week_bounds(TODAY_D)
    _submit(client, a["header"], a["project"].id, sub["id"], current_start, 10)
    _submit(client, a["header"], a["project"].id, sub["id"],
            current_start - timedelta(days=7), 20)

    for cycle, expected_start, qty in (
        ("current", current_start, 10),
        ("previous", current_start - timedelta(days=7), 20),
    ):
        ws = _load_sheet(
            client.get(f"{EXPORT_URL}?cycle={cycle}", headers=activity_admin).content
        )
        t = _sub_total_row(ws, "E-1 - Test User", "FMTL")
        assert ws.cell(t, ACT_TAGS).value == qty
        assert _cell_date(ws.cell(t, CYC_START).value) == expected_start


def test_performance_table_and_export_agree_on_the_selected_cycle(
    client, setup_author, activity_admin
):
    """The selected offset drives BOTH the comparison table and the export -
    one resolver, so they cannot disagree."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=100, name="FMTL")
    current_start, _ = compute_week_bounds(TODAY_D)
    _submit(client, a["header"], a["project"].id, sub["id"], current_start, 10)
    _submit(client, a["header"], a["project"].id, sub["id"],
            current_start - timedelta(days=14), 90)

    for offset, expected_actual in ((0, 10), (2, 90)):
        table = client.get(
            f"/api/v1/benchmarks/employees-performance?week_offset={offset}",
            headers=activity_admin,
        )
        assert table.status_code == 200, table.text
        ws = _load_sheet(
            client.get(f"{EXPORT_URL}?week_offset={offset}", headers=activity_admin).content
        )
        t = _sub_total_row(ws, "E-1 - Test User", "FMTL")
        assert ws.cell(t, ACT_TAGS).value == expected_actual


@pytest.mark.parametrize("bad", ["-1", "4", "next"])
def test_performance_table_rejects_invalid_week_offset(client, activity_admin, bad):
    res = client.get(
        f"/api/v1/benchmarks/employees-performance?week_offset={bad}",
        headers=activity_admin,
    )
    assert res.status_code == 422, res.text


# --- table <-> export agreement across the four selectable cycles -----------
#
# The comparison table and the workbook are two different code paths over the
# same ledger. They must resolve the SAME Fri..Thu window for a given offset:
# a regression where one of them re-derived bounds from date.today() would show
# the current week's numbers under an older week's label.

PERF_URL = "/api/v1/benchmarks/employees-performance"


def _table_row(client, header, *, week_offset, employee_code):
    """One employee's row from the comparison table for the selected cycle."""
    res = client.get(
        f"{PERF_URL}?page=1&page_size=100&sort=name&order=asc&week_offset={week_offset}",
        headers=header,
    )
    assert res.status_code == 200, res.text
    for row in res.json()["items"]:
        if row["employee_code"] == employee_code:
            return row
    raise AssertionError(f"{employee_code} missing from the table (offset {week_offset})")


def test_table_and_export_agree_for_every_selectable_cycle(
    client, setup_author, activity_admin
):
    """Seed one employee with a different actual in each of the four cycles,
    then assert table actual == export actual == the value seeded for that
    cycle, for every offset."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=500, name="FMTL")
    current_start, _ = compute_week_bounds(TODAY_D)

    # offset -> actual seeded in that cycle
    seeded = {0: 100, 1: 200, 2: 300, 3: 400}
    for offset, qty in seeded.items():
        _submit(client, a["header"], a["project"].id, sub["id"],
                current_start - timedelta(days=7 * offset), qty)

    for offset, qty in seeded.items():
        start = current_start - timedelta(days=7 * offset)
        end = start + timedelta(days=6)

        # 1) the comparison table
        row = _table_row(client, activity_admin, week_offset=offset, employee_code="E-1")
        assert float(row["actual"]) == qty, (
            f"table actual for offset {offset} was {row['actual']}, expected {qty}"
        )
        assert float(row["target"]) == 500
        assert float(row["pending"]) == max(0, 500 - qty)

        # 2) the workbook, same offset
        ws = _load_sheet(
            client.get(f"{EXPORT_URL}?week_offset={offset}", headers=activity_admin).content
        )
        t = _sub_total_row(ws, "E-1 - Test User", "FMTL")
        assert ws.cell(t, ACT_TAGS).value == qty, (
            f"export actual for offset {offset} was {ws.cell(t, ACT_TAGS).value}, expected {qty}"
        )

        # 3) they agree with each other, and name the same window
        assert float(row["actual"]) == ws.cell(t, ACT_TAGS).value
        assert _cell_date(ws.cell(t, CYC_START).value) == start
        assert _cell_date(ws.cell(t, CYC_END).value) == end

        # 4) no neighbouring cycle leaked in: only this cycle's date appears
        dates = {
            _cell_date(ws.cell(r, DATE_C).value)
            for r in range(3, ws.max_row + 1)
            if ws.cell(r, DATE_C).value is not None
        }
        assert dates == {start}, f"offset {offset} leaked dates {dates}"


def test_each_cycle_returns_a_distinct_table_actual(client, setup_author, activity_admin):
    """The four offsets must be distinguishable end to end - if the table
    re-derived its own bounds they would all collapse to one value."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=500, name="FMTL")
    current_start, _ = compute_week_bounds(TODAY_D)
    seeded = {0: 100, 1: 200, 2: 300, 3: 400}
    for offset, qty in seeded.items():
        _submit(client, a["header"], a["project"].id, sub["id"],
                current_start - timedelta(days=7 * offset), qty)

    actuals = {
        offset: float(
            _table_row(client, activity_admin, week_offset=offset, employee_code="E-1")["actual"]
        )
        for offset in (0, 1, 2, 3)
    }
    assert actuals == {0: 100.0, 1: 200.0, 2: 300.0, 3: 400.0}
    assert len(set(actuals.values())) == 4, "offsets are not distinguishable"


def test_legacy_cycle_alias_matches_its_offset_on_the_table(
    client, setup_author, activity_admin
):
    """?cycle=current|previous must land on the same rows as offsets 0 and 1."""
    a = setup_author()
    _, sub = _make_sub_activity(client, activity_admin, benchmark_value=500, name="FMTL")
    current_start, _ = compute_week_bounds(TODAY_D)
    _submit(client, a["header"], a["project"].id, sub["id"], current_start, 100)
    _submit(client, a["header"], a["project"].id, sub["id"],
            current_start - timedelta(days=7), 200)

    for alias, offset in (("current", 0), ("previous", 1)):
        by_alias = client.get(
            f"{PERF_URL}?page=1&page_size=100&sort=name&order=asc&cycle={alias}",
            headers=activity_admin,
        )
        assert by_alias.status_code == 200, by_alias.text
        alias_row = next(
            r for r in by_alias.json()["items"] if r["employee_code"] == "E-1"
        )
        offset_row = _table_row(
            client, activity_admin, week_offset=offset, employee_code="E-1"
        )
        assert alias_row == offset_row
