"""PM Weekly Activity Report — PAGES and RECORDS reach the preview API and the
Excel export, in the exact column order the four legacy counts already use:

    Tags | Docs | BOM | Spares | Pages | Records   (Remarks last)

The service (build_activity_rows / build_activity_groups) already produced the
two values; the gap this proves closed is the ActivityCell schema surfacing them
and the export carrying two extra count columns. Nothing about the separate
28-column Benchmark export is touched here.

The sheet is one row per activity (see test_pm_activity_report_layout.py); this
file only pins where the six count columns sit within that row.
"""
from datetime import date

import openpyxl
import pytest

from app.modules.projects.models import ProjectStatus
from app.modules.reports_export import export
from app.modules.users.models import UserRole

BASE = "/api/v1/work-reports"
AM = "/api/v1/activity-master"
ROWS = "/api/v1/reports-export/activity-rows"
XLSX = "/api/v1/reports-export/activity-rows.xlsx"
TODAY = date.today().isoformat()


@pytest.fixture()
def setup_author(make_user, make_employee, make_project, make_project_member, login):
    def _make(*, email="emp@x.com", code="E-1", proj_code="P-1"):
        u = make_user(email, role=UserRole.employee)
        e = make_employee(employee_code=code, user_id=u.id)
        p = make_project(code=proj_code, status=ProjectStatus.active)
        make_project_member(project_id=p.id, employee_id=e.id)
        return {"user": u, "emp": e, "project": p, "header": login(email)}

    return _make


@pytest.fixture()
def pm_header(auth_header):
    return auth_header(email="pm@x.com", role=UserRole.project_manager)


def _make_sub(client, admin_header, *, activity_name="MTL", name="Sub", **body):
    a = client.post(
        f"{AM}/activities", json={"name": activity_name}, headers=admin_header
    ).json()
    res = client.post(
        f"{AM}/activities/{a['id']}/sub-activities",
        json={"name": name, **body},
        headers=admin_header,
    )
    assert res.status_code in (200, 201), res.text
    return a, res.json()


def _submit(client, header, project_id, tasks):
    created = client.post(
        BASE, headers=header, json={"report_date": TODAY, "tasks": tasks}
    ).json()
    client.post(f"{BASE}/{created['id']}/submit", headers=header)
    return created


def _first_activity(client, header):
    body = client.get(ROWS, headers=header).json()
    assert body["rows"], "expected at least one report row"
    acts = body["rows"][0]["activities"]
    assert acts, "expected at least one activity on the row"
    return acts[0]


# ── API surface ────────────────────────────────────────────────────────────


def test_activity_rows_api_exposes_all_six_counts(client, setup_author, pm_header):
    a = setup_author()
    _, sub = _make_sub(
        client, pm_header, name="ALL SIX", benchmark_type="NUMERIC_DAILY",
        benchmark_value=500, relevant_count_field="pages",
    )
    _submit(
        client, a["header"], a["project"].id,
        [{"project_id": str(a["project"].id), "description": "w",
          "sub_activity_id": sub["id"], "tags_count": 1, "docs_count": 2,
          "bom_count": 3, "spares_count": 4, "pages_count": 40, "records_count": 25}],
    )
    cell = _first_activity(client, pm_header)
    # Every count key is present and carries the exact submitted value.
    assert cell["tags"] == 1
    assert cell["docs"] == 2
    assert cell["bom"] == 3
    assert cell["spares"] == 4
    assert cell["pages"] == 40      # PAGES displays its own value
    assert cell["records"] == 25    # RECORDS displays its own value


def test_activity_rows_api_zeroes_legacy_pages_records(client, setup_author, pm_header):
    """A row submitted with no pages/records (as every pre-0058 row effectively
    was) reads back as 0, never null, matching tags/docs/bom/spares."""
    a = setup_author()
    _, sub = _make_sub(client, pm_header, name="LEGACY", benchmark_type="TASK_STATUS_ONLY")
    _submit(
        client, a["header"], a["project"].id,
        [{"project_id": str(a["project"].id), "description": "w",
          "sub_activity_id": sub["id"], "tags_count": 7, "count_field": "tags"}],
    )
    cell = _first_activity(client, pm_header)
    assert cell["tags"] == 7      # unrelated existing count unchanged
    assert cell["pages"] == 0
    assert cell["records"] == 0


# ── Excel export ────────────────────────────────────────────────────────────


def _sample_row(**counts):
    base = {"tags": 0, "docs": 0, "bom": 0, "spares": 0, "pages": 0, "records": 0}
    base.update(counts)
    return {
        "employee_label": "E-1 - Test User",
        "report_date": date.today(),
        "day_status": None,
        "remarks": "the day remark",
        "activities": [{
            "project_code": "P-1",
            "activity_type": "MTL",
            "sub_activity_type": "ALL SIX",
            **base,
        }],
    }


def test_export_header_order_places_pages_records_after_spares():
    wb = openpyxl.load_workbook(export.build_workbook([_sample_row()], max_activities=1))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    # One row per activity: identity columns, then the activity's own columns.
    assert headers[:8] == [
        "EMPLOYEE ID & NAME", "DATE", "DAY", "DAY STATUS", "HALF",
        "PROJECT CODE", "ACTIVITY TYPE", "SUB ACTIVITY TYPE",
    ]
    assert headers[8:14] == [
        "NO. OF TAGS", "NO. OF DOCS", "NO. OF BOM HEADER", "NO. OF SPARES",
        "NO. OF PAGES", "NO. OF RECORDS",
    ]
    # Pages/Records land immediately after Spares, before Benchmark + Remarks.
    assert headers.index("NO. OF PAGES") == headers.index("NO. OF SPARES") + 1
    assert headers.index("NO. OF RECORDS") == headers.index("NO. OF PAGES") + 1
    assert headers[-2:] == ["BENCHMARK", "DAY REMARKS"]
    # No numbered repeats survive — that is the whole point of the new layout.
    assert not [h for h in headers if h and h.rstrip().endswith(" 2")]


def test_export_writes_pages_records_values_without_shifting_remarks():
    row = _sample_row(tags=1, docs=2, bom=3, spares=4, pages=40, records=25)
    wb = openpyxl.load_workbook(export.build_workbook([row], max_activities=1))
    ws = wb.active
    header_to_col = {c.value: c.column for c in ws[1]}
    data = ws[2]

    def val(label):
        return data[header_to_col[label] - 1].value

    assert val("NO. OF TAGS") == 1
    assert val("NO. OF DOCS") == 2
    assert val("NO. OF BOM HEADER") == 3
    assert val("NO. OF SPARES") == 4
    assert val("NO. OF PAGES") == 40
    assert val("NO. OF RECORDS") == 25
    # The remark stayed in the remarks column — no off-by-two shift. Uppercased
    # like every other text cell in the sheet; the stored remark is untouched.
    assert val("DAY REMARKS") == "THE DAY REMARK"


def test_export_zeroes_missing_pages_records():
    """A legacy activity dict without pages/records still exports 0, not blank."""
    row = _sample_row()
    del row["activities"][0]["pages"]
    del row["activities"][0]["records"]
    wb = openpyxl.load_workbook(export.build_workbook([row], max_activities=1))
    ws = wb.active
    header_to_col = {c.value: c.column for c in ws[1]}
    data = ws[2]
    assert data[header_to_col["NO. OF PAGES"] - 1].value == 0
    assert data[header_to_col["NO. OF RECORDS"] - 1].value == 0
