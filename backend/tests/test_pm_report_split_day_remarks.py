"""PM Weekly Activity Report — split-day remarks reach the single Remarks column
(preview) and Day Remarks cell (Excel), combined and labelled.

A split-day report stores its remark PER HALF in work_report_periods.remarks; the
header `remarks` is empty. The report/export must therefore combine the two half
remarks — labelled "First Half:" / "Second Half:", ordered first-then-second
regardless of task/DB order, newline-separated — rather than reading only the
(empty) header remark. Full-day reports are unchanged.
"""
from datetime import date

import openpyxl
import pytest

from app.modules.projects.models import ProjectStatus
from app.modules.reports_export import export
from app.modules.users.models import UserRole
from app.modules.work_reports.service import format_report_remarks

BASE = "/api/v1/work-reports"
ROWS = "/api/v1/reports-export/activity-rows"
TODAY = date.today().isoformat()

FIRST = "worked on this docs in first half"
SECOND = "worked on two activities"
COMBINED = f"First Half: {FIRST}\nSecond Half: {SECOND}"


# ── the helper in isolation ────────────────────────────────────────────────


def test_format_both_halves_combined_in_order_with_newline():
    out = format_report_remarks(
        "split_day", None, {"first_half": FIRST, "second_half": SECOND}
    )
    assert out == COMBINED
    assert "\n" in out
    assert out.index("First Half") < out.index("Second Half")


def test_format_ordering_independent_of_dict_insertion_order():
    out = format_report_remarks(
        "split_day", None, {"second_half": SECOND, "first_half": FIRST}
    )
    assert out == COMBINED


def test_format_first_half_only_has_no_empty_second_line():
    out = format_report_remarks("split_day", None, {"first_half": FIRST})
    assert out == f"First Half: {FIRST}"
    assert "Second Half" not in out


def test_format_second_half_only_has_no_empty_first_line():
    out = format_report_remarks("split_day", None, {"second_half": SECOND})
    assert out == f"Second Half: {SECOND}"
    assert "First Half" not in out


def test_format_both_blank_returns_empty_string():
    assert format_report_remarks("split_day", None, {}) == ""
    assert format_report_remarks(
        "split_day", None, {"first_half": "  ", "second_half": None}
    ) == ""


def test_format_full_day_remark_unchanged_and_unprefixed():
    assert format_report_remarks("full_day", "just my day note", {}) == "just my day note"
    # A stray period map is ignored for a full-day report.
    assert format_report_remarks("full_day", " trimmed ", {"first_half": "x"}) == "trimmed"


def test_format_identical_halves_preserved_twice():
    out = format_report_remarks(
        "split_day", None, {"first_half": "same", "second_half": "same"}
    )
    assert out == "First Half: same\nSecond Half: same"
    assert out.count("same") == 2


def test_format_trims_each_half_but_keeps_internal_text():
    out = format_report_remarks(
        "split_day", None,
        {"first_half": "  a b  c  ", "second_half": "\td e\n"},
    )
    assert out == "First Half: a b  c\nSecond Half: d e"


# ── through the API + Excel ─────────────────────────────────────────────────


@pytest.fixture()
def day_parts_on():
    from app.core.config import settings

    prev = settings.REPORT_DAY_PARTS_ENABLED
    settings.REPORT_DAY_PARTS_ENABLED = True
    try:
        yield
    finally:
        settings.REPORT_DAY_PARTS_ENABLED = prev


@pytest.fixture()
def setup_author(make_user, make_employee, make_project, make_project_member, login):
    def _make(*, email="emp@x.com", code="E-1", proj_code="P-1"):
        u = make_user(email, role=UserRole.employee)
        e = make_employee(employee_code=code, user_id=u.id)
        p = make_project(code=proj_code, status=ProjectStatus.active)
        make_project_member(project_id=p.id, employee_id=e.id)
        return {"user": u, "emp": e, "project": p, "header": login(email)}

    return _make


@pytest.fixture()
def pm_header(auth_header):
    return auth_header(email="pm@x.com", role=UserRole.project_manager)


def _working(project_id, remarks, **counts):
    return {
        "period_status": "work_at_office",
        "location": "chennai",
        "remarks": remarks,
        "tasks": [{"project_id": str(project_id), "description": "work", **counts}],
    }


def _create_split(client, header, project_id, first, second):
    res = client.post(BASE, headers=header, json={
        "report_date": TODAY,
        "report_mode": "split_day",
        "periods": [
            {"day_part": "first_half", **first},
            {"day_part": "second_half", **second},
        ],
    })
    assert res.status_code == 201, res.text
    created = res.json()
    client.post(f"{BASE}/{created['id']}/submit", headers=header)
    return created


def _report_row(client, header):
    body = client.get(ROWS, headers=header).json()
    assert body["rows"], "expected a report row"
    return body["rows"][0]


def test_api_returns_combined_split_day_remarks(
    client, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    pid = a["project"].id
    _create_split(
        client, a["header"], pid,
        _working(pid, FIRST, tags_count=1, docs_count=2, bom_count=3,
                 spares_count=4, pages_count=5, records_count=6),
        _working(pid, SECOND, tags_count=7),
    )
    row = _report_row(client, pm_header)
    assert row["remarks"] == COMBINED

    # First-half activity block comes before the second-half one (deterministic).
    assert len(row["activities"]) == 2
    first_cell = row["activities"][0]
    # Counts are carried through untouched by the remarks change.
    assert first_cell["tags"] == 1
    assert first_cell["docs"] == 2
    assert first_cell["bom"] == 3
    assert first_cell["spares"] == 4
    assert first_cell["pages"] == 5
    assert first_cell["records"] == 6
    assert row["activities"][1]["tags"] == 7


def test_api_full_day_remark_unchanged(client, setup_author, pm_header):
    a = setup_author()
    pid = a["project"].id
    created = client.post(BASE, headers=a["header"], json={
        "report_date": TODAY,
        "day_status": "work_at_office",
        "location": "chennai",
        "remarks": "plain full day remark",
        "tasks": [{"project_id": str(pid), "description": "work", "tags_count": 9}],
    }).json()
    client.post(f"{BASE}/{created['id']}/submit", headers=a["header"])
    row = _report_row(client, pm_header)
    assert row["remarks"] == "plain full day remark"
    assert row["activities"][0]["tags"] == 9


def test_excel_day_remarks_cell_holds_combined_text_and_wraps():
    row = {
        "employee_label": "E-1 - Test User",
        "report_date": date.today(),
        "day_status": None,
        "remarks": COMBINED,
        "activities": [{
            "project_code": "P-1", "activity_type": "MTL", "sub_activity_type": "S",
            "tags": 1, "docs": 2, "bom": 3, "spares": 4, "pages": 5, "records": 6,
        }],
    }
    wb = openpyxl.load_workbook(export.build_workbook([row], max_activities=1))
    ws = wb.active
    header_to_col = {c.value: c.column for c in ws[1]}
    cell = ws.cell(row=2, column=header_to_col["Day Remarks"])
    assert cell.value == COMBINED
    assert "\n" in cell.value
    assert cell.alignment.wrap_text is True
    assert cell.alignment.vertical == "top"
