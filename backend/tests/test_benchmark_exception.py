"""Numeric benchmark availability exception — NO_FURTHER_AVAILABLE_WORK.

The rule: an employee completed every TAG that was actually available for a
numeric activity, but fewer existed than the configured target. The row is
EVALUATED as achieved (100%, 0 difference, 0 pending) while the sheet keeps
REPORTING the real count. Target 100 / actual 40 stays 40 everywhere a number is
displayed or stored; only the percentage and the pending change.

Three layers are pinned here:

1. the pure rules (benchmark_exception.py) — what makes an exception valid, and
   what "effective actual" means;
2. the API — save, submit, clear-when-invalid, and round-trip on edit/reopen;
3. the workbook — the real actual, the zeroed pending, the composed remark, the
   amber ACTUAL cell, the totals, and the unchanged column layout.

The exception is STRUCTURAL, never textual: typing the sentence into a remark
changes nothing (test_typed_remark_wording_changes_no_calculation).
"""
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest

from app.modules.activity_master.benchmark_exception import (
    BENCHMARK_EXCEPTION_NO_FURTHER_AVAILABLE_WORK,
    effective_actual,
    export_exception_remark,
    is_eligible_row,
    is_valid_exception,
)
from app.modules.activity_master.service import compute_week_bounds
from app.modules.projects.models import ProjectStatus
from app.modules.reports_export.export import (
    _EXC_AMBER_FILL,
    _EXC_AMBER_FONT,
    _export_remarks,
    _upper,
)
from app.modules.users.models import UserRole

# Local shorthand for the one Phase 1 code — it appears on nearly every line.
NO_FURTHER = BENCHMARK_EXCEPTION_NO_FURTHER_AVAILABLE_WORK

BASE = "/api/v1/work-reports"
EXPORT_URL = "/api/v1/benchmarks/pending-export.xlsx"
SYSTEM_REMARK = "[NO FURTHER TAGS WERE AVAILABLE FOR THIS ACTIVITY]"
AMBER_FILL = "FFFFF2CC"
AMBER_FONT = "FF7F6000"
AMBER_BORDER = "FFBF8F00"

# Column map — identical to test_benchmark_pending_export.py. Repeated rather
# than imported so a silent reorder there cannot silently pass here too.
EMP, DATE_C, DAY_PART, ACTIVITY, ACH, DIFF, SUB, REMARKS, PROJECT = (
    1, 2, 3, 4, 5, 6, 7, 8, 9
)
TGT_TAGS = 10
ACT_TAGS = 16
PEN_TAGS = 22
CYC_START, CYC_END = 28, 29
LAST_COL = 29

EXPECTED_HEADER = [
    "EMP CODE & NAME", "DATE", "DAY PART", "ACTIVITY", "ACHIEVEMENT %",
    "DIFFERENCE %", "SUB ACTIVITY", "REMARKS", "PROJECT CODE & TITLE",
    "TAGS", "DOCS", "BOM", "SPARES", "PAGES", "RECORDS",
    "TAGS", "DOCS", "BOM", "SPARES", "PAGES", "RECORDS",
    "TAGS", "DOCS", "BOM", "SPARES", "PAGES", "RECORDS",
    "CYCLE START", "CYCLE END",
]


# --- fixtures / helpers ------------------------------------------------------

@pytest.fixture()
def setup_author(make_user, make_employee, make_project, make_project_member, login):
    def _make(*, email="emp@x.com", code="E-1", proj_code="P-1",
              first_name="Test", last_name="User"):
        u = make_user(email, role=UserRole.employee)
        e = make_employee(
            employee_code=code, user_id=u.id, first_name=first_name, last_name=last_name
        )
        p = make_project(code=proj_code, status=ProjectStatus.active)
        make_project_member(project_id=p.id, employee_id=e.id)
        return {"user": u, "emp": e, "project": p, "header": login(email)}

    return _make


@pytest.fixture()
def activity_admin(auth_header):
    return auth_header(email="pm@x.com", role=UserRole.project_manager)


def _make_sub(client, admin, *, value, name, count_field="tags", btype="NUMERIC_DAILY"):
    a = client.post(
        "/api/v1/activity-master/activities",
        json={"name": f"Activity for {name}"}, headers=admin,
    ).json()
    body = {"name": name, "benchmark_type": btype}
    if btype in ("NUMERIC", "NUMERIC_DAILY", "TASK_WITH_QUANTITY"):
        body["benchmark_value"] = value
        body["relevant_count_field"] = count_field
    sub = client.post(
        f"/api/v1/activity-master/activities/{a['id']}/sub-activities",
        json=body, headers=admin,
    ).json()
    return a, sub


def _prev_cycle() -> tuple[date, date]:
    start, end = compute_week_bounds(date.today())
    return start - timedelta(days=7), end - timedelta(days=7)


def _create(client, header, project_id, sub_id, on_date, qty, *,
            count_field="tags", exception=None, remarks=None):
    """Create a DRAFT report with one activity row. Returns the response body."""
    task = {
        "project_id": str(project_id), "description": "work",
        "sub_activity_id": sub_id, f"{count_field}_count": qty,
    }
    if exception is not None:
        task["benchmark_exception_code"] = exception
    payload = {"report_date": on_date.isoformat(), "tasks": [task]}
    if remarks is not None:
        payload["remarks"] = remarks
    res = client.post(BASE, headers=header, json=payload)
    assert res.status_code == 201, res.text
    return res.json()


def _submit(client, header, report_id):
    res = client.post(f"{BASE}/{report_id}/submit", headers=header)
    assert res.status_code == 200, res.text
    return res.json()


def _create_and_submit(client, header, project_id, sub_id, on_date, qty, **kw):
    body = _create(client, header, project_id, sub_id, on_date, qty, **kw)
    return _submit(client, header, body["id"])


def _sheet(client, admin):
    res = client.get(EXPORT_URL, headers=admin)
    assert res.status_code == 200
    # Round-trips through openpyxl: a workbook Excel would flag for repair does
    # not load cleanly here either.
    return openpyxl.load_workbook(BytesIO(res.content)).active


def _detail(ws, sub_name, on_date):
    for r in range(3, ws.max_row + 1):
        cell = ws.cell(r, DATE_C).value
        value = cell.date() if hasattr(cell, "date") else cell
        if ws.cell(r, SUB).value == sub_name.upper() and value == on_date:
            return r
    raise AssertionError(f"detail row not found: {sub_name} / {on_date}")


def _total(ws, sub_name):
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, SUB).value == sub_name.upper() and ws.cell(r, PROJECT).value == "TOTAL":
            return r
    raise AssertionError(f"total row not found: {sub_name}")


def _fill(cell):
    return cell.fill.fgColor.rgb if cell.fill and cell.fill.fill_type else None


# --- 1. the pure rules -------------------------------------------------------

def test_valid_exception_requires_every_condition():
    ok = dict(
        benchmark_type="NUMERIC_DAILY", count_field="tags",
        target=Decimal("100"), actual=40,
    )
    assert is_valid_exception(NO_FURTHER, **ok) is True

    # An unknown code is never honoured, however right the numbers look.
    assert is_valid_exception("SOMETHING_ELSE", **ok) is False
    assert is_valid_exception(None, **ok) is False
    # Task modes — including the one that carries a quantity.
    for bad_type in ("TASK_BASED", "TASK_STATUS_ONLY", "TASK_WITH_QUANTITY", None):
        assert is_valid_exception(NO_FURTHER, **{**ok, "benchmark_type": bad_type}) is False
    # Phase 1 is TAGS only.
    for bad_unit in ("docs", "bom", "spares", "pages", "records", None):
        assert is_valid_exception(NO_FURTHER, **{**ok, "count_field": bad_unit}) is False
    # A target must exist and be positive.
    for bad_target in (None, 0, Decimal("0"), -5):
        assert is_valid_exception(NO_FURTHER, **{**ok, "target": bad_target}) is False
    # An actual must be present. 0 IS present — "none were available".
    assert is_valid_exception(NO_FURTHER, **{**ok, "actual": None}) is False
    assert is_valid_exception(NO_FURTHER, **{**ok, "actual": 0}) is True
    # …and must be strictly below the target.
    assert is_valid_exception(NO_FURTHER, **{**ok, "actual": 100}) is False
    assert is_valid_exception(NO_FURTHER, **{**ok, "actual": 140}) is False
    assert is_valid_exception(NO_FURTHER, **{**ok, "actual": 99}) is True


def test_eligibility_is_configuration_only():
    """Static eligibility ignores the numbers — it is what the save path can
    check before any target is frozen."""
    assert is_eligible_row("NUMERIC", "tags") is True
    assert is_eligible_row("NUMERIC_DAILY", "tags") is True
    assert is_eligible_row("TASK_WITH_QUANTITY", "tags") is False
    assert is_eligible_row("NUMERIC_DAILY", "pages") is False
    assert is_eligible_row(None, None) is False


def test_effective_actual_never_replaces_the_real_one():
    # With an exception the CALCULATION uses the target…
    assert effective_actual(Decimal("100"), Decimal("40"), has_exception=True) == Decimal("100")
    # …and without one it is simply the actual, untouched.
    assert effective_actual(Decimal("100"), Decimal("40"), has_exception=False) == Decimal("40")
    # Overachievement is passed through, never capped.
    assert effective_actual(Decimal("100"), Decimal("140"), has_exception=False) == Decimal("140")
    # No target to stand in for -> the actual survives rather than becoming None.
    assert effective_actual(None, Decimal("40"), has_exception=True) == Decimal("40")


def test_export_remark_wording_is_exact():
    assert export_exception_remark("tags") == SYSTEM_REMARK
    assert export_exception_remark("pages") is None


# --- 2. export-only text helpers --------------------------------------------

def test_upper_only_touches_strings():
    assert _upper("Waiting on drawings") == "WAITING ON DRAWINGS"
    assert _upper(None) is None
    assert _upper("") is None            # blank collapses to an empty cell
    assert _upper("   ") is None         # whitespace-only too
    # Numbers, dates and percentages pass through untouched and stay typed.
    assert _upper(40) == 40
    assert _upper(0.8) == 0.8
    assert _upper(Decimal("12.5")) == Decimal("12.5")
    assert _upper(date(2026, 8, 6)) == date(2026, 8, 6)
    assert _upper(True) is True
    assert _upper("=SUM(A1:A2)") == "=SUM(A1:A2)"


def test_export_remarks_composition():
    # System remark first, user remark after, " | " between.
    assert _export_remarks("waiting on vendor", NO_FURTHER, "tags") == (
        f"{SYSTEM_REMARK} | WAITING ON VENDOR"
    )
    # No user remark -> the system remark alone.
    assert _export_remarks(None, NO_FURTHER, "tags") == SYSTEM_REMARK
    assert _export_remarks("   ", NO_FURTHER, "tags") == SYSTEM_REMARK
    # No exception -> the user's remark alone, never a prefix.
    assert _export_remarks("waiting on vendor", None, "tags") == "WAITING ON VENDOR"
    assert _export_remarks(None, None, "tags") is None
    # An unknown code is not an exception.
    assert _export_remarks("x", "MADE_UP", "tags") == "X"


def test_export_remarks_never_doubles_the_system_remark():
    """Exporting the same cycle twice, or feeding an already-composed value back
    in, must not stack the prefix."""
    once = _export_remarks("waiting", NO_FURTHER, "tags")
    twice = _export_remarks(once, NO_FURTHER, "tags")
    assert twice == once
    assert twice.count(SYSTEM_REMARK) == 1
    bare = _export_remarks(None, NO_FURTHER, "tags")
    assert _export_remarks(bare, NO_FURTHER, "tags") == bare


def test_typed_wording_in_a_remark_is_not_an_exception():
    """An employee typing the sentence by hand gets no bracketed prefix — the
    composition reads the CODE, never the text."""
    typed = "no further tags were available for this activity"
    assert _export_remarks(typed, None, "tags") == typed.upper()
    assert not _export_remarks(typed, None, "tags").startswith(SYSTEM_REMARK)


# --- 3. API: save, submit, validate, round-trip ------------------------------

def test_exception_saves_and_survives_submit(client, setup_author, activity_admin):
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]

    draft = _create(client, a["header"], a["project"].id, sub["id"], d, 40,
                    exception=NO_FURTHER)
    assert draft["tasks"][0]["benchmark_exception_code"] == NO_FURTHER
    assert draft["tasks"][0]["tags_count"] == 40      # the real count is stored

    submitted = _submit(client, a["header"], draft["id"])
    assert submitted["tasks"][0]["benchmark_exception_code"] == NO_FURTHER
    assert submitted["tasks"][0]["tags_count"] == 40


def test_default_is_no_exception(client, setup_author, activity_admin):
    """A row saved without the field — every existing report — reads null and
    behaves exactly as before."""
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]

    body = _create_and_submit(client, a["header"], a["project"].id, sub["id"], d, 40)
    assert body["tasks"][0]["benchmark_exception_code"] is None


def test_exception_is_cleared_when_actual_equals_target(client, setup_author, activity_admin):
    """Nothing was left unavailable — the exception does not apply and is
    dropped at submit rather than freezing a false 100%."""
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]

    body = _create_and_submit(client, a["header"], a["project"].id, sub["id"], d, 100,
                              exception=NO_FURTHER)
    assert body["tasks"][0]["benchmark_exception_code"] is None
    assert body["tasks"][0]["tags_count"] == 100


def test_exception_is_cleared_when_actual_exceeds_target(client, setup_author, activity_admin):
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]

    body = _create_and_submit(client, a["header"], a["project"].id, sub["id"], d, 140,
                              exception=NO_FURTHER)
    assert body["tasks"][0]["benchmark_exception_code"] is None


def test_overachievement_still_reads_above_100_percent(client, setup_author, activity_admin):
    """The existing uncapped behaviour is untouched by the new field: 140/100
    still reads 140%, not a capped 100%."""
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d, 140,
                       exception=NO_FURTHER)

    ws = _sheet(client, activity_admin)
    t = _total(ws, "FMTL")
    assert ws.cell(t, ACH).value == 1.4
    assert ws.cell(t, ACT_TAGS).value == 140
    assert _fill(ws.cell(t, ACT_TAGS)) is None     # cleared -> no amber


def test_exception_is_cleared_for_a_task_activity(client, setup_author, activity_admin):
    """A task mode cannot carry an exception — rejected at save, before the
    numbers are even considered."""
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="TASKY",
                       btype="TASK_WITH_QUANTITY")
    d = _prev_cycle()[0]

    body = _create(client, a["header"], a["project"].id, sub["id"], d, 40,
                   exception=NO_FURTHER)
    assert body["tasks"][0]["benchmark_exception_code"] is None


def test_exception_is_cleared_for_a_non_tags_unit(client, setup_author, activity_admin):
    """Phase 1 is TAGS only: the same shortfall on a PAGES activity carries no
    exception."""
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=500, name="MTL-PAGES",
                       count_field="pages")
    d = _prev_cycle()[0]

    body = _create(client, a["header"], a["project"].id, sub["id"], d, 400,
                   count_field="pages", exception=NO_FURTHER)
    assert body["tasks"][0]["benchmark_exception_code"] is None


def test_unknown_exception_code_is_rejected(client, setup_author, activity_admin):
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]

    res = client.post(BASE, headers=a["header"], json={
        "report_date": d.isoformat(),
        "tasks": [{
            "project_id": str(a["project"].id), "description": "w",
            "sub_activity_id": sub["id"], "tags_count": 40,
            "benchmark_exception_code": "FREE_LUNCH",
        }],
    })
    assert res.status_code == 422


def test_exception_round_trips_through_an_edit(client, setup_author, activity_admin):
    """Reopening a draft returns the exception, and re-saving keeps it. Editing
    the count up to the target drops it on the next submit."""
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]
    draft = _create(client, a["header"], a["project"].id, sub["id"], d, 40,
                    exception=NO_FURTHER)

    # Reopened for editing: the server returns what it holds.
    fetched = client.get(f"{BASE}/{draft['id']}", headers=a["header"]).json()
    assert fetched["tasks"][0]["benchmark_exception_code"] == NO_FURTHER

    # Re-saved unchanged: still there.
    edited = client.patch(f"{BASE}/{draft['id']}", headers=a["header"], json={
        "tasks": [{
            "project_id": str(a["project"].id), "description": "work",
            "sub_activity_id": sub["id"], "tags_count": 40,
            "benchmark_exception_code": NO_FURTHER,
        }],
    })
    assert edited.status_code == 200, edited.text
    assert edited.json()["tasks"][0]["benchmark_exception_code"] == NO_FURTHER

    # Edited up to target: the exception no longer describes the row.
    raised = client.patch(f"{BASE}/{draft['id']}", headers=a["header"], json={
        "tasks": [{
            "project_id": str(a["project"].id), "description": "work",
            "sub_activity_id": sub["id"], "tags_count": 100,
            "benchmark_exception_code": NO_FURTHER,
        }],
    })
    assert raised.status_code == 200
    body = _submit(client, a["header"], draft["id"])
    assert body["tasks"][0]["benchmark_exception_code"] is None


def test_typed_remark_wording_changes_no_calculation(client, setup_author, activity_admin):
    """The exact sentence typed into the remark, with NO structured exception:
    the benchmark still reads 40/100 = 40%, and the remark carries no
    system-generated prefix."""
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]
    _create_and_submit(
        client, a["header"], a["project"].id, sub["id"], d, 40,
        remarks="No further tags were available for this activity",
    )

    ws = _sheet(client, activity_admin)
    row = _detail(ws, "FMTL", d)
    total = _total(ws, "FMTL")
    assert ws.cell(total, ACH).value == 0.4              # unchanged by the words
    assert ws.cell(total, DIFF).value == pytest.approx(0.6)
    assert ws.cell(row, PEN_TAGS).value == 60
    assert _fill(ws.cell(row, ACT_TAGS)) is None         # no amber
    assert ws.cell(row, REMARKS).value == (
        "NO FURTHER TAGS WERE AVAILABLE FOR THIS ACTIVITY"
    )
    assert not ws.cell(row, REMARKS).value.startswith("[")


# --- 4. the workbook ---------------------------------------------------------

def test_below_target_without_exception_is_unchanged(client, setup_author, activity_admin):
    """Baseline: 40 of 100 with no exception still reads 40% / 60% difference /
    60 pending, exactly as before this feature."""
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d, 40)

    ws = _sheet(client, activity_admin)
    row, total = _detail(ws, "FMTL", d), _total(ws, "FMTL")
    assert ws.cell(row, ACT_TAGS).value == 40
    assert ws.cell(row, PEN_TAGS).value == 60
    assert ws.cell(total, ACH).value == 0.4
    assert ws.cell(total, DIFF).value == pytest.approx(0.6)
    assert ws.cell(total, PEN_TAGS).value == 60


def test_exception_row_reports_real_actual_and_evaluates_as_achieved(
    client, db, setup_author, activity_admin,
):
    """The approved example, end to end. Target 100, actual 40, exception set:
    ACTUAL reads 40, PENDING reads 0, ACHIEVEMENT 100%, DIFFERENCE 0% — and the
    stored count in PostgreSQL is still 40."""
    from app.modules.work_reports.models import WorkReportTask

    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]
    body = _create_and_submit(client, a["header"], a["project"].id, sub["id"], d, 40,
                              exception=NO_FURTHER)

    ws = _sheet(client, activity_admin)
    row, total = _detail(ws, "FMTL", d), _total(ws, "FMTL")

    assert ws.cell(row, TGT_TAGS).value == 100
    assert ws.cell(row, ACT_TAGS).value == 40     # the REAL count, not the target
    assert ws.cell(row, PEN_TAGS).value == 0
    assert ws.cell(total, TGT_TAGS).value == 100
    assert ws.cell(total, ACT_TAGS).value == 40   # totals sum REAL actuals
    assert ws.cell(total, PEN_TAGS).value == 0
    assert ws.cell(total, ACH).value == 1.0
    assert ws.cell(total, ACH).number_format == "0.00%"
    assert ws.cell(total, DIFF).value == 0
    assert ws.cell(total, DIFF).number_format == "0.00%"

    # The database is untouched by the evaluation.
    import uuid as uuid_mod
    stored = db.get(WorkReportTask, uuid_mod.UUID(body["tasks"][0]["id"]))
    db.refresh(stored)
    assert stored.tags_count == 40
    assert stored.benchmark_exception_code == NO_FURTHER


def test_exception_actual_cell_carries_the_amber_style(client, setup_author, activity_admin):
    """Only the ACTUAL COMPLETED -> TAGS cell is styled: amber fill, amber bold
    font, thin amber border. Never the target, never the pending, never the row."""
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d, 40,
                       exception=NO_FURTHER)

    ws = _sheet(client, activity_admin)
    row = _detail(ws, "FMTL", d)
    cell = ws.cell(row, ACT_TAGS)
    assert _fill(cell) == AMBER_FILL == _EXC_AMBER_FILL.fgColor.rgb
    assert cell.font.color.rgb == AMBER_FONT == _EXC_AMBER_FONT.color.rgb
    assert cell.font.bold is True
    for side in ("top", "bottom", "left", "right"):
        assert getattr(cell.border, side).style == "thin", side
        assert getattr(cell.border, side).color.rgb == AMBER_BORDER, side

    # Nothing else on the detail row is filled — not the target, not the pending.
    for c in range(1, LAST_COL + 1):
        if c == ACT_TAGS:
            continue
        assert _fill(ws.cell(row, c)) is None, ws.cell(row, c).coordinate
    # And the values under/around the amber are unchanged.
    assert ws.cell(row, TGT_TAGS).value == 100
    assert ws.cell(row, ACT_TAGS).value == 40


def test_total_actual_cell_is_styled_when_a_contributing_row_excepts(
    client, setup_author, activity_admin,
):
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d, 40,
                       exception=NO_FURTHER)

    ws = _sheet(client, activity_admin)
    cell = ws.cell(_total(ws, "FMTL"), ACT_TAGS)
    assert _fill(cell) == AMBER_FILL
    assert cell.font.color.rgb == AMBER_FONT
    assert cell.font.bold is True
    assert cell.border.top.style == "thin"


def test_total_actual_cell_is_not_styled_without_an_exception(
    client, setup_author, activity_admin,
):
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d, 40)

    ws = _sheet(client, activity_admin)
    assert _fill(ws.cell(_total(ws, "FMTL"), ACT_TAGS)) is None


def test_exception_remark_precedes_the_employees_own(client, db, setup_author, activity_admin):
    """The system remark comes first, the employee's after it, and the stored
    remark in PostgreSQL is untouched (still its original case, no prefix)."""
    from app.modules.work_reports.models import DailyWorkReport

    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]
    body = _create_and_submit(client, a["header"], a["project"].id, sub["id"], d, 40,
                              exception=NO_FURTHER, remarks="Waiting on vendor")

    ws = _sheet(client, activity_admin)
    assert ws.cell(_detail(ws, "FMTL", d), REMARKS).value == (
        f"{SYSTEM_REMARK} | WAITING ON VENDOR"
    )

    import uuid as uuid_mod
    stored = db.get(DailyWorkReport, uuid_mod.UUID(body["id"]))
    db.refresh(stored)
    assert stored.remarks == "Waiting on vendor"     # original case, no prefix


def test_exception_remark_alone_when_the_employee_wrote_none(
    client, setup_author, activity_admin,
):
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d, 40,
                       exception=NO_FURTHER)

    ws = _sheet(client, activity_admin)
    assert ws.cell(_detail(ws, "FMTL", d), REMARKS).value == SYSTEM_REMARK


def test_exporting_twice_does_not_duplicate_the_system_remark(
    client, setup_author, activity_admin,
):
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d, 40,
                       exception=NO_FURTHER, remarks="Waiting on vendor")

    first = _sheet(client, activity_admin)
    second = _sheet(client, activity_admin)
    for ws in (first, second):
        value = ws.cell(_detail(ws, "FMTL", d), REMARKS).value
        assert value.count(SYSTEM_REMARK) == 1
        assert value == f"{SYSTEM_REMARK} | WAITING ON VENDOR"


def test_every_exported_text_cell_is_uppercase(client, setup_author, activity_admin):
    """Headers, employee, day part, activity, sub-activity, remarks, project and
    the TOTAL marker — every string in the sheet reads uppercase, while the
    numbers, dates and percentages stay typed."""
    a = setup_author(code="cdc-9", first_name="Anita", last_name="raj")
    _, sub = _make_sub(client, activity_admin, value=100, name="Fmtl Rework")
    d = _prev_cycle()[0]
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d, 40,
                       remarks="waiting on Vendor drawings")

    ws = _sheet(client, activity_admin)
    for r in range(1, ws.max_row + 1):
        for c in range(1, LAST_COL + 1):
            value = ws.cell(r, c).value
            if isinstance(value, str):
                assert value == value.upper(), ws.cell(r, c).coordinate

    row = _detail(ws, "FMTL REWORK", d)
    assert ws.cell(row, EMP).value == "CDC-9 - ANITA RAJ"
    assert ws.cell(row, SUB).value == "FMTL REWORK"
    assert ws.cell(row, REMARKS).value == "WAITING ON VENDOR DRAWINGS"
    assert ws.cell(row, PROJECT).value == "P-1 - TEST PROJECT"
    # Numbers and dates did not become strings.
    assert isinstance(ws.cell(row, ACT_TAGS).value, (int, float))
    assert not isinstance(ws.cell(row, DATE_C).value, str)
    assert ws.cell(row, DATE_C).number_format == "yyyy-mm-dd"


def test_columns_are_unchanged_by_the_feature(client, setup_author, activity_admin):
    """No column added, removed, renamed or reordered — and none of the
    withdrawn names (Evaluated Actual, Exception, Status, Exception Reason)
    exists anywhere in the header."""
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d = _prev_cycle()[0]
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d, 40,
                       exception=NO_FURTHER)

    ws = _sheet(client, activity_admin)
    assert [ws.cell(2, c).value for c in range(1, LAST_COL + 1)] == EXPECTED_HEADER
    # Checked BEFORE touching column 30: ws.cell() materialises a cell, which
    # would bump max_column itself.
    assert ws.max_column == LAST_COL
    assert ws.cell(2, LAST_COL + 1).value is None      # nothing in column 30
    assert ws.cell(1, TGT_TAGS).value == "BENCHMARK TARGET"
    assert ws.cell(1, ACT_TAGS).value == "ACTUAL COMPLETED"
    assert ws.cell(1, PEN_TAGS).value == "PENDING BENCHMARK"
    assert {str(m) for m in ws.merged_cells.ranges} == {"J1:O1", "P1:U1", "V1:AA1"}
    assert ws.freeze_panes == "A3"
    assert ws.auto_filter.ref == f"A2:AC{ws.max_row}"
    labels = {ws.cell(2, c).value for c in range(1, LAST_COL + 1)}
    assert labels.isdisjoint(
        {"EVALUATED ACTUAL", "EXCEPTION", "STATUS", "EXCEPTION REASON"}
    )


# --- 5. aggregation: mixed exception / normal rows ---------------------------

def test_mixed_exception_and_normal_rows_aggregate_correctly(
    client, setup_author, activity_admin,
):
    """Two days of one sub-activity: day 1 excepted (40 of 100), day 2 normal
    (60 of 100).

      ACTUAL total   = 40 + 60 = 100  (REAL values)
      effective      = 100 + 60 = 160 (day 1 evaluated at target)
      ACHIEVEMENT    = 160 / 200 = 80%
      DIFFERENCE     = 20%
      PENDING        = max(0, 200 - 160) = 40
    """
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d1 = _prev_cycle()[0]
    d2 = d1 + timedelta(days=1)
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d1, 40,
                       exception=NO_FURTHER)
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d2, 60)

    ws = _sheet(client, activity_admin)
    r1, r2, t = _detail(ws, "FMTL", d1), _detail(ws, "FMTL", d2), _total(ws, "FMTL")

    assert ws.cell(r1, ACT_TAGS).value == 40 and ws.cell(r1, PEN_TAGS).value == 0
    assert ws.cell(r2, ACT_TAGS).value == 60 and ws.cell(r2, PEN_TAGS).value == 40
    assert ws.cell(t, TGT_TAGS).value == 200
    assert ws.cell(t, ACT_TAGS).value == 100          # sum of the REAL actuals
    assert ws.cell(t, PEN_TAGS).value == 40           # from the evaluated result
    assert ws.cell(t, ACH).value == pytest.approx(0.8)
    assert ws.cell(t, DIFF).value == pytest.approx(0.2)
    assert _fill(ws.cell(t, ACT_TAGS)) == AMBER_FILL  # one contributor excepted


def test_all_exception_rows_total_to_exactly_one_hundred_percent(
    client, setup_author, activity_admin,
):
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d1 = _prev_cycle()[0]
    d2 = d1 + timedelta(days=1)
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d1, 40,
                       exception=NO_FURTHER)
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d2, 10,
                       exception=NO_FURTHER)

    ws = _sheet(client, activity_admin)
    t = _total(ws, "FMTL")
    assert ws.cell(t, ACT_TAGS).value == 50           # 40 + 10, the real counts
    assert ws.cell(t, PEN_TAGS).value == 0
    assert ws.cell(t, ACH).value == 1.0
    assert ws.cell(t, DIFF).value == 0


def test_multiple_normal_rows_are_unaffected_by_the_feature(
    client, setup_author, activity_admin,
):
    a = setup_author()
    _, sub = _make_sub(client, activity_admin, value=100, name="FMTL")
    d1 = _prev_cycle()[0]
    d2 = d1 + timedelta(days=1)
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d1, 40)
    _create_and_submit(client, a["header"], a["project"].id, sub["id"], d2, 60)

    ws = _sheet(client, activity_admin)
    t = _total(ws, "FMTL")
    assert ws.cell(t, ACT_TAGS).value == 100
    assert ws.cell(t, PEN_TAGS).value == 100
    assert ws.cell(t, ACH).value == 0.5
    assert _fill(ws.cell(t, ACT_TAGS)) is None


def test_exception_does_not_leak_across_sub_activities(client, setup_author, activity_admin):
    """An exception on one sub-activity never touches another's total, and never
    styles another's ACTUAL cell."""
    a = setup_author()
    _, one = _make_sub(client, activity_admin, value=100, name="FMTL")
    _, two = _make_sub(client, activity_admin, value=100, name="MTL")
    d1 = _prev_cycle()[0]
    d2 = d1 + timedelta(days=1)
    _create_and_submit(client, a["header"], a["project"].id, one["id"], d1, 40,
                       exception=NO_FURTHER)
    _create_and_submit(client, a["header"], a["project"].id, two["id"], d2, 40)

    ws = _sheet(client, activity_admin)
    t_one, t_two = _total(ws, "FMTL"), _total(ws, "MTL")
    assert ws.cell(t_one, ACH).value == 1.0
    assert ws.cell(t_two, ACH).value == 0.4          # untouched
    assert ws.cell(t_two, PEN_TAGS).value == 60
    assert _fill(ws.cell(t_one, ACT_TAGS)) == AMBER_FILL
    assert _fill(ws.cell(t_two, ACT_TAGS)) is None
