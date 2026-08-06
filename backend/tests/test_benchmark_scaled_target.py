"""Whole-unit benchmark targets, and the red shortfall marker in the Excel.

Two rules pinned here:

1. A scaled target is always a WHOLE number, rounded UP. Benchmarks count real
   things — tags, documents, BOM lines — so half of a 35-tag benchmark is 18,
   not 17.5 and not 17. It applies wherever a target is derived: the submit-time
   snapshot, both ledgers, and (mirrored in benchmark-target.ts) the number
   shown beside the input, so an employee is measured against exactly the figure
   they were shown.

2. An ACTUAL COMPLETED cell below its target is filled RED when nothing excuses
   it. Together with the amber accepted-exception marker this makes the ACTUAL
   column self-explaining: amber = the available work ran out, red = work is
   genuinely outstanding, unfilled = target met or beaten.
"""
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest

from app.modules.activity_master.benchmark_exception import (
    BENCHMARK_EXCEPTION_NO_FURTHER_AVAILABLE_WORK as NO_FURTHER,
)
from app.core.config import settings
from app.modules.activity_master.service import compute_week_bounds, scaled_target
from app.modules.projects.models import ProjectStatus
from app.modules.users.models import UserRole

BASE = "/api/v1/work-reports"
EXPORT_URL = "/api/v1/benchmarks/pending-export.xlsx"

SUB, PROJECT = 7, 9
TGT_TAGS, ACT_TAGS, PEN_TAGS = 10, 16, 22
DATE_C = 2
SHORT_RED = "FFFFC7CE"
SHORT_RED_FONT = "FF9C0006"
AMBER = "FFFFF2CC"


# --- 1. the rule itself ------------------------------------------------------

@pytest.mark.parametrize(
    "base, fraction, expected",
    [
        # A full day is the base, untouched.
        (66, 1, 66),
        (35, 1, 35),
        # Evenly divisible halves keep their exact value.
        (66, "0.5", 33),
        (100, "0.5", 50),
        # Odd halves round UP — the reported case is 35 -> 18.
        (35, "0.5", 18),
        (33, "0.5", 17),
        (1, "0.5", 1),
        (3, "0.5", 2),
        # Ceiling, not nearest: anything above 17 lands on 18.
        ("34.2", "0.5", 18),
        ("35.8", "0.5", 18),
        # Degenerate inputs stay arithmetic rather than raising.
        (0, "0.5", 0),
        (None, "0.5", 0),
    ],
)
def test_scaled_target_is_a_whole_number_rounded_up(base, fraction, expected):
    assert scaled_target(base, fraction) == Decimal(expected)


def test_scaled_target_returns_an_integral_decimal():
    """The value is frozen into a Numeric column and written into a numeric
    Excel cell, so it must carry no fractional part at all — not merely display
    as one."""
    value = scaled_target(35, "0.5")
    assert value == value.to_integral_value()
    assert str(value) == "18"


# --- fixtures ----------------------------------------------------------------

@pytest.fixture()
def setup_author(make_user, make_employee, make_project, make_project_member, login):
    def _make(*, email="emp@x.com", code="E-1", proj_code="P-1"):
        u = make_user(email, role=UserRole.employee)
        e = make_employee(employee_code=code, user_id=u.id,
                          first_name="Test", last_name="User")
        p = make_project(code=proj_code, status=ProjectStatus.active)
        make_project_member(project_id=p.id, employee_id=e.id)
        return {"user": u, "emp": e, "project": p, "header": login(email)}

    return _make


@pytest.fixture()
def activity_admin(auth_header):
    return auth_header(email="pm@x.com", role=UserRole.project_manager)


@pytest.fixture()
def day_parts_on():
    """Split Day is feature-flagged and conftest pins it off for determinism."""
    prev = settings.REPORT_DAY_PARTS_ENABLED
    settings.REPORT_DAY_PARTS_ENABLED = True
    try:
        yield
    finally:
        settings.REPORT_DAY_PARTS_ENABLED = prev


def _make_sub(client, admin, *, value, name, count_field="tags"):
    a = client.post("/api/v1/activity-master/activities",
                    json={"name": f"Activity for {name}"}, headers=admin).json()
    return client.post(
        f"/api/v1/activity-master/activities/{a['id']}/sub-activities",
        json={"name": name, "benchmark_type": "NUMERIC_DAILY",
              "benchmark_value": value, "relevant_count_field": count_field},
        headers=admin,
    ).json()


def _prev_cycle():
    start, end = compute_week_bounds(date.today())
    return start - timedelta(days=7), end - timedelta(days=7)


def _submit_full_day(client, header, project_id, sub_id, on_date, qty, *, exception=None):
    task = {"project_id": str(project_id), "description": "w",
            "sub_activity_id": sub_id, "tags_count": qty}
    if exception:
        task["benchmark_exception_code"] = exception
    body = client.post(BASE, headers=header, json={
        "report_date": on_date.isoformat(), "tasks": [task],
    })
    assert body.status_code == 201, body.text
    body = body.json()
    assert client.post(f"{BASE}/{body['id']}/submit", headers=header).status_code == 200
    return body


def _sheet(client, admin):
    res = client.get(EXPORT_URL, headers=admin)
    assert res.status_code == 200
    return openpyxl.load_workbook(BytesIO(res.content)).active


def _detail(ws, sub_name, on_date):
    for r in range(3, ws.max_row + 1):
        cell = ws.cell(r, DATE_C).value
        value = cell.date() if hasattr(cell, "date") else cell
        if ws.cell(r, SUB).value == sub_name.upper() and value == on_date:
            return r
    raise AssertionError(f"detail row not found: {sub_name}")


def _total(ws, sub_name):
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, SUB).value == sub_name.upper() and ws.cell(r, PROJECT).value == "TOTAL":
            return r
    raise AssertionError(f"total row not found: {sub_name}")


def _fill(cell):
    return cell.fill.fgColor.rgb if cell.fill and cell.fill.fill_type else None


# --- 2. the whole-unit target reaches the Excel ------------------------------

def test_half_day_target_is_whole_in_the_export(
    client, setup_author, activity_admin, day_parts_on,
):
    """A 35-tag benchmark worked as a split day carries an 18 target per half in
    the workbook — never 17.5. The BENCHMARK TARGET cell stays numeric."""
    a = setup_author()
    sub = _make_sub(client, activity_admin, value=35, name="ODD-HALF")
    d = _prev_cycle()[0]
    created = client.post(BASE, headers=a["header"], json={
        "report_date": d.isoformat(),
        "report_mode": "split_day",
        "periods": [
            {"day_part": "first_half", "period_status": "work_at_office",
             "location": "chennai", "remarks": "am",
             "tasks": [{"project_id": str(a["project"].id), "description": "w",
                        "sub_activity_id": sub["id"], "tags_count": 10}]},
            {"day_part": "second_half", "period_status": "leave", "tasks": []},
        ],
    })
    assert created.status_code == 201, created.text
    assert client.post(
        f"{BASE}/{created.json()['id']}/submit", headers=a["header"]
    ).status_code == 200

    ws = _sheet(client, activity_admin)
    row = _detail(ws, "ODD-HALF", d)
    target = ws.cell(row, TGT_TAGS).value
    assert target == 18                       # 35 x 0.5 = 17.5 -> 18
    assert isinstance(target, (int, float)) and float(target).is_integer()
    # Pending follows the whole target: 18 - 10.
    assert ws.cell(row, PEN_TAGS).value == 8


def test_full_day_target_is_unchanged_by_the_rounding_rule(
    client, setup_author, activity_admin,
):
    a = setup_author()
    sub = _make_sub(client, activity_admin, value=35, name="ODD-FULL")
    d = _prev_cycle()[0]
    _submit_full_day(client, a["header"], a["project"].id, sub["id"], d, 10)

    ws = _sheet(client, activity_admin)
    assert ws.cell(_detail(ws, "ODD-FULL", d), TGT_TAGS).value == 35


def test_a_legacy_half_unit_snapshot_is_rounded_up_on_read(
    client, db, setup_author, activity_admin,
):
    """A report submitted BEFORE the whole-unit rule existed has a fractional
    effective target frozen on its row — a 65-tag benchmark halved is 32.5. The
    export must still read 33: the stored snapshot is left exactly as it was,
    only its presentation is made whole, so no historical report needs a
    migration or a resubmission.

    This is the reported case: "32.5" appearing in a BENCHMARK TARGET cell."""
    import uuid as uuid_mod

    from app.modules.work_reports.models import WorkReportTask

    a = setup_author()
    sub = _make_sub(client, activity_admin, value=65, name="LEGACY-HALF")
    d = _prev_cycle()[0]
    body = _submit_full_day(client, a["header"], a["project"].id, sub["id"], d, 30)

    # Reproduce a pre-rule snapshot: the fractional value the old code froze.
    row = db.get(WorkReportTask, uuid_mod.UUID(body["tasks"][0]["id"]))
    row.benchmark_value_snapshot = Decimal("32.5")
    row.benchmark_fraction_snapshot = Decimal("0.5")
    db.commit()

    ws = _sheet(client, activity_admin)
    detail = _detail(ws, "LEGACY-HALF", d)
    target = ws.cell(detail, TGT_TAGS).value
    assert target == 33                       # not 32.5
    assert isinstance(target, int)
    # Pending follows the whole target: 33 - 30.
    assert ws.cell(detail, PEN_TAGS).value == 3
    # …and the stored snapshot is untouched.
    db.refresh(row)
    assert row.benchmark_value_snapshot == Decimal("32.5")


def test_every_unit_cell_is_a_whole_number(client, setup_author, activity_admin):
    """Sheet-wide: no BENCHMARK TARGET / ACTUAL COMPLETED / PENDING cell ever
    carries a fractional part — they count real things."""
    a = setup_author()
    odd = _make_sub(client, activity_admin, value=35, name="ODD")
    even = _make_sub(client, activity_admin, value=100, name="EVEN")
    d1, _ = _prev_cycle()
    d2 = d1 + timedelta(days=1)
    _submit_full_day(client, a["header"], a["project"].id, odd["id"], d1, 17)
    _submit_full_day(client, a["header"], a["project"].id, even["id"], d2, 120)

    ws = _sheet(client, activity_admin)
    for r in range(3, ws.max_row + 1):
        for c in list(range(TGT_TAGS, TGT_TAGS + 6)) + \
                 list(range(ACT_TAGS, ACT_TAGS + 6)) + \
                 list(range(PEN_TAGS, PEN_TAGS + 6)):
            value = ws.cell(r, c).value
            if value is None:
                continue
            assert isinstance(value, int), ws.cell(r, c).coordinate


# --- 3. the red shortfall marker ---------------------------------------------

def test_actual_below_target_is_red(client, setup_author, activity_admin):
    """The reported case: target 100, actual 80, no exception -> the ACTUAL
    COMPLETED cell is red. Only that cell; the value under it is untouched."""
    a = setup_author()
    sub = _make_sub(client, activity_admin, value=100, name="SHORT")
    d = _prev_cycle()[0]
    _submit_full_day(client, a["header"], a["project"].id, sub["id"], d, 80)

    ws = _sheet(client, activity_admin)
    row = _detail(ws, "SHORT", d)
    cell = ws.cell(row, ACT_TAGS)
    assert _fill(cell) == SHORT_RED
    assert cell.font.color.rgb == SHORT_RED_FONT
    assert cell.font.bold is True
    for side in ("top", "bottom", "left", "right"):
        assert getattr(cell.border, side).style == "thin"
    assert cell.value == 80                    # value untouched
    # The target and the pending beside it stay unfilled.
    assert _fill(ws.cell(row, TGT_TAGS)) is None
    assert _fill(ws.cell(row, PEN_TAGS)) is None


def test_actual_at_or_above_target_is_not_red(client, setup_author, activity_admin):
    a = setup_author()
    exact = _make_sub(client, activity_admin, value=100, name="EXACT")
    over = _make_sub(client, activity_admin, value=100, name="OVER")
    d1, _ = _prev_cycle()
    d2 = d1 + timedelta(days=1)
    _submit_full_day(client, a["header"], a["project"].id, exact["id"], d1, 100)
    _submit_full_day(client, a["header"], a["project"].id, over["id"], d2, 140)

    ws = _sheet(client, activity_admin)
    assert _fill(ws.cell(_detail(ws, "EXACT", d1), ACT_TAGS)) is None
    assert _fill(ws.cell(_detail(ws, "OVER", d2), ACT_TAGS)) is None


def test_an_excepted_shortfall_is_amber_not_red(client, setup_author, activity_admin):
    """The two markers are mutually exclusive: an accepted exception reads amber
    even though the actual is below target."""
    a = setup_author()
    sub = _make_sub(client, activity_admin, value=100, name="EXCUSED")
    d = _prev_cycle()[0]
    _submit_full_day(client, a["header"], a["project"].id, sub["id"], d, 40,
                     exception=NO_FURTHER)

    ws = _sheet(client, activity_admin)
    cell = ws.cell(_detail(ws, "EXCUSED", d), ACT_TAGS)
    assert _fill(cell) == AMBER
    assert _fill(cell) != SHORT_RED


def test_total_row_actual_is_red_when_the_cycle_falls_short(
    client, setup_author, activity_admin,
):
    a = setup_author()
    sub = _make_sub(client, activity_admin, value=100, name="SHORT")
    d1, _ = _prev_cycle()
    d2 = d1 + timedelta(days=1)
    _submit_full_day(client, a["header"], a["project"].id, sub["id"], d1, 80)
    _submit_full_day(client, a["header"], a["project"].id, sub["id"], d2, 90)

    ws = _sheet(client, activity_admin)
    total = _total(ws, "SHORT")
    assert ws.cell(total, ACT_TAGS).value == 170      # real sum
    assert _fill(ws.cell(total, ACT_TAGS)) == SHORT_RED


def test_total_row_actual_is_clean_when_the_cycle_nets_out(
    client, setup_author, activity_admin,
):
    """A day's overachievement paying off another day's shortfall leaves the
    CYCLE whole, so the total's ACTUAL cell carries no marker even though one
    detail row is red."""
    a = setup_author()
    sub = _make_sub(client, activity_admin, value=100, name="NETS")
    d1, _ = _prev_cycle()
    d2 = d1 + timedelta(days=1)
    _submit_full_day(client, a["header"], a["project"].id, sub["id"], d1, 80)
    _submit_full_day(client, a["header"], a["project"].id, sub["id"], d2, 130)

    ws = _sheet(client, activity_admin)
    assert _fill(ws.cell(_detail(ws, "NETS", d1), ACT_TAGS)) == SHORT_RED
    assert _fill(ws.cell(_detail(ws, "NETS", d2), ACT_TAGS)) is None
    total = _total(ws, "NETS")
    assert ws.cell(total, ACT_TAGS).value == 210      # 210 vs a 200 target
    assert _fill(ws.cell(total, ACT_TAGS)) is None
