"""Weekly Activity Report export — a Split Day's NO-ACTIVITY half still gets a row.

A Split Day records TWO reporting periods. When one of them is a working half
and the other is Leave / Week Off / Company Holiday, only the working half owns
a task line, so flattening the day's `activities` alone exported ONE row and the
other half disappeared — even though the Report Detail page shows both.

The fix renders the recorded period, not an invented activity: a status-only row
carrying its own DAY STATUS and HALF, with project / activity / sub-activity /
counts / benchmark left empty. It is a reporting PERIOD, not an activity, so it
never enters `activities`, `max_activities`, or the two-activity-per-report rule.

Phase B's Full-Day rules are pinned here too, so this can't leak into them:
FULL DAY for one activity, ACTIVITY 1 / ACTIVITY 2 for two, and FIRST/SECOND
HALF reserved for a genuine Split Day.
"""
from datetime import date
from io import BytesIO

import openpyxl
import pytest

from app.modules.projects.models import ProjectStatus
from app.modules.reports_export import export
from app.modules.users.models import UserRole

BASE = "/api/v1/work-reports"
ROWS = "/api/v1/reports-export/activity-rows"
XLSX = "/api/v1/reports-export/activity-rows.xlsx"
MONDAY = date(2026, 8, 24)

# The 16 columns, in order — the export's published contract.
EXPECTED_HEADERS = [
    "EMPLOYEE ID & NAME", "DATE", "DAY", "DAY STATUS", "HALF", "PROJECT CODE",
    "ACTIVITY TYPE", "SUB ACTIVITY TYPE", "NO. OF TAGS", "NO. OF DOCS",
    "NO. OF BOM HEADER", "NO. OF SPARES", "NO. OF PAGES", "NO. OF RECORDS",
    "BENCHMARK", "DAY REMARKS",
]
# Everything a status-only period row must leave alone.
ACTIVITY_COLUMNS = [
    "PROJECT CODE", "ACTIVITY TYPE", "SUB ACTIVITY TYPE", "BENCHMARK",
    "NO. OF TAGS", "NO. OF DOCS", "NO. OF BOM HEADER", "NO. OF SPARES",
    "NO. OF PAGES", "NO. OF RECORDS",
]


# ── pure export layer ────────────────────────────────────────────────────────


def _activity(**over):
    base = {
        "day_part": "full_day",
        "period_status": None,
        "project_code": "NAAA",
        "activity_type": "TRAINING",
        "sub_activity_type": "PM-FAMILIARIZATION",
        "tags": 0, "docs": 0, "bom": 0, "spares": 0, "pages": 0, "records": 0,
        "benchmark_type": None,
        "benchmark_value": None,
        "benchmark_unit": None,
    }
    base.update(over)
    return base


def _day(activities=(), status_periods=(), **over):
    row = {
        "employee_label": "CDC019 - Karthikeyan K",
        "report_date": MONDAY,
        "day_status": "Work at Office",
        "remarks": None,
        "activities": list(activities),
        "status_periods": list(status_periods),
    }
    row.update(over)
    return row


def _sheet(rows):
    return openpyxl.load_workbook(export.build_workbook(rows)).active


def _cells(ws):
    """{(row, HEADER): value} — the sheet addressed by column name."""
    headers = {c.column: c.value for c in ws[1]}
    return {
        (c.row, headers[c.column]): c.value
        for row in ws.iter_rows(min_row=2)
        for c in row
    }


def test_day_entries_orders_first_half_before_second_half():
    """Ordering comes from the recorded period, never from which list the entry
    arrived in — a Leave FIRST HALF still leads the day."""
    activities = [_activity(day_part="second_half")]
    periods = [{"day_part": "first_half", "period_status": "Leave"}]
    assert [
        (e.get("day_part"), status_only)
        for e, _i, status_only in export.day_entries(activities, periods)
    ] == [("first_half", True), ("second_half", False)]


def test_day_entries_without_status_periods_is_the_untouched_old_order():
    activities = [_activity(day_part="full_day"), _activity(day_part="full_day")]
    assert export.day_entries(activities, []) == [
        (activities[0], 0, False), (activities[1], 1, False),
    ]


# CASE 2 — activity in the First Half, Leave in the Second.
def test_split_day_leave_second_half_gets_its_own_row():
    ws = _sheet([_day(
        [_activity(day_part="first_half", period_status="Work at Office")],
        [{"day_part": "second_half", "period_status": "Leave"}],
    )])
    c = _cells(ws)
    assert ws.max_row == 3
    assert c[(2, "HALF")] == "FIRST HALF"
    assert c[(2, "DAY STATUS")] == "WORK AT OFFICE"
    assert c[(2, "PROJECT CODE")] == "NAAA"
    assert c[(2, "ACTIVITY TYPE")] == "TRAINING"
    assert c[(2, "SUB ACTIVITY TYPE")] == "PM-FAMILIARIZATION"
    assert c[(3, "HALF")] == "SECOND HALF"
    assert c[(3, "DAY STATUS")] == "LEAVE"


# CASE 3 — the mirror image: Leave first, work second.
def test_split_day_leave_first_half_gets_its_own_row_and_leads():
    ws = _sheet([_day(
        [_activity(day_part="second_half", period_status="Work at Office")],
        [{"day_part": "first_half", "period_status": "Leave"}],
    )])
    c = _cells(ws)
    assert ws.max_row == 3
    assert (c[(2, "HALF")], c[(2, "DAY STATUS")]) == ("FIRST HALF", "LEAVE")
    assert (c[(3, "HALF")], c[(3, "DAY STATUS")]) == ("SECOND HALF", "WORK AT OFFICE")
    assert c[(3, "PROJECT CODE")] == "NAAA"


def test_status_only_row_invents_no_project_activity_benchmark_or_count():
    """The row is a reporting period, not work: a zero in a count column would
    read as 'nothing was produced against a target', which is a claim this
    report never made."""
    ws = _sheet([_day(
        [_activity(day_part="first_half", period_status="Work at Office", tags=120,
                   benchmark_type="NUMERIC_DAILY", benchmark_value=120,
                   benchmark_unit="tags")],
        [{"day_part": "second_half", "period_status": "Leave"}],
    )])
    c = _cells(ws)
    for header in ACTIVITY_COLUMNS:
        assert c[(3, header)] is None, header
    # The activity-bearing half is untouched by its neighbour.
    assert c[(2, "NO. OF TAGS")] == 120
    assert c[(2, "BENCHMARK")] == "120 TAGS"


# CASE 4 — any existing non-work status, not just Leave.
@pytest.mark.parametrize("status", ["Week Off", "Company Holiday", "Comp-off"])
def test_other_non_work_statuses_use_the_same_representation(status):
    ws = _sheet([_day(
        [_activity(day_part="first_half", period_status="Work at Office")],
        [{"day_part": "second_half", "period_status": status}],
    )])
    c = _cells(ws)
    assert c[(3, "HALF")] == "SECOND HALF"
    assert c[(3, "DAY STATUS")] == status.upper()


def test_both_halves_no_activity_still_shows_two_periods():
    ws = _sheet([_day([], [
        {"day_part": "first_half", "period_status": "Leave"},
        {"day_part": "second_half", "period_status": "Week Off"},
    ], day_status=None)])
    c = _cells(ws)
    assert ws.max_row == 3
    assert [c[(r, "HALF")] for r in (2, 3)] == ["FIRST HALF", "SECOND HALF"]
    assert [c[(r, "DAY STATUS")] for r in (2, 3)] == ["LEAVE", "WEEK OFF"]


def test_day_identity_still_merges_across_a_status_period_row():
    ws = _sheet([_day(
        [_activity(day_part="first_half", period_status="Work at Office")],
        [{"day_part": "second_half", "period_status": "Leave"}],
        remarks="First Half: idb work\nSecond Half: on leave",
    )])
    merged = {str(r) for r in ws.merged_cells.ranges}
    # Employee / Date / Day / Day Remarks span the day; DAY STATUS differs
    # between the halves, so it stays two separately filterable cells.
    assert {"A2:A3", "B2:B3", "C2:C3", "P2:P3"} <= merged
    assert "D2:D3" not in merged


def test_autofilter_covers_the_added_period_row():
    ws = _sheet([_day(
        [_activity(day_part="first_half")],
        [{"day_part": "second_half", "period_status": "Leave"}],
    )])
    assert ws.auto_filter.ref == "A1:P3"


def test_column_names_and_order_are_unchanged():
    ws = _sheet([_day(
        [_activity(day_part="first_half")],
        [{"day_part": "second_half", "period_status": "Leave"}],
    )])
    assert [c.value for c in ws[1]] == EXPECTED_HEADERS


# ── Phase B regressions — none of the above may touch these ──────────────────


def test_full_day_one_activity_still_reads_full_day():
    ws = _sheet([_day([_activity(day_part="full_day")])])
    assert ws.max_row == 2
    assert _cells(ws)[(2, "HALF")] == "FULL DAY"


def test_full_day_two_activities_still_read_activity_1_and_2():
    ws = _sheet([_day([
        _activity(day_part="full_day", project_code="PROJECT A"),
        _activity(day_part="full_day", project_code="PROJECT B"),
    ])])
    c = _cells(ws)
    assert [c[(r, "HALF")] for r in (2, 3)] == ["ACTIVITY 1", "ACTIVITY 2"]
    assert "FIRST HALF" not in {c[(2, "HALF")], c[(3, "HALF")]}
    assert "SECOND HALF" not in {c[(2, "HALF")], c[(3, "HALF")]}


def test_split_day_with_two_activities_is_unchanged():
    ws = _sheet([_day([
        _activity(day_part="first_half", project_code="PROJECT A"),
        _activity(day_part="second_half", project_code="PROJECT B"),
    ])])
    c = _cells(ws)
    assert ws.max_row == 3
    assert [c[(r, "HALF")] for r in (2, 3)] == ["FIRST HALF", "SECOND HALF"]
    assert [c[(r, "PROJECT CODE")] for r in (2, 3)] == ["PROJECT A", "PROJECT B"]


def test_ls_and_numeric_benchmarks_still_render_beside_a_status_period():
    """A status-only half must not disturb the benchmark of the half that has
    one — LS stays LS and a numeric target stays numeric."""
    ls = _sheet([_day(
        [_activity(day_part="first_half", benchmark_type="TASK_STATUS_ONLY")],
        [{"day_part": "second_half", "period_status": "Leave"}],
    )])
    assert _cells(ls)[(2, "BENCHMARK")] == "LS"
    numeric = _sheet([_day(
        [_activity(day_part="first_half", benchmark_type="NUMERIC_DAILY",
                   benchmark_value=250, benchmark_unit="tags")],
        [{"day_part": "second_half", "period_status": "Leave"}],
    )])
    assert _cells(numeric)[(2, "BENCHMARK")] == "250 TAGS"
    lsq = _sheet([_day([_activity(
        day_part="first_half", benchmark_type="TASK_WITH_QUANTITY",
        benchmark_value=300, benchmark_unit="spares",
    )], [{"day_part": "second_half", "period_status": "Leave"}])])
    assert _cells(lsq)[(2, "BENCHMARK")] == "LS - 300 SPARES"


def test_whole_leave_day_row_is_unchanged():
    """A full-day leave report has no periods to render and keeps its existing
    single FULL DAY row with the status echoed into ACTIVITY TYPE."""
    ws = _sheet([_day(day_status="Leave")])
    c = _cells(ws)
    assert ws.max_row == 2
    assert (c[(2, "HALF")], c[(2, "DAY STATUS")]) == ("FULL DAY", "LEAVE")
    assert c[(2, "ACTIVITY TYPE")] == "LEAVE"


def test_rows_without_the_new_key_still_build():
    """Every other caller hands in rows that have no `status_periods` at all."""
    ws = _sheet([{
        "employee_label": "CDC019 - Karthikeyan K",
        "report_date": MONDAY,
        "day_status": "Work at Office",
        "remarks": None,
        "activities": [_activity()],
    }])
    assert ws.max_row == 2
    assert _cells(ws)[(2, "HALF")] == "FULL DAY"


# ── end to end, through the real report + export ─────────────────────────────


@pytest.fixture()
def day_parts_on():
    from app.core.config import settings

    prev = settings.REPORT_DAY_PARTS_ENABLED
    settings.REPORT_DAY_PARTS_ENABLED = True
    try:
        yield
    finally:
        settings.REPORT_DAY_PARTS_ENABLED = prev


@pytest.fixture()
def setup_author(make_user, make_employee, make_project, make_project_member, login):
    def _make(*, email="emp@x.com", code="CDC019", proj_code="NAAA"):
        u = make_user(email, role=UserRole.employee)
        e = make_employee(employee_code=code, user_id=u.id)
        p = make_project(code=proj_code, status=ProjectStatus.active)
        make_project_member(project_id=p.id, employee_id=e.id)
        return {"user": u, "emp": e, "project": p, "header": login(email)}

    return _make


@pytest.fixture()
def pm_header(auth_header):
    return auth_header(email="pm@x.com", role=UserRole.project_manager)


def _working(project_id):
    return {
        "period_status": "work_at_office",
        "location": "chennai",
        "tasks": [{"project_id": str(project_id), "description": "work"}],
    }


def _create_split(client, header, first, second, report_date=None):
    res = client.post(BASE, headers=header, json={
        "report_date": (report_date or date.today()).isoformat(),
        "report_mode": "split_day",
        "periods": [
            {"day_part": "first_half", **first},
            {"day_part": "second_half", **second},
        ],
    })
    assert res.status_code == 201, res.text
    created = res.json()
    submitted = client.post(f"{BASE}/{created['id']}/submit", headers=header)
    assert submitted.status_code == 200, submitted.text
    return created


def _export_cells(client, header):
    res = client.get(XLSX, headers=header)
    assert res.status_code == 200, res.text
    ws = openpyxl.load_workbook(BytesIO(res.content)).active
    return ws, _cells(ws)


def test_e2e_split_day_work_then_leave_exports_both_halves(
    client, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    _create_split(
        client, a["header"], _working(a["project"].id), {"period_status": "leave"}
    )
    ws, c = _export_cells(client, pm_header)
    assert ws.max_row == 3
    assert c[(2, "HALF")] == "FIRST HALF"
    assert c[(2, "DAY STATUS")] == "WORK AT OFFICE"
    assert c[(2, "PROJECT CODE")] == "NAAA"
    assert c[(3, "HALF")] == "SECOND HALF"
    assert c[(3, "DAY STATUS")] == "LEAVE"
    for header in ACTIVITY_COLUMNS:
        assert c[(3, header)] is None, header


def test_e2e_split_day_leave_then_work_exports_both_halves(
    client, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    _create_split(
        client, a["header"], {"period_status": "leave"}, _working(a["project"].id)
    )
    ws, c = _export_cells(client, pm_header)
    assert ws.max_row == 3
    assert (c[(2, "HALF")], c[(2, "DAY STATUS")]) == ("FIRST HALF", "LEAVE")
    assert (c[(3, "HALF")], c[(3, "DAY STATUS")]) == ("SECOND HALF", "WORK AT OFFICE")
    assert c[(3, "PROJECT CODE")] == "NAAA"


def test_e2e_split_day_week_off_half_exports_its_own_row(
    client, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    _create_split(
        client, a["header"], _working(a["project"].id), {"period_status": "week_off"}
    )
    _ws, c = _export_cells(client, pm_header)
    assert (c[(3, "HALF")], c[(3, "DAY STATUS")]) == ("SECOND HALF", "WEEK OFF")


def test_e2e_leave_half_is_not_an_activity_in_the_preview(
    client, setup_author, pm_header, day_parts_on
):
    """The JSON preview and the two-activity rule count ACTIVITIES. A Leave half
    is a period, so the day still reports exactly one activity."""
    a = setup_author()
    _create_split(
        client, a["header"], _working(a["project"].id), {"period_status": "leave"}
    )
    body = client.get(ROWS, headers=pm_header).json()
    assert body["max_activities"] == 1
    assert len(body["rows"]) == 1
    assert len(body["rows"][0]["activities"]) == 1


def test_e2e_split_day_with_both_halves_worked_is_unchanged(
    client, setup_author, pm_header, day_parts_on
):
    a = setup_author()
    pid = a["project"].id
    _create_split(client, a["header"], _working(pid), _working(pid))
    ws, c = _export_cells(client, pm_header)
    assert ws.max_row == 3
    assert [c[(r, "HALF")] for r in (2, 3)] == ["FIRST HALF", "SECOND HALF"]
    assert [c[(r, "PROJECT CODE")] for r in (2, 3)] == ["NAAA", "NAAA"]


def test_e2e_full_day_report_is_unchanged(client, setup_author, pm_header):
    a = setup_author()
    pid = a["project"].id
    created = client.post(BASE, headers=a["header"], json={
        "report_date": date.today().isoformat(),
        "day_status": "work_at_office",
        "location": "chennai",
        "tasks": [
            {"project_id": str(pid), "description": "one"},
            {"project_id": str(pid), "description": "two"},
        ],
    }).json()
    client.post(f"{BASE}/{created['id']}/submit", headers=a["header"])
    ws, c = _export_cells(client, pm_header)
    assert ws.max_row == 3
    assert [c[(r, "HALF")] for r in (2, 3)] == ["ACTIVITY 1", "ACTIVITY 2"]


def test_e2e_full_day_leave_report_still_exports_one_full_day_row(
    client, setup_author, pm_header
):
    a = setup_author()
    created = client.post(BASE, headers=a["header"], json={
        "report_date": date.today().isoformat(),
        "day_status": "leave",
    }).json()
    client.post(f"{BASE}/{created['id']}/submit", headers=a["header"])
    ws, c = _export_cells(client, pm_header)
    assert ws.max_row == 2
    assert (c[(2, "HALF")], c[(2, "DAY STATUS")]) == ("FULL DAY", "LEAVE")
