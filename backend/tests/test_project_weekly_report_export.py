"""Project Weekly Report — the .xlsx download (Phase 7).

The rule this file exists to enforce is dataset parity: the workbook is a
RENDERING of the preview payload, not a second query. Both endpoints call
projects.service.get_project_weekly_report, so a Head can never see 25 rows on
screen and 27 in the file. The tests below check that by reading the workbook
back and comparing it cell-for-cell against the JSON the preview returned for
the same project and cycle.

Everything else here is about the file being genuinely usable: real Excel dates
and real numbers (so the Head can sort and filter), the project CODE alone in
the PROJECT column, long text preserved rather than truncated, and a filename
that names the cycle it came from.
"""
from datetime import date, datetime, timedelta
from io import BytesIO

import openpyxl
import pytest

from app.modules.activity_master.models import ActivityMaster
from app.modules.projects import weekly_report
from app.modules.projects.models import ProjectMember, ProjectStatus
from app.modules.users.models import UserRole

BASE = "/api/v1/projects"
REPORTS = "/api/v1/work-reports"
XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

TODAY = date.today()
LAST_CYCLE = TODAY - timedelta(days=7)

EXPECTED_HEADERS = [
    "SL.NO", "DATE", "WORK PERIOD", "EMPLOYEE NAME", "PROJECT", "ACTIVITY",
    "SUB-ACTIVITY", "BENCHMARK", "TAGS", "DOCS", "BOM", "SPARES", "PAGES",
    "RECORDS", "TASK STATUS", "REMARKS",
]

LONG_REMARK = (
    "FAHN MTL ASSET PHOTO POPULATION FOR 45 TAGS (BALANCE TAG-296), QC FOR AI "
    "DERIVED OUTPUT FROM ASSET PHOTO AND E-NAME PLATE DESIGN FIELD UPDATION"
)


@pytest.fixture()
def wr(db, make_project, make_user, make_employee, login):
    project = make_project(code="WR-XL", name="Excel Works", status=ProjectStatus.active)

    fmtl = ActivityMaster(name="FMTL", level="activity")
    meeting = ActivityMaster(name="PROJECT MEETING", level="activity")
    db.add_all([fmtl, meeting])
    db.commit()
    tags_sub = ActivityMaster(
        name="FMTL-TAG DESCRIPTION FROM P&ID", level="sub_activity", parent_id=fmtl.id,
        benchmark_type="NUMERIC_DAILY", benchmark_value=300, relevant_count_field="tags",
    )
    meeting_sub = ActivityMaster(
        name="PROJECT MEETING-MTL", level="sub_activity", parent_id=meeting.id,
    )
    db.add_all([tags_sub, meeting_sub])
    db.commit()

    def person(email, code, first, last):
        user = make_user(email, "password123", UserRole.employee)
        emp = make_employee(
            employee_code=code, first_name=first, last_name=last, user_id=user.id
        )
        db.add(ProjectMember(project_id=project.id, employee_id=emp.id))
        db.commit()
        return emp, login(email)

    head_emp, head = person("xl.head@x.com", "XL-H", "Hari", "Krishnan")
    _, alice = person("xl.alice@x.com", "XL-A", "Alice", "Anand")
    _, bob = person("xl.bob@x.com", "XL-B", "Bala", "Murugan")

    project.head_employee_id = head_emp.id
    db.add(project)
    db.commit()

    return {
        "project": project, "tags_sub": tags_sub, "meeting_sub": meeting_sub,
        "head": head, "alice": alice, "bob": bob,
    }


def _task(project, sub, **over):
    task = {
        "project_id": str(project.id),
        "sub_activity_id": str(sub.id),
        "minutes_spent": 240,
        "description": "worked on it",
    }
    task.update(over)
    return task


def _file(client, headers, day, tasks):
    res = client.post(REPORTS, json={
        "report_date": day.isoformat(), "day_status": "work_at_office", "tasks": tasks,
    }, headers=headers)
    assert res.status_code == 201, res.text
    res = client.post(f"{REPORTS}/{res.json()['id']}/submit", headers=headers)
    assert res.status_code == 200, res.text


def _preview(client, wr, cycle="current"):
    res = client.get(f"{BASE}/{wr['project'].id}/weekly-report?cycle={cycle}",
                     headers=wr["head"])
    assert res.status_code == 200, res.text
    return res.json()


def _download(client, wr, cycle="current"):
    res = client.get(f"{BASE}/{wr['project'].id}/weekly-report.xlsx?cycle={cycle}",
                     headers=wr["head"])
    assert res.status_code == 200, res.text
    return res


def _sheet(res):
    wb = openpyxl.load_workbook(BytesIO(res.content))
    return wb.active


def _values(ws):
    """Data rows only (row 1 is the header), as lists of cell values."""
    return [
        [c.value for c in row]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row)
        if any(c.value is not None for c in row)
    ]


@pytest.fixture()
def a_week(client, wr):
    """A representative week: two employees, two dates, three activity shapes,
    including one row whose remark is longer than its column."""
    _file(client, wr["alice"], LAST_CYCLE, [
        _task(wr["project"], wr["tags_sub"], tags_count=250, description=LONG_REMARK),
        _task(wr["project"], wr["meeting_sub"], description="weekly sync"),
    ])
    _file(client, wr["bob"], LAST_CYCLE, [
        _task(wr["project"], wr["tags_sub"], tags_count=69, description="FAHN-MTL 69 TAGS"),
    ])
    _file(client, wr["alice"], LAST_CYCLE + timedelta(days=2), [
        _task(wr["project"], wr["tags_sub"], tags_count=80, description="80 tags"),
    ])
    return wr


# ---------- the file itself -------------------------------------------------
def test_a_valid_workbook_is_returned(client, a_week):
    res = _download(client, a_week, "previous")
    assert res.headers["content-type"].startswith(XLSX_MEDIA)
    assert "attachment;" in res.headers["content-disposition"]
    ws = _sheet(res)
    assert ws.title == "Weekly Report"


def test_the_filename_is_the_project_number_and_the_week(client, a_week):
    res = _download(client, a_week, "previous")
    payload = _preview(client, a_week, "previous")
    start = date.fromisoformat(payload["period"]["start_date"])
    end = date.fromisoformat(payload["period"]["end_date"])
    expected = weekly_report.export_filename("4460-GC22104900", start, end)
    assert expected.startswith("4460 WEEKLY REPORT [")
    # The fixture project's code carries no leading number, so its own file
    # drops the prefix entirely rather than inventing one.
    assert 'filename="WEEKLY REPORT [' in res.headers["content-disposition"]
    # Neither the project NAME nor its full code reaches the filename.
    assert "Excel Works" not in res.headers["content-disposition"]


@pytest.mark.parametrize("code,expected", [
    ("4460-GC22104900", "4460 WEEKLY REPORT [7 AUG - 13 AUG].xlsx"),
    ("4716-LC25102900", "4716 WEEKLY REPORT [7 AUG - 13 AUG].xlsx"),
    # No code, or one that does not start with a number: no prefix, and no
    # stray leading space either.
    (None, "WEEKLY REPORT [7 AUG - 13 AUG].xlsx"),
    ("", "WEEKLY REPORT [7 AUG - 13 AUG].xlsx"),
    ("WR-XL", "WEEKLY REPORT [7 AUG - 13 AUG].xlsx"),
])
def test_the_filename_format(code, expected):
    assert weekly_report.export_filename(
        code, date(2026, 8, 7), date(2026, 8, 13)
    ) == expected


def test_the_filename_day_is_not_zero_padded():
    assert weekly_report.export_filename(
        "4460-X", date(2026, 8, 7), date(2026, 8, 13)
    ) == "4460 WEEKLY REPORT [7 AUG - 13 AUG].xlsx"


def test_the_filename_follows_the_selected_cycle(client, a_week):
    current = _download(client, a_week, "current").headers["content-disposition"]
    previous = _download(client, a_week, "previous").headers["content-disposition"]
    assert current != previous


def test_a_project_code_with_unsafe_characters_cannot_break_the_header():
    name = weekly_report.export_filename('44/6"0', date(2026, 8, 7), date(2026, 8, 13))
    assert '"' not in name and "/" not in name
    assert name == "44 WEEKLY REPORT [7 AUG - 13 AUG].xlsx"


def test_the_headers_are_the_agreed_columns(client, a_week):
    ws = _sheet(_download(client, a_week, "previous"))
    got = [ws.cell(row=1, column=i).value for i in range(1, len(EXPECTED_HEADERS) + 1)]
    assert got == EXPECTED_HEADERS
    assert ws.cell(row=1, column=len(EXPECTED_HEADERS) + 1).value is None


def test_the_header_is_styled_and_frozen(client, a_week):
    """Matches the company's own Head Report workbook: bold on yellow, with the
    header row frozen so a long week stays readable while scrolling."""
    ws = _sheet(_download(client, a_week, "previous"))
    head = ws.cell(row=1, column=1)
    assert head.font.bold is True
    assert head.fill.fgColor.rgb == "FFFFFF00"
    assert ws.freeze_panes == "A2"


# ---------- parity with the preview ----------------------------------------
def test_row_counts_match_the_preview(client, a_week):
    for cycle in ("current", "previous"):
        payload = _preview(client, a_week, cycle)
        rows = _values(_sheet(_download(client, a_week, cycle)))
        assert len(rows) == payload["row_count"], cycle
        assert len(rows) == len(payload["rows"]), cycle


def test_every_cell_matches_the_preview_row(client, a_week):
    """Column by column, in order, for the whole week. The only Excel-only
    column is SL.NO."""
    payload = _preview(client, a_week, "previous")
    rows = _values(_sheet(_download(client, a_week, "previous")))
    assert rows, "the fixture week must not be empty"

    for serial, (cells, row) in enumerate(zip(rows, payload["rows"]), start=1):
        assert cells[0] == serial
        assert cells[1].date().isoformat() == row["report_date"]
        assert cells[2] == row["work_period_label"]
        assert cells[3] == row["employee_name"]
        assert cells[4] == row["project_code"]
        assert cells[5] == row["activity_name"]
        assert cells[6] == row["sub_activity_name"]
        expected_benchmark = (
            row["benchmark"] if row["benchmark"] is not None
            else (row["benchmark_label"] or "-")
        )
        assert cells[7] == expected_benchmark
        for offset, unit in enumerate(
            ["tags", "docs", "bom", "spares", "pages", "records"], start=8
        ):
            assert cells[offset] == (row[unit] if row[unit] is not None else "-"), unit
        assert cells[14] == (row["task_status_label"] or "-")
        assert cells[15] == (row["remarks"] or "-")


def test_the_export_ordering_is_the_preview_ordering(client, a_week):
    payload = _preview(client, a_week, "previous")
    rows = _values(_sheet(_download(client, a_week, "previous")))
    assert [(r[1].date().isoformat(), r[3]) for r in rows] == [
        (r["report_date"], r["employee_name"]) for r in payload["rows"]
    ]


def test_the_serial_numbers_run_from_one(client, a_week):
    rows = _values(_sheet(_download(client, a_week, "previous")))
    assert [r[0] for r in rows] == list(range(1, len(rows) + 1))


def test_current_and_previous_produce_different_files(client, wr):
    _file(client, wr["alice"], TODAY, [_task(wr["project"], wr["tags_sub"], tags_count=11)])
    _file(client, wr["alice"], LAST_CYCLE,
          [_task(wr["project"], wr["tags_sub"], tags_count=22)])

    current = _values(_sheet(_download(client, wr, "current")))
    previous = _values(_sheet(_download(client, wr, "previous")))
    assert [r[8] for r in current] == [11]
    assert [r[8] for r in previous] == [22]


# ---------- cell types ------------------------------------------------------
def test_dates_are_real_excel_dates(client, a_week):
    ws = _sheet(_download(client, a_week, "previous"))
    cell = ws.cell(row=2, column=2)
    # openpyxl reads a real date cell back as a datetime, never as a string.
    assert isinstance(cell.value, datetime)
    assert cell.number_format == "dd-mm-yyyy"


def test_counts_and_benchmarks_stay_numeric(client, a_week):
    ws = _sheet(_download(client, a_week, "previous"))
    numeric = [
        (r[7], r[8]) for r in _values(ws) if r[8] != "-"
    ]
    assert numeric, "the fixture week must contain a counted row"
    for benchmark, tags in numeric:
        assert isinstance(tags, (int, float)) and not isinstance(tags, bool)
        assert isinstance(benchmark, (int, float)) and not isinstance(benchmark, bool)


def test_an_inapplicable_cell_is_a_dash_not_a_zero(client, wr):
    """The distinction matters most in Excel, where a 0 would be summed."""
    _file(client, wr["alice"], TODAY, [_task(wr["project"], wr["meeting_sub"])])

    row = _values(_sheet(_download(client, wr, "current")))[0]
    assert row[7] == "-"                       # BENCHMARK
    assert row[8:14] == ["-"] * 6              # the six unit columns
    assert row[14] == "-"                      # TASK STATUS
    assert 0 not in row[7:15]


def test_the_project_column_holds_the_code_only(client, a_week):
    rows = _values(_sheet(_download(client, a_week, "previous")))
    assert {r[4] for r in rows} == {"WR-XL"}
    assert not any("Excel Works" in str(cell) for row in rows for cell in row)


def test_long_text_is_preserved_in_full(client, a_week):
    ws = _sheet(_download(client, a_week, "previous"))
    remarks = [r[15] for r in _values(ws)]
    assert LONG_REMARK in remarks
    # Wrapped rather than clipped, and the sub-activity column too.
    long_row = next(i for i, r in enumerate(_values(ws), start=2) if r[15] == LONG_REMARK)
    assert ws.cell(row=long_row, column=16).alignment.wrap_text is True
    assert ws.cell(row=long_row, column=7).alignment.wrap_text is True
    assert ws.cell(row=long_row, column=7).value == "FMTL-TAG DESCRIPTION FROM P&ID"


# ---------- the empty week --------------------------------------------------
def test_an_empty_week_still_downloads_a_valid_header_only_workbook(client, wr):
    """The established CoreOps export behaviour: a workbook is always produced.
    The UI keeps the button disabled when there is nothing to download, so this
    is a fallback rather than the normal path."""
    res = _download(client, wr, "previous")
    ws = _sheet(res)
    assert [ws.cell(row=1, column=i).value for i in range(1, 17)] == EXPECTED_HEADERS
    assert _values(ws) == []
    assert _preview(client, wr, "previous")["row_count"] == 0
