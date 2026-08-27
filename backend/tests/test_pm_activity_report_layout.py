"""PM Weekly Activity Report — the one-row-per-activity Excel layout.

The old export laid a day's second activity out HORIZONTALLY ("Project Code 2",
"No. of Tags 2" …), which no Excel filter or pivot could reach. This proves the
replacement: one header row, one row per logical activity — labelled from the
RECORDED PERIOD, so a Full Day report holding two activities reads ACTIVITY 1 /
ACTIVITY 2 and only a genuine Split Day reads FIRST HALF / SECOND HALF — the
day's shared identity merged vertically WITHOUT losing the value the filter
needs, and the three benchmark shapes — "120 TAGS" / "LS" / "LS - 300 SPARES" —
rendered from the snapshot frozen on the task.

Nothing here recomputes a benchmark: `benchmark_display` reads benchmark_type /
benchmark_value / benchmark_unit exactly as work_report_tasks froze them, and
`_export_benchmark` only decides WHICH classification reaches it — the frozen
snapshot, or the sub-activity's own on a row that never went through a submit.
"""
from datetime import date
from decimal import Decimal
from types import SimpleNamespace

import openpyxl
import pytest

from app.modules.reports_export import export
from app.modules.work_reports.service import _export_benchmark

MONDAY = date(2026, 8, 17)


def _activity(**over):
    base = {
        "day_part": "full_day",
        "period_status": None,
        "project_code": "P-1",
        "activity_type": "MTL",
        "sub_activity_type": "MTL-DATA POPULATION",
        "tags": 0, "docs": 0, "bom": 0, "spares": 0, "pages": 0, "records": 0,
        "benchmark_type": None,
        "benchmark_value": None,
        "benchmark_unit": None,
    }
    base.update(over)
    return base


def _day(*activities, **over):
    row = {
        "employee_label": "CDC019 - Arthi S",
        "report_date": MONDAY,
        "day_status": "Work at Office",
        "remarks": "worked on the idb",
        "activities": list(activities),
    }
    row.update(over)
    return row


def _sheet(rows):
    return openpyxl.load_workbook(export.build_workbook(rows)).active


def _cells(ws):
    """{(row, HEADER): value} — the sheet addressed by column name."""
    headers = {c.column: c.value for c in ws[1]}
    return {
        (c.row, headers[c.column]): c.value
        for row in ws.iter_rows(min_row=2)
        for c in row
    }


# ── benchmark rendering (pure) ───────────────────────────────────────────────


@pytest.mark.parametrize(
    ("benchmark_type", "value", "unit", "expected"),
    [
        # CASE 1 — count-based: the target with its own unit. Every unit, so a
        # counted activity can never be mistaken for a lumpsum one.
        ("NUMERIC_DAILY", 120, "tags", "120 TAGS"),
        ("NUMERIC", 176, "spares", "176 SPARES"),
        ("NUMERIC_DAILY", 250, "tags", "250 TAGS"),
        ("NUMERIC_DAILY", 100, "docs", "100 DOCS"),
        ("NUMERIC_DAILY", 50, "bom", "50 BOM"),
        ("NUMERIC_DAILY", 200, "records", "200 RECORDS"),
        ("NUMERIC_DAILY", 300, "pages", "300 PAGES"),
        # CASE 2 — lumpsum with no quantity at all.
        ("TASK_STATUS_ONLY", None, None, "LS"),
        ("TASK_BASED", None, None, "LS"),
        # CASE 3 — lumpsum that DOES carry a quantity: the count must survive.
        ("TASK_WITH_QUANTITY", 300, "spares", "LS - 300 SPARES"),
        ("TASK_WITH_QUANTITY", 120, "tags", "LS - 120 TAGS"),
        ("TASK_WITH_QUANTITY", 50, "records", "LS - 50 RECORDS"),
        # No benchmark configured (e.g. LEAVE, TRAINING) — an empty cell.
        (None, None, None, None),
    ],
)
def test_benchmark_display_shapes(benchmark_type, value, unit, expected):
    assert export.benchmark_display(
        _activity(benchmark_type=benchmark_type, benchmark_value=value, benchmark_unit=unit)
    ) == expected


def test_lumpsum_with_a_quantity_never_collapses_to_bare_ls():
    """The requirement that is easiest to lose: an LS activity that has a count
    must print the count, in exactly 'LS - <COUNT> <UNIT>' form."""
    out = export.benchmark_display(
        _activity(benchmark_type="TASK_WITH_QUANTITY", benchmark_value=300, benchmark_unit="spares")
    )
    assert out == "LS - 300 SPARES"
    assert out != "LS"
    assert out not in ("LS 300", "300 LS", "Lumpsum - 300 Spares")


# ── benchmark CLASSIFICATION (pure) ──────────────────────────────────────────
#
# Which mode/target/unit reaches benchmark_display above. The snapshot frozen at
# submit wins; a row that never went through a submit — an approved LS
# continuation added to an already-submitted report, an edit-granted report
# still open, a legacy row predating the snapshot columns — falls back to the
# sub-activity's OWN Activity Master classification instead of exporting blank.


FULL_DAY = Decimal("1.0")
HALF_DAY = Decimal("0.5")


def _task(mode=None, value=None, unit=None):
    return SimpleNamespace(
        benchmark_type_snapshot=mode,
        benchmark_value_snapshot=value,
        relevant_count_field_snapshot=unit,
    )


def _master(mode=None, value=None, unit=None):
    return SimpleNamespace(
        benchmark_type=mode, benchmark_value=value, relevant_count_field=unit
    )


def test_frozen_snapshot_always_wins():
    """The value the day was actually measured against is never second-guessed,
    even when the master has since been reconfigured underneath it."""
    task = _task("NUMERIC_DAILY", 120, "tags")
    reconfigured = _master("TASK_STATUS_ONLY", None, None)
    assert _export_benchmark(task, reconfigured, FULL_DAY) == (
        "NUMERIC_DAILY", 120, "tags",
    )


def test_unsubmitted_lumpsum_row_still_classifies_as_ls():
    """The Phase B defect: an LS row with no frozen snapshot exported an EMPTY
    benchmark cell, so one report could show the same lumpsum sub-activity as
    'LS' on one line and blank on the next. Classification comes from the
    activity — not from approval, continuation or due-date state."""
    resolved = _export_benchmark(_task(), _master("TASK_STATUS_ONLY"), FULL_DAY)
    assert export.benchmark_display(dict(zip(
        ("benchmark_type", "benchmark_value", "benchmark_unit"), resolved
    ))) == "LS"


def test_unsubmitted_lumpsum_keeps_the_employees_own_unit():
    """A lumpsum row's unit is the EMPLOYEE's choice, stored at save time. The
    fallback must not overwrite it with the master's (which a lumpsum has none
    of anyway)."""
    task = _task(unit="pages")
    assert _export_benchmark(task, _master("TASK_BASED"), FULL_DAY) == (
        "TASK_BASED", None, "pages",
    )


def test_unsubmitted_counted_row_keeps_its_number_not_ls():
    """The mirror-image requirement: a counted activity must NOT become LS just
    because its snapshot is missing."""
    task, master = _task(), _master("NUMERIC_DAILY", 250, "tags")
    assert _export_benchmark(task, master, FULL_DAY) == ("NUMERIC_DAILY", 250, "tags")


def test_fallback_scales_a_half_day_target_the_way_submit_would():
    """Same rule as _apply_benchmarks (whole units), so the fallback and the
    snapshot can never disagree about a half-day row's effective target."""
    task, master = _task(), _master("NUMERIC_DAILY", 35, "tags")
    assert _export_benchmark(task, master, HALF_DAY) == ("NUMERIC_DAILY", 18, "tags")


def test_no_benchmark_stays_blank():
    """A sub-activity with no benchmark configured (LEAVE, TRAINING), and a row
    naming no sub-activity at all, both keep the existing empty cell — nothing
    is invented for them."""
    assert _export_benchmark(_task(), _master(), FULL_DAY) == (None, None, None)
    assert _export_benchmark(_task(), None, FULL_DAY) == (None, None, None)


def test_benchmark_quantity_is_whole_where_it_is_whole():
    """A 300 target reads '300 SPARES', never '300.0 SPARES'. A genuinely
    fractional stored target is shown as it is rather than silently rounded."""
    whole = _activity(benchmark_type="NUMERIC_DAILY", benchmark_value=300.0, benchmark_unit="spares")
    assert export.benchmark_display(whole) == "300 SPARES"
    half = _activity(benchmark_type="NUMERIC_DAILY", benchmark_value=32.5, benchmark_unit="tags")
    assert export.benchmark_display(half) == "32.5 TAGS"


# ── HALF labelling ───────────────────────────────────────────────────────────


def test_half_label_follows_the_recorded_period_when_there_is_one():
    # Two tasks logged in the SAME recorded half both read FIRST HALF — the
    # period is authoritative, so nothing is renumbered behind the employee.
    assert export.half_label("first_half", 0, 2) == "FIRST HALF"
    assert export.half_label("first_half", 1, 2) == "FIRST HALF"
    assert export.half_label("second_half", 1, 2) == "SECOND HALF"


def test_full_day_activities_are_numbered_never_halved():
    """A Full Day report is a Full Day report however many activities it holds.

    The two-activity rule lets an employee put TWO activities in one Full Day
    report; that is two activities, not two halves. Labelling them FIRST/SECOND
    HALF would claim a Split Day report the employee never filed, so the count
    of activities never decides the day format."""
    assert export.half_label("full_day", 0, 1) == "FULL DAY"
    assert export.half_label(None, 0, 1) == "FULL DAY"      # legacy, no period
    assert export.half_label("full_day", 0, 2) == "ACTIVITY 1"
    assert export.half_label("full_day", 1, 2) == "ACTIVITY 2"
    assert export.half_label(None, 1, 2) == "ACTIVITY 2"    # legacy, no period
    assert export.half_label("full_day", 2, 3) == "ACTIVITY 3"
    for index in (0, 1):
        assert export.half_label("full_day", index, 2) not in (
            "FIRST HALF", "SECOND HALF",
        )


# ── sheet structure ──────────────────────────────────────────────────────────


def test_single_header_row_with_autofilter_and_frozen_panes():
    ws = _sheet([_day(_activity())])
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:P2"
    # Exactly one header row: no employee banner, no repeat further down.
    assert ws.cell(1, 1).value == "EMPLOYEE ID & NAME"
    assert ws.cell(2, 1).value != "EMPLOYEE ID & NAME"


def test_two_activities_become_two_rows_first_then_second_half():
    """CASE 5 + CASE 6 — the structural change. Two activities on one day are
    two rows, each keeping its own project, so both stay filterable."""
    ws = _sheet([_day(
        _activity(day_part="first_half", project_code="PROJECT A",
                  sub_activity_type="TAGGING", tags=120,
                  benchmark_type="NUMERIC_DAILY", benchmark_value=120, benchmark_unit="tags"),
        _activity(day_part="second_half", project_code="PROJECT B",
                  sub_activity_type="SUPPORT", benchmark_type="TASK_STATUS_ONLY"),
    )])
    c = _cells(ws)
    assert ws.max_row == 3
    assert c[(2, "HALF")] == "FIRST HALF"
    assert c[(3, "HALF")] == "SECOND HALF"
    assert c[(2, "PROJECT CODE")] == "PROJECT A"
    assert c[(3, "PROJECT CODE")] == "PROJECT B"
    assert c[(2, "BENCHMARK")] == "120 TAGS"
    assert c[(3, "BENCHMARK")] == "LS"
    assert c[(2, "NO. OF TAGS")] == 120
    assert c[(3, "NO. OF TAGS")] == 0


def test_one_activity_produces_one_row_and_no_blank_half():
    ws = _sheet([_day(_activity())])
    assert ws.max_row == 2
    assert _cells(ws)[(2, "HALF")] == "FULL DAY"


def test_full_day_with_two_activities_is_two_activities_not_two_halves():
    """The Full Day / Split Day distinction, end to end.

    Both rows carry day_part 'full_day' — the employee filed ONE Full Day report
    holding two activities, which the two-activity rule allows. The sheet must
    say so: two numbered activity rows, and the words FIRST HALF / SECOND HALF
    nowhere on the day."""
    ws = _sheet([_day(
        _activity(day_part="full_day", project_code="PROJECT A",
                  sub_activity_type="TAGGING"),
        _activity(day_part="full_day", project_code="PROJECT B",
                  sub_activity_type="SUPPORT"),
    )])
    c = _cells(ws)
    assert ws.max_row == 3
    assert c[(2, "HALF")] == "ACTIVITY 1"
    assert c[(3, "HALF")] == "ACTIVITY 2"
    assert "FIRST HALF" not in {c[(2, "HALF")], c[(3, "HALF")]}
    assert "SECOND HALF" not in {c[(2, "HALF")], c[(3, "HALF")]}
    # Each activity still keeps its own identity columns.
    assert c[(2, "PROJECT CODE")] == "PROJECT A"
    assert c[(3, "PROJECT CODE")] == "PROJECT B"


def test_split_day_keeps_first_and_second_half_unchanged():
    """The other side of the same rule: a GENUINE Split Day still reads FIRST
    HALF / SECOND HALF, because the recorded period says so."""
    ws = _sheet([_day(
        _activity(day_part="first_half"), _activity(day_part="second_half")
    )])
    c = _cells(ws)
    assert c[(2, "HALF")] == "FIRST HALF"
    assert c[(3, "HALF")] == "SECOND HALF"


def test_full_day_two_activities_keep_their_own_benchmarks():
    """A counted activity and a lumpsum activity in the SAME Full Day report:
    each renders its own classification, and neither is relabelled by the other
    or by the half-day vocabulary."""
    ws = _sheet([_day(
        _activity(day_part="full_day", benchmark_type="NUMERIC_DAILY",
                  benchmark_value=250, benchmark_unit="tags"),
        _activity(day_part="full_day", benchmark_type="TASK_STATUS_ONLY"),
    )])
    c = _cells(ws)
    assert (c[(2, "HALF")], c[(2, "BENCHMARK")]) == ("ACTIVITY 1", "250 TAGS")
    assert (c[(3, "HALF")], c[(3, "BENCHMARK")]) == ("ACTIVITY 2", "LS")


def test_day_identity_is_merged_but_every_cell_keeps_its_value():
    """The merge is visual only. Excel's AutoFilter reads the underlying cell,
    so blanking the second-half row would drop it from an 'Employee = X' filter
    — the exact failure this guards."""
    ws = _sheet([_day(
        _activity(day_part="first_half"), _activity(day_part="second_half")
    )])
    merged = {str(r) for r in ws.merged_cells.ranges}
    assert {"A2:A3", "B2:B3", "C2:C3", "D2:D3", "P2:P3"} <= merged
    # The activity columns are NEVER merged — they must stay individually
    # filterable and sortable.
    assert not [m for m in merged if m[0] in "EFGHIJKLMNO"]
    # openpyxl masks a merged follower on read; the written cell still holds it.
    written = {c.coordinate: c.value for c in ws["A"]}
    assert written["A2"] == "CDC019 - ARTHI S"


def test_day_status_is_not_merged_when_the_halves_differ():
    ws = _sheet([_day(
        _activity(day_part="first_half", period_status="Work at Office"),
        _activity(day_part="second_half", period_status="Leave"),
    )])
    c = _cells(ws)
    assert c[(2, "DAY STATUS")] == "WORK AT OFFICE"
    assert c[(3, "DAY STATUS")] == "LEAVE"
    assert "D2:D3" not in {str(r) for r in ws.merged_cells.ranges}


def test_leave_day_shows_leave_and_no_activity_detail():
    """CASE 4 — a day with no task lines at all."""
    ws = _sheet([_day(day_status="Leave", remarks=None)])
    c = _cells(ws)
    assert ws.max_row == 2
    assert c[(2, "DAY STATUS")] == "LEAVE"
    assert c[(2, "ACTIVITY TYPE")] == "LEAVE"
    assert c[(2, "HALF")] == "FULL DAY"
    # No project, sub-activity, counts or benchmark are invented for a day that
    # was not worked.
    for header in ("PROJECT CODE", "SUB ACTIVITY TYPE", "BENCHMARK",
                   "NO. OF TAGS", "NO. OF SPARES", "DAY REMARKS"):
        assert c[(2, header)] is None


def test_day_and_date_columns():
    ws = _sheet([_day(_activity())])
    c = _cells(ws)
    assert c[(2, "DAY")] == "MONDAY"  # 2026-08-17 is a Monday
    # A real Excel date, not text (openpyxl reads one back as a datetime).
    assert c[(2, "DATE")].date() == MONDAY
    assert ws.cell(2, 2).number_format == "yyyy-mm-dd"


def test_every_text_cell_is_uppercase():
    """CASE 7. Numbers and dates are left alone — uppercasing those would be
    meaningless at best and destructive at worst."""
    ws = _sheet([_day(
        _activity(day_part="first_half", project_code="p-1", activity_type="mtl",
                  sub_activity_type="mtl-data population", tags=51,
                  benchmark_type="NUMERIC_DAILY", benchmark_value=50, benchmark_unit="tags"),
        _activity(day_part="second_half", project_code="p-2", activity_type="doc idb"),
    )])
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert cell.value == cell.value.upper(), cell.coordinate
    c = _cells(ws)
    assert c[(2, "EMPLOYEE ID & NAME")] == "CDC019 - ARTHI S"
    assert c[(2, "SUB ACTIVITY TYPE")] == "MTL-DATA POPULATION"
    assert c[(2, "NO. OF TAGS")] == 51          # the count is untouched


def test_multiple_employees_share_one_continuous_table():
    """No per-employee banner or repeated header — those made the old sheet
    unfilterable as a single range."""
    rows = [
        _day(_activity(), employee_label="CDC019 - Arthi S"),
        _day(_activity(), employee_label="CDC021 - Yuva Shree"),
    ]
    ws = _sheet(rows)
    assert ws.max_row == 3
    assert ws.auto_filter.ref == "A1:P3"
    assert [ws.cell(r, 1).value for r in (2, 3)] == [
        "CDC019 - ARTHI S", "CDC021 - YUVA SHREE",
    ]


def test_empty_report_still_produces_a_usable_header():
    ws = _sheet([])
    assert ws.max_row == 1
    assert ws.auto_filter.ref == "A1:P1"
